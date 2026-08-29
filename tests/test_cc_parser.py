"""Tests for credit_card_parser hardening.

Covers two robustness fixes:
- zero-amount referenced lines must NOT be classified as spend (they would
  demand an impossible R0.00 receipt and inflate the outstanding count);
- a Bank Statement sheet with no 'Reconciled' column must raise loudly rather
  than silently marking every historical (settled) line as outstanding.
"""
import openpyxl
import pytest

from northwind.cards import parser as cc


def test_classify_zero_amount_is_transfer_not_spend():
    assert cc.classify("SOME MERCHANT", 0) == "transfer"
    assert cc.classify("SOME MERCHANT", -6000) == "spend"
    assert cc.classify("PAYMENT RECEIVED", 12000) == "transfer"
    assert cc.classify("#DECLINED AUTH FEE", -100) == "fee"
    # zero-amount line does not need a receipt
    assert cc.needs_receipt(cc.classify("X", 0)) is False


def _ws(headers, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for r in rows:
        ws.append(r)
    return ws


def test_parse_raises_when_no_reconciled_column():
    ws = _ws(["Date", "Reference", "Amount"],
             [["2025-01-05", "SOME SHOP", -60.0]])
    with pytest.raises(ValueError, match="Reconciled"):
        list(cc._parse_bank_statement(ws))


def test_summary_unreconciled_fallback_section():
    # Some Xero exports leave the Bank Statement sheet with only balance rows and
    # itemise the chargeable lines in the summary's 'Plus Unreconciled Statement
    # Lines' section. The parser must read those, stop at the Total row, and
    # classify a positive BALANCE TRANSFERRED as a (hidden) transfer.
    import datetime as dt
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Plus Unreconciled Statement Lines'])
    ws.append([dt.datetime(2026, 6, 24), None, 'SUMMIT CITY PARKING    SUMMIT', -10.0])
    ws.append([dt.datetime(2026, 6, 29), None, 'BALANCE TRANSFERRED', 5172.36])
    ws.append(['Total Unreconciled Statement Lines', None, None, 0])
    ws.append([dt.datetime(2026, 7, 2), None, 'AFTER TOTAL SHOULD NOT APPEAR', -99.0])

    out = [(ref, cents) for _, ref, cents, recon in cc._parse_summary_unreconciled(ws)]
    assert ('SUMMIT CITY PARKING SUMMIT', -1000) in out
    assert ('BALANCE TRANSFERRED', 517236) in out
    assert all('AFTER TOTAL' not in ref for ref, _ in out)      # stopped at the Total row
    assert cc.classify('BALANCE TRANSFERRED', 517236) == 'transfer'  # positive -> hidden


def test_parse_reads_reconciled_flag_when_present():
    ws = _ws(["Date", "Reference", "Amount", "Reconciled"],
             [["2025-01-05", "SETTLED SHOP", -60.0, "Yes"],
              ["2025-01-06", "OPEN SHOP", -40.0, "No"]])
    out = list(cc._parse_bank_statement(ws))
    assert len(out) == 2
    # (date, ref, amount_cents, reconciled, import_batch)
    recon_by_ref = {ref: reconciled for _, ref, _, reconciled, _batch in out}
    assert recon_by_ref["SETTLED SHOP"] is True
    assert recon_by_ref["OPEN SHOP"] is False


def test_display_name_trims_dangling_separator():
    assert cc._display_name("Operations Credit Card") == "Operations"
    assert cc._display_name("Primary Credit Card") == "Primary"


def test_union_collapses_reimport_but_keeps_genuine_duplicates():
    import datetime as dt
    d = dt.date(2026, 6, 1)

    # Re-import artifact: the SAME charge listed under two different import
    # batches (and re-listed on the Summary). It is ONE transaction.
    bank = [(d, "SHOP A", -5000, False, "2026-06-01"),
            (d, "SHOP A", -5000, False, "2026-06-08")]   # re-import, later batch
    summ = [(d, "SHOP A", -5000, False)]                 # Summary re-lists it too
    a = [r for r in cc._union_rows(bank, summ) if r[1] == "SHOP A"]
    assert len(a) == 1                                   # collapsed, not 2 or 3

    # Genuine same-day double charge: two lines in the SAME import batch -> kept.
    bank2 = [(d, "COFFEE", -3000, False, "2026-06-01"),
             (d, "COFFEE", -3000, False, "2026-06-01")]
    c = [r for r in cc._union_rows(bank2, []) if r[1] == "COFFEE"]
    assert len(c) == 2

    # A line only the Summary lists (carried over from an earlier month) still
    # comes through, once, as unreconciled.
    assert cc._union_rows([], [(d, "SHOP B", -7000, False)]) == \
        [(d, "SHOP B", -7000, False)]


def test_as_at_report_unions_bank_and_summary_across_months():
    """An 'as at' export lists only the current period on the Bank Statement
    sheet but every still-outstanding line (incl. earlier months) in the Summary.
    parse_workbook must union both so nothing carried over is dropped."""
    import io
    import datetime as dt
    wb = openpyxl.Workbook()
    # Summary sheet (self-identifies the card, carries meta + the full list).
    s = wb.active
    s.title = "Zed - Credit Card Reconci..."
    s["A1"] = "Zed - Credit Card Reconciliation Summary"
    s["A3"] = "As at 9 July 2026"
    s["A4"] = "Zed - Credit Card"
    s.append([])  # row 5
    s.append(["For the period 1 July 2026 to 9 July 2026"])  # row 6 (meta)
    s.append([])
    s.append(["Plus Unreconciled Statement Lines"])
    s.append([dt.datetime(2026, 6, 15), None, "OLD JUNE SHOP        NTH", -120.0])
    s.append([dt.datetime(2026, 7, 2), None, "NEW JULY SHOP        NTH", -50.0])
    s.append(["Total Unreconciled Statement Lines", None, None, 0])

    # Bank Statement sheet: only the July period line (+ its Reconciled flag).
    b = wb.create_sheet("Bank Statement")
    b.append(["Date", "Description", "Date imported", "Reference", "Reconciled",
              "Source", "Amount", "Balance"])
    b.append([dt.datetime(2026, 7, 2), None, None, "NEW JULY SHOP        NTH",
              "No", "Imported", -50.0, 0])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    snap = cc.parse_workbook(buf, source_filename="zed.xlsx")

    refs = sorted(l.reference for l in snap.lines)
    assert refs == ["NEW JULY SHOP NTH", "OLD JUNE SHOP NTH"]   # both, deduped
    months = {l.reference: (l.line_date.year, l.line_date.month) for l in snap.lines}
    assert months["OLD JUNE SHOP NTH"] == (2026, 6)
    assert months["NEW JULY SHOP NTH"] == (2026, 7)
    assert snap.display_name == "Zed"
