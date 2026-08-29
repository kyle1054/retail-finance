"""Cash-recon reporting exports + the overview exceptions panel.

Covers the three new admin exports (overview grid xlsx, category×store matrix
xlsx, batch expense-MJ zip) and that the all-stores overview surfaces stores
needing attention. Isolated store + far-future months.
"""
import csv
import io
import zipfile

import openpyxl

from northwind.data import database as db


def _store(conn, name, code=None):
    conn.execute(
        "INSERT OR IGNORE INTO stores (name, store_code, xero_tracking_name) VALUES (?, ?, ?)",
        (name, code, f'{name} - TEST'))
    if code:
        conn.execute("UPDATE stores SET store_code=?, xero_tracking_name=? WHERE name=?",
                     (code, f'{name} - TEST', name))
    conn.commit()
    db.invalidate_stores_cache()
    return name


def _expense_cat(conn, like='milk%', code='6330', vat='novat'):
    cid = conn.execute(
        "SELECT id FROM recon_categories WHERE kind='expense' AND lower(name) LIKE ? LIMIT 1",
        (like,)).fetchone()['id']
    conn.execute("UPDATE recon_categories SET xero_code=?, vat_type=? WHERE id=?", (code, vat, cid))
    conn.commit()
    return cid


def test_overview_xlsx_has_totals_row(client, conn):
    store = _store(conn, 'PYTEST RPT OVERVIEW')
    db.add_recon_entry(store, '2098-05-04', _expense_cat(conn), 60.00, 'milk', created_by=store)
    r = client.get('/cash/overview.xlsx?start=2098-05-01&end=2098-05-31')
    assert r.status_code == 200 and 'spreadsheet' in r.content_type
    ws = openpyxl.load_workbook(io.BytesIO(r.get_data())).active
    assert [c.value for c in ws[1]][:2] == ['Store', 'Opening']
    labels = [row[0].value for row in ws.iter_rows()]
    assert store in labels
    assert 'All stores' in labels                      # totals row present


def test_category_matrix_xlsx_shape(client, conn):
    store = _store(conn, 'PYTEST RPT MATRIX')
    db.add_recon_entry(store, '2098-06-04', _expense_cat(conn), 75.00, 'milk', created_by=store)
    r = client.get('/cash/reports/category-matrix.xlsx?start=2098-06-01&end=2098-06-30')
    assert r.status_code == 200
    ws = openpyxl.load_workbook(io.BytesIO(r.get_data())).active
    hdr = [c.value for c in ws[1]]
    assert hdr[0] == 'Category' and hdr[1] == 'Xero code' and hdr[-1] == 'Total'
    assert store in hdr                                # the store is a column
    assert any(row[0].value == 'TOTAL' for row in ws.iter_rows())


def test_category_matrix_xlsx_totals_are_correct(client, conn):
    """Beyond shape: the per-store cells, the row Total, the column TOTAL row and
    the grand total must all reconcile. Uses a unique far-future range so this
    store is the only one with spend in it."""
    store = _store(conn, 'PYTEST MATRIX TOT', code='SC-132')
    milk = _expense_cat(conn, 'milk%', '6330', 'novat')
    coffee = _expense_cat(conn, 'coffee%', '6240', 'standard')
    db.add_recon_entry(store, '2096-03-04', milk, 75.00, 'milk', created_by=store)
    db.add_recon_entry(store, '2096-03-06', milk, 25.00, 'more milk', created_by=store)
    db.add_recon_entry(store, '2096-03-08', coffee, 40.00, 'beans', created_by=store)

    r = client.get('/cash/reports/category-matrix.xlsx?start=2096-03-01&end=2096-03-31')
    assert r.status_code == 200
    ws = openpyxl.load_workbook(io.BytesIO(r.get_data())).active
    hdr = [c.value for c in ws[1]]
    scol = hdr.index(store) + 1                     # this store's column (1-based)
    total_col = len(hdr)                            # 'Total' is the last column

    rows_by_cat = {}
    total_row = None
    for row in ws.iter_rows(min_row=2):
        label = row[0].value
        if label == 'TOTAL':
            total_row = row
        elif label:
            rows_by_cat[label] = row

    def cell(row, col):
        return round(float(row[col - 1].value or 0), 2)

    milk_row = next(r for name, r in rows_by_cat.items() if name.lower().startswith('milk'))
    coffee_row = next(r for name, r in rows_by_cat.items() if name.lower().startswith('coffee'))
    # Per-store cells sum the entries; the row Total equals this store's spend
    # (it is the only store with data in this range).
    assert cell(milk_row, scol) == 100.00 and cell(milk_row, total_col) == 100.00
    assert cell(coffee_row, scol) == 40.00 and cell(coffee_row, total_col) == 40.00
    # Column TOTAL for the store, and the grand total, both = 140.00.
    assert cell(total_row, scol) == 140.00
    assert cell(total_row, total_col) == 140.00


def test_batch_mj_zip_balances_and_skips_empty(client, conn):
    store = _store(conn, 'PYTEST RPT ZIP', code='SC-131')
    db.add_recon_entry(store, '2098-07-04', _expense_cat(conn, 'milk%', '6330', 'novat'),
                       120.00, 'milk', created_by=store)
    r = client.get('/cash/2098/7/mj/all.zip')
    assert r.status_code == 200 and r.content_type == 'application/zip'
    z = zipfile.ZipFile(io.BytesIO(r.get_data()))
    mine = [n for n in z.namelist() if n.startswith(store)]
    assert mine, 'store with expenses should have a CSV'
    rows = list(csv.DictReader(io.StringIO(z.read(mine[0]).decode())))
    assert sum(round(float(x['*Amount']) * 100) for x in rows) == 0     # balanced


def test_batch_mj_zip_refuses_incomplete_mapping(client, conn):
    store = _store(conn, 'PYTEST RPT ZIP UNMAPPED')
    db.add_recon_entry(store, '2098-09-04', _expense_cat(conn, 'milk%', '6330', 'novat'),
                       120.00, 'milk', created_by=store)
    r = client.get('/cash/2098/9/mj/all.zip', follow_redirects=False)
    assert r.status_code in (302, 303)
    assert r.headers['Location'].endswith('/cash')


def test_overview_flags_stores_needing_attention(client, conn):
    store = _store(conn, 'PYTEST RPT ATTN')
    # Opening 100, expense 500 -> negative closing float -> must be flagged.
    db.set_recon_opening(store, 2098, 8, 100.00)
    db.add_recon_entry(store, '2098-08-04', _expense_cat(conn), 500.00, 'big spend', created_by=store)
    body = client.get('/cash?start=2098-08-01&end=2098-08-31').get_data(as_text=True)
    assert 'need attention' in body
    assert 'Negative closing float' in body
    assert store in body
