"""Consolidated cash-sales journal + Shopify-compare export.

Covers the pure builder (balancing + contra), the per-store aggregation
(only 'Cash Sale' income counts), and the three routes (CSV journal, xlsx
compare, settings round-trip). Isolated store + far-future months (2098-xx).
"""
import csv
import hashlib
import io

import openpyxl
import pytest

from northwind.cash import mj as cash_mj
from northwind.data import database as db
from northwind.cash.routes import (CASH_SHOPIFY_EXCLUDED_KEY, _cash_sales_context,
                             _excluded_shopify_locations)
from northwind.cash.shopify import parse_shopify_cash_csv


def _sales_store(conn, name='PYTEST SALES', code='SC-139'):
    conn.execute("INSERT OR IGNORE INTO stores (name, store_code) VALUES (?, ?)", (name, code))
    conn.execute("UPDATE stores SET store_code=? WHERE name=?", (code, name))
    conn.commit()
    db.invalidate_stores_cache()
    return name


def _income_cat(conn, like):
    return conn.execute(
        "SELECT id FROM recon_categories WHERE kind='income' AND lower(name) LIKE ? LIMIT 1",
        (like,)).fetchone()['id']


# ── pure builder ──────────────────────────────────────────────────────────────

def test_build_cash_sales_rows_matches_finance_split_template():
    rows = cash_mj.build_cash_sales_rows(
        'Retail cash sales journal - July 2026', '31/07/2026', '8990.9',
        [{'store': 'Riverbend', 'store_code': 'SC-117', 'sales_cents': 10000},
         {'store': 'Unknown', 'store_code': None, 'sales_cents': 5000}])
    # Seven columns, exactly the workbook's "To copy" sheet.
    assert list(rows[0].keys()) == cash_mj.CASH_SALES_HEADER
    assert len(cash_mj.CASH_SALES_HEADER) == 7
    assert len(rows) == 3                                   # 2 store debits + 1 contra
    assert rows[0]['*AccountCode'] == 'SC-117' and rows[0]['*Amount'] == '100.00'
    assert rows[1]['*AccountCode'] == cash_mj.UNMAPPED_CODE  # None -> ####
    contra = rows[-1]
    assert contra['*AccountCode'] == '8990.9'
    assert contra['*Amount'] == '-150.00'                   # balancing credit
    # This journal carries no tracking — the store is in the account code.
    assert all(r['TrackingName1'] == '' for r in rows)
    assert all(r['*TaxRate'] == 'No VAT (0%)' for r in rows)


def test_cash_sales_journal_balances_to_zero():
    """Xero rejects an unbalanced manual journal, so Σ*Amount must be exactly 0."""
    rows = cash_mj.build_cash_sales_rows(
        'Retail cash sales journal - July 2026', '31/07/2026', '8990.9',
        [{'store': 'A', 'store_code': 'SC-117', 'sales_cents': 4668750},
         {'store': 'B', 'store_code': 'SC-121', 'sales_cents': 4783750},
         {'store': 'C', 'store_code': 'SC-110', 'sales_cents': 0},
         {'store': 'D', 'store_code': 'SC-100', 'sales_cents': 2010260}])
    cents = [round(float(r['*Amount']) * 100) for r in rows]
    assert sum(cents) == 0
    assert cents[-1] == -(4668750 + 4783750 + 2010260)


# ── aggregation ────────────────────────────────────────────────────────────────

def test_cash_sales_by_store_counts_only_cash_sales(conn):
    store = _sales_store(conn)
    sale = _income_cat(conn, 'cash sale%')
    topup = _income_cat(conn, '%top up%')
    exp = conn.execute("SELECT id FROM recon_categories WHERE kind='expense' LIMIT 1").fetchone()['id']

    db.add_recon_entry(store, '2098-04-03', sale, 1200.00, 'till', created_by=store)
    db.add_recon_entry(store, '2098-04-09', sale, 800.00, 'till', created_by=store)
    db.add_recon_entry(store, '2098-04-10', topup, 500.00, 'float', created_by=store)   # excluded
    db.add_recon_entry(store, '2098-04-11', exp, 60.00, 'milk', created_by=store)        # excluded

    got = [r for r in db.get_cash_sales_journal_stores(2098, 4) if r['store'] == store]
    assert len(got) == 1
    assert got[0]['sales_cents'] == 200000        # 1200 + 800 only
    assert got[0]['store_code'] == 'SC-139'


# ── routes ───────────────────────────────────────────────────────────────────

def test_cash_sales_line_items_carry_sale_number(client, conn):
    store = _sales_store(conn, 'PYTEST SALES DRILL', 'SC-133')
    sale = _income_cat(conn, 'cash sale%')
    db.add_recon_entry(store, '2098-08-03', sale, 1500.00, 'Z-report #55501', created_by=store)
    db.add_recon_entry(store, '2098-08-07', sale, 900.00, 'Z-report #55510', created_by=store)

    items = [it for it in db.get_cash_sales_line_items(2098, 8) if it['store'] == store]
    assert len(items) == 2
    assert {it['sale_no'] for it in items} == {'Z-report #55501', 'Z-report #55510'}
    # The preview page drills down to the individual sales + their numbers.
    body = client.get('/cash/cash-sales/2098/8').get_data(as_text=True)
    assert 'Z-report #55501' in body and 'Z-report #55510' in body
    assert 'Sale / receipt no.' in body


def test_cash_sales_mj_csv_matches_reference_and_is_month_scoped(client, conn):
    store = _sales_store(conn, 'PYTEST SALES CSV', 'SC-135')
    db.set_store_xero(store, None, 'SC-135', 'Finance Store Label')
    sale = _income_cat(conn, 'cash sale%')
    db.add_recon_entry(store, '2098-05-04', sale, 3333.33, 'till', created_by=store)
    db.add_recon_entry(store, '2098-04-30', sale, 9999.99, 'other month', created_by=store)

    r = client.get('/cash/cash-sales/2098/5/mj.csv')
    assert r.status_code == 200 and 'text/csv' in r.content_type
    rows = list(csv.DictReader(io.StringIO(r.get_data(as_text=True))))
    assert list(rows[0].keys()) == cash_mj.CASH_SALES_HEADER
    mine = next(x for x in rows if x['*AccountCode'] == 'SC-135')
    assert mine['*Amount'] == '3333.33'
    assert mine['TrackingName1'] == ''          # cash_sales_label is not tracking
    assert rows[-1]['*AccountCode'] == '8990.9'
    # The control row balances the whole journal, including every other store.
    assert sum(round(float(x['*Amount']) * 100) for x in rows) == 0


def _shopify_csv(location, gross='1500.00', refund='0', net='1500.00'):
    return (b'POS location name,Payment gateway,Order name,Transactions,Gross payments,'
            b'Refunded payments,Net payments\n' +
            f'{location},cash,NORTHWIND9001,2,{gross},{refund},{net}\n'.encode())


def test_shopify_parser_uses_cash_rows_only():
    data = (_shopify_csv('NORTHWIND Riverbend') +
            b'NORTHWIND Riverbend,card,NORTHWIND9002,1,999.00,0,999.00\n')
    rows, digest = parse_shopify_cash_csv(data)
    assert len(rows) == 1 and rows[0]['net_cents'] == 150000
    assert rows[0]['transactions'] == 2 and len(digest) == 64


def test_shopify_parser_rejects_bad_headers_and_bad_net():
    with pytest.raises(ValueError, match='missing'):
        parse_shopify_cash_csv(b'location,amount\nNORTHWIND Riverbend,10\n')
    with pytest.raises(ValueError, match='does not equal'):
        parse_shopify_cash_csv(_shopify_csv('NORTHWIND Riverbend', gross='100', refund='-10', net='95'))


def test_shopify_upload_is_month_bound_replace_confirmed_and_reasons_persist(client):
    first = _shopify_csv('NORTHWIND Riverbend', gross='100', net='100')
    second = _shopify_csv('NORTHWIND Riverbend', gross='200', net='200')
    r = client.post('/cash/cash-sales/2098/10/shopify', data={
        'shopify_csv': (io.BytesIO(first), 'october.csv'),
    }, content_type='multipart/form-data')
    assert r.status_code in (302, 303)
    assert db.get_cash_shopify_upload(2098, 10)['source_filename'] == 'october.csv'

    # A repeat without explicit confirmation leaves the first upload untouched.
    client.post('/cash/cash-sales/2098/10/shopify', data={
        'shopify_csv': (io.BytesIO(second), 'replacement.csv'),
    }, content_type='multipart/form-data')
    assert db.get_cash_shopify_upload(2098, 10)['source_filename'] == 'october.csv'
    client.post('/cash/cash-sales/2098/10/shopify', data={
        'confirm_replace': '1',
        'shopify_csv': (io.BytesIO(second), 'replacement.csv'),
    }, content_type='multipart/form-data')
    assert db.get_cash_shopify_upload(2098, 10)['source_filename'] == 'replacement.csv'
    assert db.get_cash_shopify_summary(2098, 10)[0]['net_cents'] == 20000
    assert db.get_cash_shopify_upload(2098, 11) is None
    assert client.get('/cash/cash-sales/2098/10/recon.xlsx').status_code in (302, 303)

    client.post('/cash/cash-sales/2098/10/reasons', data={
        'row_count': '1', 'store_0': 'Riverbend', 'reason_0': 'Payment method corrected',
    })
    assert db.get_cash_sales_variance_reasons(2098, 10)['Riverbend'] == 'Payment method corrected'
    assert client.get('/cash/cash-sales/2098/10/recon.xlsx').status_code == 200


def test_cash_sales_recon_xlsx_shape(client, conn):
    store = _sales_store(conn, 'PYTEST SALES XLSX', 'SC-134')
    sale = _income_cat(conn, 'cash sale%')
    db.add_recon_entry(store, '2098-06-04', sale, 999.00, 'INV-9001', created_by=store)
    db.add_recon_entry(store, '2098-06-08', sale, 501.00, 'INV-9002', created_by=store)

    # Map the synthetic Shopify location, then upload the month-specific CSV.
    db.save_cash_shopify_mapping('NORTHWIND PYTEST SALES XLSX', store)
    up = client.post('/cash/cash-sales/2098/6/shopify', data={
        'shopify_csv': (io.BytesIO(_shopify_csv('NORTHWIND PYTEST SALES XLSX')), 'shopify-june.csv'),
    }, content_type='multipart/form-data')
    assert up.status_code in (302, 303)

    r = client.get('/cash/cash-sales/2098/6/recon.xlsx')
    assert r.status_code == 200 and 'spreadsheet' in r.content_type
    wb = openpyxl.load_workbook(io.BytesIO(r.get_data()))
    ws = wb.worksheets[0]
    assert [c.value for c in ws[1]] == [
        'POS location name', 'POS location name', 'Payment gateway', 'Transactions',
        'Gross payments', 'Refunded payments', 'Net payments', None,
        'Cash Sheet', None, 'Diff', None, 'Reason']
    mine = [row for row in ws.iter_rows(values_only=True) if row[1] == store]
    assert mine and mine[0][6] == 1500.00 and mine[0][8] == 1500.00
    assert str(mine[0][10]).startswith('=I')

    # Second sheet lists EVERY individual sale with its receipt number, grouped
    # per store: a store header row carrying the store subtotal, its sales below.
    ws2 = wb.worksheets[1]
    assert [c.value for c in ws2[1]] == ['Store', 'Date', 'Sale / receipt no.', 'Amount']
    rows = list(ws2.iter_rows(values_only=True))
    # Store header row: store name in col A + subtotal (1500) in col D, no date.
    hdr = next(row for row in rows if row[0] == store)
    assert hdr[1] is None and hdr[3] == 1500.00        # subtotal on the header row
    # Its sales appear as detail rows (no store in col A, receipt no. in col C).
    nums = {row[2] for row in rows if row[2] in ('INV-9001', 'INV-9002')}
    assert nums == {'INV-9001', 'INV-9002'}
    assert any(row[0] == 'GRAND TOTAL' for row in rows)


def test_cash_sales_settings_round_trip(client):
    client.post('/cash/xero-setup/cash-sales',
                data={'contra_code': '7777.7', 'narration': 'Custom journal'},
                follow_redirects=True)
    assert db.get_setting('cash_sales_contra_code') == '7777.7'
    assert db.get_setting('cash_sales_narration') == 'Custom journal'
    # Preview reflects the saved narration + contra.
    store = None
    r = client.get('/cash/cash-sales/2098/7')
    assert r.status_code == 200
    body = r.get_data(as_text=True)
    assert '7777.7' in body and 'Custom journal' in body
    # Restore defaults so we don't bleed into other tests.
    client.post('/cash/xero-setup/cash-sales',
                data={'contra_code': '8990.9', 'narration': 'Retail cash sales journal'})


def test_cash_sales_empty_month_exports_nothing(client):
    r = client.get('/cash/cash-sales/2097/1/mj.csv', follow_redirects=False)
    assert r.status_code in (302, 303)      # warned + redirected, no empty journal


def test_cash_sales_xero_export_refuses_unmapped_store(client, conn):
    store = _sales_store(conn, 'PYTEST SALES UNMAPPED', '')
    sale = _income_cat(conn, 'cash sale%')
    db.add_recon_entry(store, '2098-09-04', sale, 250.00, 'INV-UNMAPPED', created_by=store)

    preview = client.get('/cash/cash-sales/2098/9').get_data(as_text=True)
    assert 'Xero journal needs mapping' in preview
    assert 'The Xero journal stays unavailable' in preview

    r = client.get('/cash/cash-sales/2098/9/mj.csv', follow_redirects=False)
    assert r.status_code in (302, 303)
    assert r.headers['Location'].endswith('/cash/cash-sales/2098/9')


def test_cash_sales_is_listed_with_other_cash_exports(client):
    body = client.get('/cash?start=2098-06-01&end=2098-06-30').get_data(as_text=True)
    assert 'Cash sales &amp; Xero' in body
    assert 'Cash-sales Xero MJ' in body


# ── Shopify parser rejection paths ────────────────────────────────────────────
# The parser is fail-loud by contract: every unreadable cell must surface as a
# row-numbered ValueError the route can flash verbatim. An OverflowError or an
# InvalidOperation escaping here is a 500, and a silently-zeroed amount would
# corrupt the reconciliation, so both are asserted against explicitly.

# Cells that parse as a Decimal but are not a usable finite money/count value.
NON_FINITE_CELLS = ('Infinity', '-Infinity', 'NaN', '-NaN', 'sNaN', '1e400')


def test_shopify_parser_rejects_non_finite_transactions():
    for cell in NON_FINITE_CELLS:
        payload = (b'POS location name,Payment gateway,Order name,Transactions,'
                   b'Gross payments,Refunded payments,Net payments\n' +
                   f'NORTHWIND Riverbend,cash,NORTHWIND9001,{cell},100.00,0,100.00\n'.encode())
        with pytest.raises(ValueError, match=r'Row 2: Transactions') as exc:
            parse_shopify_cash_csv(payload)
        # Not an OverflowError, and not a stray ValueError from int()/Decimal.
        assert type(exc.value) is ValueError, cell


def test_shopify_parser_rejects_non_finite_money_columns():
    columns = {'Gross payments': 4, 'Refunded payments': 5, 'Net payments': 6}
    for label, index in columns.items():
        for cell in NON_FINITE_CELLS:
            fields = ['NORTHWIND Riverbend', 'cash', 'NORTHWIND9001', '2', '100.00', '0', '100.00']
            fields[index] = cell
            payload = (b'POS location name,Payment gateway,Order name,Transactions,'
                       b'Gross payments,Refunded payments,Net payments\n' +
                       (','.join(fields) + '\n').encode())
            with pytest.raises(ValueError, match=r'Row 2: %s is not a valid amount' % label):
                parse_shopify_cash_csv(payload)


def test_shopify_parser_rejects_absurd_but_finite_magnitudes():
    # Finite, whole, non-negative — but far past any real month of retail cash,
    # and past what SQLite can store as an integer once converted to cents.
    payload = (b'POS location name,Payment gateway,Order name,Transactions,Gross payments,'
               b'Refunded payments,Net payments\n'
               b'NORTHWIND Riverbend,cash,NORTHWIND9001,1e18,100.00,0,100.00\n')
    with pytest.raises(ValueError, match=r'Row 2: Transactions must be a non-negative'):
        parse_shopify_cash_csv(payload)
    with pytest.raises(ValueError, match=r'Row 2: Gross payments is not a valid amount'):
        parse_shopify_cash_csv(_shopify_csv(
            'NORTHWIND Riverbend', gross='99999999999999999999', net='99999999999999999999'))


def test_shopify_parser_rejects_oversized_payload():
    from northwind.cash.shopify import MAX_UPLOAD_BYTES
    with pytest.raises(ValueError, match='larger than 10 MB'):
        parse_shopify_cash_csv(b'x' * (MAX_UPLOAD_BYTES + 1))


def test_shopify_parser_rejects_non_utf8_bytes():
    payload = (b'POS location name,Payment gateway,Order name,Transactions,Gross payments,'
               b'Refunded payments,Net payments\n' +
               'NORTHWIND Café,cash,NORTHWIND9001,1,100.00,0,100.00\n'.encode('latin-1'))
    with pytest.raises(ValueError, match='must be UTF-8 encoded'):
        parse_shopify_cash_csv(payload)


def test_shopify_parser_rejects_empty_payload():
    with pytest.raises(ValueError, match='Choose a Shopify CSV'):
        parse_shopify_cash_csv(b'')


def test_shopify_parser_rejects_file_without_cash_rows():
    payload = (b'POS location name,Payment gateway,Order name,Transactions,Gross payments,'
               b'Refunded payments,Net payments\n'
               b'NORTHWIND Riverbend,card,NORTHWIND9001,1,100.00,0,100.00\n'
               b'NORTHWIND Riverbend,shopify_payments,NORTHWIND9002,1,50.00,0,50.00\n')
    with pytest.raises(ValueError, match='contains no cash-payment rows'):
        parse_shopify_cash_csv(payload)


def test_shopify_parser_rejects_cash_row_without_location():
    payload = (b'POS location name,Payment gateway,Order name,Transactions,Gross payments,'
               b'Refunded payments,Net payments\n'
               b',cash,NORTHWIND9001,1,100.00,0,100.00\n')
    with pytest.raises(ValueError, match='Row 2: POS location name is required'):
        parse_shopify_cash_csv(payload)


def test_shopify_parser_reads_a_clean_multi_row_file():
    """Positive control: the rejection paths above must not swallow real data."""
    payload = (
        b'POS location name,Payment gateway,Order name,Transactions,Gross payments,'
        b'Refunded payments,Net payments\n'
        b'NORTHWIND Riverbend,cash,NORTHWIND9101,3,1250.50,-50.50,1200.00\n'
        b'NORTHWIND Riverbend,card,NORTHWIND9102,7,4000.00,0,4000.00\n'          # dropped: not cash
        b'NORTHWIND Northgate,Cash,NORTHWIND9103,1,99.99,0,99.99\n'              # gateway is casefolded
        b'NORTHWIND Parkview,cash,NORTHWIND9104,0,0,0,0\n'                       # a real zero stays
    )
    rows, digest = parse_shopify_cash_csv(payload)
    assert rows == [
        {'source_row': 2, 'pos_location_name': 'NORTHWIND Riverbend', 'payment_gateway': 'cash',
         'order_name': 'NORTHWIND9101', 'transactions': 3, 'gross_cents': 125050,
         'refunded_cents': -5050, 'net_cents': 120000},
        {'source_row': 4, 'pos_location_name': 'NORTHWIND Northgate', 'payment_gateway': 'Cash',
         'order_name': 'NORTHWIND9103', 'transactions': 1, 'gross_cents': 9999,
         'refunded_cents': 0, 'net_cents': 9999},
        {'source_row': 5, 'pos_location_name': 'NORTHWIND Parkview', 'payment_gateway': 'cash',
         'order_name': 'NORTHWIND9104', 'transactions': 0, 'gross_cents': 0,
         'refunded_cents': 0, 'net_cents': 0},
    ]
    # The digest covers the raw uploaded bytes, not the decoded text or the kept rows.
    assert digest == hashlib.sha256(payload).hexdigest()


# ── demo stores, orphans, stale mappings, exclusions (months 2095-xx) ─────────
# Each of these guards a way the month-end close can be silently wrong rather
# than loudly broken: fake money in an imported ledger workbook, a store that fell out of
# `stores`, a mapping pointing at a name nobody uses any more, and a POS location
# that can never be a store. All four used to produce a plausible-looking export.

DEMO_STORE = db.DEMO_STORE_PREFIX + 'PYTEST'    # never reaches a Xero export
NEAR_DEMO_STORE = 'PYTEST DEMO ROOM'            # merely CONTAINS "DEMO" — a real store


def _sheet1_cash_total(ws):
    """Σ of sheet 1's 'Cash Sheet' column over the comparison rows.

    Not the TOTAL row — openpyxl reads its `=SUM(...)` back as a formula string.
    Every comparison row carries the literal 'cash' gateway in column C."""
    return sum(row[8] for row in ws.iter_rows(values_only=True)
               if row[2] == 'cash' and isinstance(row[8], (int, float)))


def _sheet2_grand_total(ws2):
    return next(row[3] for row in ws2.iter_rows(values_only=True) if row[0] == 'GRAND TOTAL')


@pytest.fixture
def demo_store(conn):
    """A 'ZZ DEMO - …' store carrying a POS code, removed again afterwards.

    It must not outlive the test: the DB copy is session-scoped and
    tests/test_demo_seed.py counts every `ZZ DEMO%` store in it."""
    name = _sales_store(conn, DEMO_STORE, 'SC-129')
    yield name
    conn.execute("DELETE FROM cash_recon_entries WHERE store = ?", (name,))
    conn.execute("DELETE FROM stores WHERE name = ?", (name,))
    conn.commit()
    db.invalidate_stores_cache()


@pytest.fixture
def orphan_store(conn):
    """A store name with cash sales but no `stores` row, cleaned up afterwards —
    a lingering orphan fails tests/test_db_integrity.py for the whole suite."""
    name = 'PYTEST STORE THAT VANISHED'
    assert name not in db.get_stores()
    yield name
    conn.execute("DELETE FROM cash_recon_entries WHERE store = ?", (name,))
    conn.execute("DELETE FROM cash_sales_variance_reasons WHERE store = ?", (name,))
    conn.commit()


def _orphan_sale(conn, store, entry_date, cents, note):
    """Insert a cash sale directly — add_recon_entry would reject an unknown store."""
    sale = _income_cat(conn, 'cash sale%')
    conn.execute(
        "INSERT INTO cash_recon_entries (store, entry_date, category_id, description, "
        "direction, amount_cents, note, created_by) "
        "SELECT ?, ?, ?, name, 'in', ?, ?, 'pytest' FROM recon_categories WHERE id = ?",
        (store, entry_date, sale, cents, note, sale))
    conn.commit()


def test_demo_store_is_excluded_from_both_cash_sales_helpers(client, conn, demo_store):
    """`scripts/seed_demo_cash_recon.py` money must never reach a Xero import.

    Both helpers have to filter, not one: the journal drives sheet 1 of the recon
    workbook and the line items drive sheet 2, so filtering either alone puts the
    demo's takings in one sheet and not the other."""
    demo = demo_store
    near = _sales_store(conn, NEAR_DEMO_STORE, 'SC-130')
    sale = _income_cat(conn, 'cash sale%')
    db.add_recon_entry(demo, '2095-01-04', sale, 5000.00, 'DEMO-1', created_by=demo)
    db.add_recon_entry(near, '2095-01-05', sale, 1000.00, 'REAL-1', created_by=near)

    journal = {r['store'] for r in db.get_cash_sales_journal_stores(2095, 1)}
    assert demo not in journal
    assert near in journal, 'only the "ZZ DEMO - " PREFIX is a demo, not the word DEMO'
    items = {i['store'] for i in db.get_cash_sales_line_items(2095, 1)}
    assert demo not in items
    assert near in items

    # A demo store carrying a POS code must not block the export either — the
    # point of the filter is that the store is simply not in the journal at all.
    r = client.get('/cash/cash-sales/2095/1/mj.csv')
    assert r.status_code == 200 and 'text/csv' in r.content_type
    rows = list(csv.DictReader(io.StringIO(r.get_data(as_text=True))))
    codes = {x['*AccountCode'] for x in rows}
    assert 'SC-129' not in codes and 'SC-130' in codes
    assert cash_mj.UNMAPPED_CODE not in codes
    assert sum(round(float(x['*Amount']) * 100) for x in rows) == 0


def test_demo_store_absent_from_recon_workbook_whose_sheets_still_tie(client, conn,
                                                                      demo_store):
    sale = _income_cat(conn, 'cash sale%')
    db.add_recon_entry(demo_store, '2095-01-04', sale, 5000.00, 'DEMO-2',
                       created_by=demo_store)
    _sales_store(conn, NEAR_DEMO_STORE, 'SC-130')
    db.save_cash_shopify_mapping('NORTHWIND PYTEST DEMO ROOM', NEAR_DEMO_STORE)
    up = client.post('/cash/cash-sales/2095/1/shopify', data={
        'shopify_csv': (io.BytesIO(_shopify_csv('NORTHWIND PYTEST DEMO ROOM',
                                                gross='1000.00', net='1000.00')),
                        'demo-month.csv')}, content_type='multipart/form-data')
    assert up.status_code in (302, 303)

    r = client.get('/cash/cash-sales/2095/1/recon.xlsx')
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.get_data()))
    ws, ws2 = wb.worksheets[0], wb.worksheets[1]
    assert DEMO_STORE not in {row[1] for row in ws.iter_rows(values_only=True)}
    assert DEMO_STORE not in {row[0] for row in ws2.iter_rows(values_only=True)}
    # Sheet 1 (journal-driven) and sheet 2 (line-item-driven) must agree; they
    # differ by the demo's R5,000 the moment one filter is dropped.
    assert _sheet1_cash_total(ws) == _sheet2_grand_total(ws2) == 1000.00


def test_orphaned_cash_sale_is_journalled_and_fails_the_csv_closed(client, conn,
                                                                   orphan_store):
    """An entry whose store row is gone must be visible and must block Xero.

    Dropping it would make sheet 2 out-total sheet 1 with no warning; exporting
    it would emit the `####` placeholder into a file finance pastes into Xero."""
    orphan = orphan_store
    _orphan_sale(conn, orphan, '2095-02-04', 80000, 'ORPHAN-1')

    row = next(r for r in db.get_cash_sales_journal_stores(2095, 2) if r['store'] == orphan)
    assert row['store_code'] is None and row['sales_cents'] == 80000
    assert orphan in {i['store'] for i in db.get_cash_sales_line_items(2095, 2)}

    r = client.get('/cash/cash-sales/2095/2/mj.csv', follow_redirects=False)
    assert r.status_code in (302, 303), 'a NULL POS code must not export as ####'
    assert 'text/csv' not in r.content_type
    assert r.headers['Location'].endswith('/cash/cash-sales/2095/2')


def test_orphaned_cash_sale_keeps_the_two_recon_sheets_tying(client, conn, orphan_store):
    orphan = orphan_store
    _orphan_sale(conn, orphan, '2095-02-04', 80000, 'ORPHAN-1')
    _sales_store(conn, NEAR_DEMO_STORE, 'SC-130')
    db.save_cash_shopify_mapping('NORTHWIND PYTEST ORPHAN MONTH', NEAR_DEMO_STORE)
    up = client.post('/cash/cash-sales/2095/2/shopify', data={
        'shopify_csv': (io.BytesIO(_shopify_csv('NORTHWIND PYTEST ORPHAN MONTH',
                                                gross='0', net='0')),
                        'orphan-month.csv')}, content_type='multipart/form-data')
    assert up.status_code in (302, 303)
    # The orphan's R800 is a difference needing a reason. It is written at the
    # data layer on purpose: cash_sales_save_reasons only accepts a store that is
    # still in `stores`, so the page cannot currently clear this one.
    db.save_cash_sales_variance_reasons(2095, 2, {orphan: 'Store renamed in Cash Recon'})

    r = client.get('/cash/cash-sales/2095/2/recon.xlsx')
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.get_data()))
    ws, ws2 = wb.worksheets[0], wb.worksheets[1]
    assert orphan in {row[1] for row in ws.iter_rows(values_only=True)}
    assert orphan in {row[0] for row in ws2.iter_rows(values_only=True)}
    assert _sheet1_cash_total(ws) == _sheet2_grand_total(ws2) == 800.00


def test_stale_shopify_mapping_is_unresolved_not_a_phantom_comparison_row(client, conn):
    """A mapping pointing at a renamed store must block, not invent a pair.

    Trusting it produced a dead-name row (all Shopify, no cash) beside the real
    store (all cash, no Shopify) — two variances that net to zero and read as
    ordinary noise, while the download gate was satisfied."""
    ghost = 'PYTEST STORE RENAMED AWAY'
    _sales_store(conn, NEAR_DEMO_STORE, 'SC-130')
    db.save_cash_shopify_mapping('NORTHWIND PYTEST RENAMED', NEAR_DEMO_STORE)
    conn.execute("UPDATE cash_shopify_store_mappings SET store=? WHERE shopify_location=?",
                 (ghost, 'NORTHWIND PYTEST RENAMED'))
    conn.commit()
    assert ghost not in db.get_stores()
    up = client.post('/cash/cash-sales/2095/3/shopify', data={
        'shopify_csv': (io.BytesIO(_shopify_csv('NORTHWIND PYTEST RENAMED')), 'renamed.csv'),
    }, content_type='multipart/form-data')
    assert up.status_code in (302, 303)

    ctx = _cash_sales_context(2095, 3)
    unresolved = [r['pos_location_name'] for r in ctx['unmatched_shopify']]
    assert unresolved == ['NORTHWIND PYTEST RENAMED']
    assert 'no longer a Northwind store' in ctx['unmatched_shopify'][0]['issue']
    assert ghost not in {r['store'] for r in ctx['comparison']}
    assert not ctx['comparison_ready']
    assert client.get('/cash/cash-sales/2095/3/recon.xlsx',
                      follow_redirects=False).status_code in (302, 303)

    # Positive control: re-mapping it to a store that exists clears the block.
    fix = client.post('/cash/cash-sales/2095/3/mapping',
                      data={'shopify_location': 'NORTHWIND PYTEST RENAMED',
                            'store': NEAR_DEMO_STORE})
    assert fix.status_code in (302, 303)
    assert not _cash_sales_context(2095, 3)['unmatched_shopify']


def test_excluding_a_non_store_location_unblocks_the_workbook_and_is_reversible(client, conn):
    """"NORTHWIND Sample Sale" can never be mapped, so it would block the month forever."""
    store = _sales_store(conn, 'PYTEST EXCL STORE', 'SC-136')
    sale = _income_cat(conn, 'cash sale%')
    db.add_recon_entry(store, '2095-04-04', sale, 1500.00, 'EXCL-1', created_by=store)
    db.save_cash_shopify_mapping('NORTHWIND PYTEST EXCL', store)
    payload = (_shopify_csv('NORTHWIND PYTEST EXCL') +
               b'NORTHWIND PYTEST SAMPLE SALE,cash,NORTHWIND9002,1,123.45,0,123.45\n')
    up = client.post('/cash/cash-sales/2095/4/shopify', data={
        'shopify_csv': (io.BytesIO(payload), 'exclude-month.csv'),
    }, content_type='multipart/form-data')
    assert up.status_code in (302, 303)
    try:
        assert client.get('/cash/cash-sales/2095/4/recon.xlsx',
                          follow_redirects=False).status_code in (302, 303)

        r = client.post('/cash/cash-sales/2095/4/exclude',
                        data={'shopify_location': 'NORTHWIND PYTEST SAMPLE SALE'},
                        follow_redirects=True)
        body = r.get_data(as_text=True)
        # Changed 2026-08-12: an exclusion is scoped to the (year, month) it was
        # made in, so both the flash and the panel now name the month. The old
        # month-less wording read as a global rule — and was one: excluding while
        # viewing May silently rewrote June.
        assert 'excluded from the April 2095 comparison' in body
        assert 'NORTHWIND PYTEST SAMPLE SALE' in body and '123.45' in body

        ctx = _cash_sales_context(2095, 4)
        assert [x['pos_location_name'] for x in ctx['excluded_shopify']] == \
            ['NORTHWIND PYTEST SAMPLE SALE']
        assert not ctx['unmatched_shopify']
        # Excluded from the COMPARISON only — its money still counts in the
        # headline Shopify net, so the KPI never understates the upload.
        assert ctx['shopify_totals']['net'] == 1623.45
        assert ctx['comparison_totals']['net'] == 1500.00
        assert ctx['shopify_totals']['outside'] == 123.45
        # Scoped to the month it was made in: May must be untouched.
        assert _excluded_shopify_locations(2095, 5) == []
        assert client.get('/cash/cash-sales/2095/4/recon.xlsx').status_code == 200
        # …and the downloaded workbook carries it too, or the artefact finance
        # keeps would be missing R123.45 the screen shows.
        wb = openpyxl.load_workbook(io.BytesIO(
            client.get('/cash/cash-sales/2095/4/recon.xlsx').get_data()))
        rows = list(wb.worksheets[0].iter_rows(values_only=True))
        assert ('NORTHWIND PYTEST SAMPLE SALE', 123.45) in {(x[0], x[6]) for x in rows}
        assert 'TOTAL (all Shopify cash)' in {x[0] for x in rows}

        back = client.post('/cash/cash-sales/2095/4/exclude',
                           data={'shopify_location': 'NORTHWIND PYTEST SAMPLE SALE',
                                 'restore': '1'})
        assert back.status_code in (302, 303)
        ctx = _cash_sales_context(2095, 4)
        assert not ctx['excluded_shopify']
        assert [x['pos_location_name'] for x in ctx['unmatched_shopify']] == \
            ['NORTHWIND PYTEST SAMPLE SALE']
        assert client.get('/cash/cash-sales/2095/4/recon.xlsx',
                          follow_redirects=False).status_code in (302, 303)
    finally:
        db.set_setting(CASH_SHOPIFY_EXCLUDED_KEY, '')


def test_exclude_route_refuses_a_store_session_a_bad_month_and_a_blank_name(client,
                                                                           staff_client):
    """Same guards the sibling cash-sales POSTs carry — it writes app_settings."""
    def excluded():
        return db.get_setting(CASH_SHOPIFY_EXCLUDED_KEY) or ''

    assert excluded() == ''
    store_client, _ = staff_client
    denied = store_client.post('/cash/cash-sales/2095/4/exclude',
                               data={'shopify_location': 'NORTHWIND PYTEST SAMPLE SALE'},
                               follow_redirects=False)
    assert denied.status_code in (302, 303) and excluded() == ''

    for year, month in ((2095, 13), (1999, 4)):
        r = client.post(f'/cash/cash-sales/{year}/{month}/exclude',
                        data={'shopify_location': 'NORTHWIND PYTEST SAMPLE SALE'},
                        follow_redirects=False)
        assert r.status_code in (302, 303)
        assert r.headers['Location'].endswith('/cash'), (year, month)
        assert excluded() == ''

    blank = client.post('/cash/cash-sales/2095/4/exclude',
                        data={'shopify_location': '   '}, follow_redirects=True)
    assert 'Choose a Shopify location to exclude.' in blank.get_data(as_text=True)
    assert excluded() == ''


def test_exclude_route_requires_a_csrf_token(db_copy):
    import app as a
    a.app.config['TESTING'] = True
    a.app.config['WTF_CSRF_ENABLED'] = True
    try:
        c = a.app.test_client()
        with c.session_transaction() as sess:
            sess['admin'] = True
            sess['admin_role'] = 'super'
        r = c.post('/cash/cash-sales/2095/4/exclude',
                   data={'shopify_location': 'NORTHWIND PYTEST SAMPLE SALE'},
                   headers={'Accept': 'application/json'})
        assert r.status_code == 400
        assert (db.get_setting(CASH_SHOPIFY_EXCLUDED_KEY) or '') == ''
    finally:
        a.app.config['WTF_CSRF_ENABLED'] = False


def test_set_store_xero_keeps_the_cash_sales_label_unless_told_otherwise(conn):
    """Three-argument callers set the EXPENSE mapping; the cash-split label is a
    different piece of finance's setup and must not be wiped as a side effect."""
    def label():
        return conn.execute("SELECT cash_sales_label FROM stores WHERE name=?",
                            (store,)).fetchone()['cash_sales_label']

    store = _sales_store(conn, 'PYTEST LABEL STORE', 'SC-137')
    db.set_store_xero(store, 'Tracking A', 'SC-137', 'Finance Short Label')
    assert label() == 'Finance Short Label'

    db.set_store_xero(store, 'Tracking B', 'SC-138')          # label omitted
    assert label() == 'Finance Short Label'
    assert db.get_store_xero(store) == {'tracking_name': 'Tracking B',
                                        'store_code': 'SC-138'}
    # The label is what the journal helper reports as the store's finance name.
    assert next(r['tracking_name'] for r in db.get_cash_sales_journal_stores(2095, 5)
                if r['store'] == store) == 'Finance Short Label'

    db.set_store_xero(store, 'Tracking B', 'SC-138', '')      # explicit clear
    assert label() is None
    assert next(r['tracking_name'] for r in db.get_cash_sales_journal_stores(2095, 5)
                if r['store'] == store) == store
