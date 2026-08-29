"""Cross-card finance-pack exports preserve atomic transaction evidence."""
import datetime as dt
import io
import uuid
import zipfile

import openpyxl

from northwind.data import database as db
from northwind.cards.parser import CardSnapshot, StatementLine


def _make_export_card():
    token = uuid.uuid4().hex[:10]
    name = f'ZZZ Finance Pack {token} Credit Card'
    snapshot = CardSnapshot(
        card_name=name,
        display_name='Finance Pack Tester',
        period_start=dt.date(2026, 9, 1),
        period_end=dt.date(2026, 9, 30),
        as_at=dt.date(2026, 9, 30),
        statement_balance_cents=None,
        lines=[
            StatementLine(
                line_date=dt.date(2026, 9, 8),
                reference='PACK READY MERCHANT',
                amount_cents=-12345,
                category='spend',
                reconciled=False,
                fingerprint=f'finance-pack-ready-{token}',
                occurrence=0,
            ),
            StatementLine(
                line_date=dt.date(2026, 9, 9),
                reference='PACK WAITING MERCHANT',
                amount_cents=-6789,
                category='spend',
                reconciled=False,
                fingerprint=f'finance-pack-waiting-{token}',
                occurrence=0,
            ),
        ],
        duplicates_removed_by_xero=0,
        source_filename='finance-pack-test.xlsx',
    )
    result = db.import_card_snapshot(snapshot)
    card_id = result['card_id']
    conn = db.get_db()
    try:
        statement_id = conn.execute(
            "SELECT id FROM cc_statements WHERE card_id=? AND year=2026 AND month=9",
            (card_id,),
        ).fetchone()['id']
    finally:
        conn.close()
    lines = db.get_cc_statement_lines(statement_id)
    ready_id, waiting_id = [row['id'] for row in lines]
    receipt_id = db.add_cc_receipt(
        card_id,
        statement_id,
        f'{card_id}/2026-09/pack-ready.pdf',
        '../../unsafe pack receipt.pdf',
        'application/pdf',
        'pytest',
        content_hash=f'finance-pack-{card_id}-{token}',
    )
    db.link_cc_receipt(receipt_id, ready_id)
    db.set_cc_receipt_download_name(receipt_id, '../../unsafe pack receipt.pdf')
    db.set_cc_line_reason(ready_id, '=HYPERLINK("https://example.test")')
    db.set_cc_line_location(ready_id, 'HQ')
    db.set_cc_lines_submitted([ready_id], 'cardholder@test.co')
    db.set_cc_line_ai_coding(
        ready_id,
        '6230',
        'Motor Vehicle Expenses',
        'high',
        False,
        'Likely fuel purchase',
        'ai',
    )
    return card_id, statement_id, ready_id, waiting_id, receipt_id


def _query(card_id, status='unreconciled'):
    return [
        ('cards_present', '1'),
        ('card_id', str(card_id)),
        ('period', '2026-09'),
        ('status', status),
    ]


def test_review_offers_export_of_the_current_filters(client):
    card_id, _, _, _, _ = _make_export_card()
    page = client.get('/cards/review', query_string=_query(card_id)).get_data(as_text=True)

    assert 'Export current view' in page
    assert 'Finance pack' in page
    assert 'Transactions only' in page
    assert '/cards/review/export.zip' in page
    assert '/cards/review/export.xlsx' in page
    assert 'no Xero API or import-format dependency' in page


def test_transactions_workbook_is_filtered_atomic_and_excel_safe(client):
    card_id, _, ready_id, waiting_id, _ = _make_export_card()
    response = client.get(
        '/cards/review/export.xlsx',
        query_string=_query(card_id, status='ready'),
    )

    assert response.status_code == 200
    assert response.mimetype == (
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    workbook = openpyxl.load_workbook(io.BytesIO(response.data), data_only=False)
    sheet = workbook['Transactions']
    headers = {cell.value: index for index, cell in enumerate(sheet[1], start=1)}
    assert sheet.max_row == 2
    assert sheet.cell(2, headers['Transaction ID']).value == ready_id
    assert sheet.cell(2, headers['Workflow status']).value == 'Ready for Xero'
    assert sheet.cell(2, headers['Statement amount (R)']).value == -123.45
    assert sheet.cell(2, headers['Reason']).value.startswith("'=")
    assert sheet.cell(2, headers['Account source']).value == (
        'AI suggestion — check in Xero'
    )
    assert waiting_id not in {
        sheet.cell(row, headers['Transaction ID']).value
        for row in range(2, sheet.max_row + 1)
    }
    assert workbook['Pack summary']['B5'].value == 'Finance Pack Tester'


def test_transactions_workbook_surfaces_open_vat_invoice_request(client):
    card_id, _, ready_id, _, _ = _make_export_card()
    db.set_cc_line_vat_invoice_required(ready_id, True, 'pytest-finance')

    response = client.get(
        '/cards/review/export.xlsx',
        query_string=_query(card_id),
    )

    workbook = openpyxl.load_workbook(io.BytesIO(response.data), data_only=True)
    sheet = workbook['Transactions']
    headers = {cell.value: index for index, cell in enumerate(sheet[1], start=1)}
    row_number = next(
        row for row in range(2, sheet.max_row + 1)
        if sheet.cell(row, headers['Transaction ID']).value == ready_id
    )
    assert sheet.cell(
        row_number, headers['Workflow status']).value == 'VAT invoice requested'
    assert sheet.cell(
        row_number, headers['VAT invoice requested']).value == 'Yes'
    assert sheet.cell(
        row_number, headers['VAT invoice requested at']).value
    assert sheet.cell(
        row_number, headers['VAT invoice requested by']).value == 'pytest-finance'


def test_finance_pack_groups_receipts_per_transaction_and_sanitizes_names(
        client, monkeypatch):
    card_id, _, ready_id, waiting_id, _ = _make_export_card()
    monkeypatch.setattr(
        'northwind.cards.admin_review.storage.read',
        lambda path: b'%PDF finance pack test',
    )

    response = client.get(
        '/cards/review/export.zip',
        query_string=_query(card_id),
    )

    assert response.status_code == 200
    assert response.mimetype == 'application/zip'
    assert response.headers['X-NORTHWIND-Receipt-Files'] == '1'
    assert response.headers['X-NORTHWIND-Missing-Receipt-Files'] == '0'
    with zipfile.ZipFile(io.BytesIO(response.data)) as archive:
        names = archive.namelist()
        assert 'transactions.xlsx' in names
        assert 'README.txt' in names
        receipt_name = next(name for name in names if name.startswith('receipts/'))
        assert f'/{ready_id}_Finance_Pack_Tester_' in receipt_name
        assert receipt_name.endswith('unsafe_pack_receipt.pdf')
        assert '..' not in receipt_name
        assert not any(f'/{waiting_id}_' in name for name in names)
        workbook = openpyxl.load_workbook(
            io.BytesIO(archive.read('transactions.xlsx')),
            data_only=True,
        )
        assert workbook['Transactions'].max_row == 3


def test_missing_receipt_blob_is_reported_without_breaking_pack(client, monkeypatch):
    card_id, _, ready_id, _, _ = _make_export_card()

    def missing(_path):
        raise FileNotFoundError

    monkeypatch.setattr('northwind.cards.admin_review.storage.read', missing)
    response = client.get(
        '/cards/review/export.zip',
        query_string=_query(card_id, status='ready'),
    )

    assert response.status_code == 200
    assert response.headers['X-NORTHWIND-Receipt-Files'] == '0'
    assert response.headers['X-NORTHWIND-Missing-Receipt-Files'] == '1'
    with zipfile.ZipFile(io.BytesIO(response.data)) as archive:
        readme = archive.read('README.txt').decode('utf-8')
        assert f'Transaction {ready_id}: Missing stored file' in readme
        workbook = openpyxl.load_workbook(
            io.BytesIO(archive.read('transactions.xlsx')),
            data_only=True,
        )
        sheet = workbook['Transactions']
        headers = {cell.value: index for index, cell in enumerate(sheet[1], start=1)}
        assert 'Missing stored file' in sheet.cell(
            2, headers['Receipt export note']).value
