"""Store cash reconciliation — ledger + admin views.

Two audiences:
  • A logged-in STORE gets a single-page monthly ledger (cash_ledger) scoped to
    its own store, can add entries (every entry now needs a reason), and can
    delete its own entries. It may NOT adjust the opening float.
  • An ADMIN gets a drill-down: an all-stores overview grid (cash_home) →
    a per-store daily summary (cash_summary) → a single day (cash_day) →
    the full month journal (cash_ledger, with float-adjust + edit/delete).

Money is integer cents in the DB; amounts are Rands at the template boundary.
Store sessions are scoped to their own store via _deny_other_store(); admin-only
endpoints additionally bounce any store session (defense in depth). Entry deletes
and edits are scoped to the entry's real owner store, never a posted field.
"""
import calendar
import io
import json
import re
import zipfile
from datetime import datetime, date, timedelta
from decimal import Decimal, ROUND_HALF_UP
from flask import render_template, request, redirect, url_for, flash, session, Response, send_file
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from northwind.data import database as db
from northwind.services import money
from northwind.cash import mj as cash_mj
from northwind.cash import shopify as shopify_csv
from northwind.core import app

# App-settings keys for the consolidated cash-sales journal.
CASH_SALES_CONTRA_KEY = 'cash_sales_contra_code'
CASH_SALES_NARRATION_KEY = 'cash_sales_narration'
DEFAULT_CASH_SALES_NARRATION = 'Retail cash sales journal'
# Shopify POS locations finance has confirmed are not a Northwind store at all, keyed
# by month — see _load_excluded_shopify() for the shape and why it lives here.
CASH_SHOPIFY_EXCLUDED_KEY = 'cash_shopify_excluded_locations'


def _bounce(message, endpoint, category='danger', **values):
    """Flash one message and redirect back — every guard in this file does this.

    It was written out seventeen times, three of them back to back, which buried
    the rule each one was actually enforcing."""
    flash(message, category)
    return redirect(url_for(endpoint, **values))


def _stores():
    return db.get_stores()


def _store_session():
    """The store a non-admin (store) session is locked to, else None.

    Admins (and local dev with auth disabled) return None — full access."""
    if session.get('admin'):
        return None
    return session.get('staff_store')


def _deny_other_store(store):
    """Redirect a store session away from another store's cash recon."""
    mine = _store_session()
    if mine and store != mine:
        return _bounce('You can only access your own store’s cash recon.', 'cash_store', store=mine)
    return None


def _deny_demo_journal(store):
    """Keep a demo-seeded store out of every Xero journal path.

    Its ledger and dashboards stay open — only the exports refuse it, because
    `scripts/seed_demo_cash_recon.py` money is fake and would be pasted into an
    imported ledger workbook. Returns a redirect to block, or None to allow."""
    if db.is_demo_store(store):
        return _bounce('Demo stores are excluded from Xero journals — their cash is '
                       'seeded test data.', 'cash_store', 'warning', store=store)
    return None


def _require_admin():
    """Bounce a store session away from an admin-only view (defense in depth).

    Returns a redirect response to block, or None to allow."""
    mine = _store_session()
    if mine is not None:
        return _bounce('That area is for administrators only.', 'cash_store', store=mine)
    return None


def _month_from_request(default_now=True):
    """Resolve ?year/?month from the query string, falling back to now."""
    now = datetime.now()
    year = request.args.get('year', type=int) or now.year
    month = request.args.get('month', type=int) or now.month
    if db.validate_month_year(year, month) is None:
        year, month = now.year, now.month
    return year, month


def _month_bounds(y, m):
    """(first, last) date objects of month m in year y."""
    return date(y, m, 1), date(y, m, calendar.monthrange(y, m)[1])


def _range_from_request():
    """Resolve an inclusive [start, end] date range from ?start/?end (ISO), or
    default to the current calendar month. Returns (start_iso, end_iso, s, e)."""
    now = datetime.now()

    def parse(v):
        try:
            return datetime.strptime((v or '').strip(), '%Y-%m-%d').date()
        except ValueError:
            return None

    s = parse(request.args.get('start'))
    e = parse(request.args.get('end'))
    if s is None or e is None:
        s, e = _month_bounds(now.year, now.month)
    if e < s:
        s, e = e, s
    return s.isoformat(), e.isoformat(), s, e


@app.route('/cash')
def cash_home():
    # A logged-in store skips the overview — straight to its own ledger.
    mine = _store_session()
    if mine:
        return redirect(url_for('cash_store', store=mine))
    # Admin: all-stores overview grid for the chosen date range (default = this
    # month). Rows expand inline (AJAX -> cash_store_days) for a day breakdown.
    start_iso, end_iso, s, e = _range_from_request()
    rows = db.get_recon_overview_range(start_iso, end_iso)
    # Estate-wide totals for the grid footer.
    totals = {k: round(sum(r[k] for r in rows), 2) for k in
              ('opening', 'total_in', 'total_expense', 'total_banked', 'total_adjust', 'closing')}
    totals['entry_count'] = sum(r['entry_count'] for r in rows)

    # Exceptions across all stores (mirrors the dashboard's attention logic) so
    # the default overview flags problems at a glance, not only the /cash/dashboard.
    activity = db.get_recon_activity_summary(db.get_stores(), start_iso, end_iso)
    stale_before = min(e, date.today()) - timedelta(days=3)
    attention = []
    for row in rows:
        reasons = []
        latest = activity['by_store'].get(row['store'], {}).get('latest_entry_date')
        if not row['entry_count']:
            reasons.append('No entries in this range')
        elif latest and date.fromisoformat(latest) < stale_before:
            reasons.append(f"Stale — last entry {latest}")
        if row['closing'] < 0:
            reasons.append('Negative closing float')
        if row['total_adjust']:
            reasons.append(f"Adjustment R {row['total_adjust']:,.2f}")
        if reasons:
            attention.append({'store': row['store'], 'reasons': reasons})

    # Month-stepping chevrons are anchored on the start date's month.
    p_first, _ = _month_bounds(*( (s.year - 1, 12) if s.month == 1 else (s.year, s.month - 1) ))
    prev_first, prev_last = _month_bounds(p_first.year, p_first.month)
    n_year, n_month = (s.year + 1, 1) if s.month == 12 else (s.year, s.month + 1)
    next_first, next_last = _month_bounds(n_year, n_month)

    # Friendly label: a whole calendar month shows as "July 2026", else a range.
    _, last = _month_bounds(s.year, s.month)
    is_full_month = (s.day == 1 and e == last)
    if is_full_month:
        range_label = f"{db.MONTH_FULL[s.month]} {s.year}"
    else:
        range_label = f"{s.strftime('%d %b %Y')} – {e.strftime('%d %b %Y')}"

    return render_template(
        'cash_overview.html', rows=rows, start=start_iso, end=end_iso,
        range_label=range_label, is_admin=True,
        prev_start=prev_first.isoformat(), prev_end=prev_last.isoformat(),
        next_start=next_first.isoformat(), next_end=next_last.isoformat(),
        cash_sales_year=s.year, cash_sales_month=s.month,
        cash_sales_label=f"{db.MONTH_FULL[s.month]} {s.year}", totals=totals,
        attention=attention,
        attention_stores={item['store'] for item in attention})


@app.route('/cash/<store>/days')
def cash_store_days(store):
    """Admin-only HTML fragment: one store's per-day breakdown for a date range,
    injected inline under its overview row (no page navigation)."""
    blocked = _require_admin()
    if blocked:
        return blocked
    if store not in _stores():
        return ('Unknown store.', 404)
    start_iso, end_iso, s, e = _range_from_request()
    data = db.get_recon_range(store, start_iso, end_iso)
    return render_template('cash_store_days.html', store=store, data=data,
                           start=start_iso, end=end_iso)


@app.route('/cash/<store>')
def cash_store(store):
    blocked = _deny_other_store(store)
    if blocked:
        return blocked
    now = datetime.now()
    # Store session → its month journal; admin → the per-store daily summary.
    if _store_session() is not None:
        return redirect(url_for('cash_ledger', store=store, year=now.year, month=now.month))
    return redirect(url_for('cash_summary', store=store, year=now.year, month=now.month))


@app.route('/cash/<store>/<int:year>/<int:month>/summary')
def cash_summary(store, year, month):
    blocked = _require_admin()
    if blocked:
        return blocked
    if store not in _stores():
        return _bounce('Unknown store.', 'cash_home')
    if db.validate_month_year(year, month) is None:
        return redirect(url_for('cash_home'))

    data = db.get_recon_daily_summary(store, year, month)
    prev_year, prev_month = (year - 1, 12) if month == 1 else (year, month - 1)
    next_year, next_month = (year + 1, 1) if month == 12 else (year, month + 1)
    return render_template(
        'cash_summary.html', store=store, year=year, month=month,
        month_name=db.MONTH_FULL[month], data=data, is_admin=True,
        prev_year=prev_year, prev_month=prev_month,
        next_year=next_year, next_month=next_month)


@app.route('/cash/<store>/day/<date>')
def cash_day(store, date):
    blocked = _require_admin()
    if blocked:
        return blocked
    if store not in _stores():
        return _bounce('Unknown store.', 'cash_home')
    try:
        parsed = datetime.strptime(date, '%Y-%m-%d')
    except ValueError:
        return _bounce('Invalid date.', 'cash_summary', store=store,
                       year=datetime.now().year, month=datetime.now().month)

    data = db.get_recon_day(store, date)
    return render_template(
        'cash_day.html', store=store, date=date, data=data, is_admin=True,
        year=parsed.year, month=parsed.month,
        month_name=db.MONTH_FULL[parsed.month])


@app.route('/cash/<store>/<int:year>/<int:month>')
def cash_ledger(store, year, month):
    blocked = _deny_other_store(store)
    if blocked:
        return blocked
    if store not in _stores():
        return _bounce('Unknown store.', 'cash_home')
    if db.validate_month_year(year, month) is None:
        return redirect(url_for('cash_store', store=store))

    data = db.get_recon_month(store, year, month)
    categories = db.get_recon_categories()
    prev_year, prev_month = (year - 1, 12) if month == 1 else (year, month - 1)
    next_year, next_month = (year + 1, 1) if month == 12 else (year, month + 1)
    # The date field must stay inside the month on screen. It used to default to
    # TODAY whatever month was being viewed, so an entry added from (say) the July
    # page was filed under today's date in August, the page it redirected back to
    # never showed it, and the store read that as "it didn't save".
    first, last = _month_bounds(year, month)
    today = datetime.now().date()
    default_date = min(max(today, first), last)
    return render_template(
        'cash_ledger.html', store=store, year=year, month=month,
        month_name=db.MONTH_FULL[month], data=data, categories=categories,
        today=default_date.isoformat(),
        date_min=first.isoformat(),
        # No future-dating, but never an impossible range on a future month.
        date_max=(min(last, today) if today >= first else last).isoformat(),
        is_admin=(_store_session() is None),
        prev_year=prev_year, prev_month=prev_month,
        next_year=next_year, next_month=next_month)


@app.route('/cash/<store>/<int:year>/<int:month>/add', methods=['POST'])
def cash_add_entry(store, year, month):
    blocked = _deny_other_store(store)
    if blocked:
        return blocked
    if db.validate_month_year(year, month) is None:
        return _bounce('That month is not valid.', 'cash_store', store=store)

    def _back(message):
        return _bounce(message, 'cash_ledger', store=store, year=year, month=month)

    # A missing date falls back to a day INSIDE the posted month (today when that
    # month is the current one), never a bare today — see the in-month check below.
    first, last = _month_bounds(year, month)
    fallback_date = min(max(datetime.now().date(), first), last).isoformat()
    try:
        category_id = int(request.form['category_id'])
        amount = float(request.form['amount'])
        entry_date = request.form.get('entry_date', '').strip() or fallback_date
        note = request.form.get('note', '')
    except (KeyError, ValueError):
        return _back('Please choose a category and enter a valid amount.')

    if amount <= 0:
        return _back('Amount must be greater than zero.')

    cat = next((c for c in db.get_recon_categories() if c['id'] == category_id), None)
    if cat is None:
        return _back('Please choose a valid category.')

    # Every entry now needs a reason (stored in the note column) — enforce it
    # server-side too, client validation can be bypassed.
    if not note.strip():
        return _back('Please enter a reason for this entry.')

    # The entry must be a REAL day inside the month it was posted from. Nothing
    # downstream parses entry_date — add_recon_entry stores the string verbatim —
    # so a prefix check alone would let '2099-07' or '2099-07-99' through and leave
    # a row whose date breaks the day grouping, the admin edit form's date input and
    # the day-drilldown link. Out-of-month is a real mistake too: the entry lands on
    # a page the store isn't looking at, the ledger they're on doesn't change, and it
    # reads as a save that silently failed. The date input is bounded to the month;
    # this is the server-side half of that.
    try:
        parsed_date = date.fromisoformat(entry_date)
    except ValueError:
        return _back('That date is not a real date.')
    if not (first <= parsed_date <= last):
        return _back(f'That date is outside {db.MONTH_FULL[month]} {year} — '
                     f'switch to the right month first.')
    entry_date = parsed_date.isoformat()   # normalised, so the stored form is exact

    db.add_recon_entry(store, entry_date, category_id, amount, note, created_by=store)
    # ?added=1 tells the page this POST actually landed, so the browser can drop
    # its saved draft (cash-ledger.js). Every OTHER way of arriving back here —
    # a validation bounce above, an expired-session CSRF redirect, a tablet
    # reloading an evicted tab — keeps the draft and refills the form.
    return _bounce('Entry added.', 'cash_ledger', 'success',
                   store=store, year=year, month=month, added=1)


@app.route('/cash/entry/<int:entry_id>/edit', methods=['POST'])
def cash_edit_entry(entry_id):
    blocked = _require_admin()
    if blocked:
        return blocked
    # Scope to the entry's REAL owner store, never a posted field.
    entry_store = db.get_recon_entry_store(entry_id)
    if entry_store is None:
        return _bounce('Entry not found.', 'cash_home')

    year = request.form.get('year', type=int)
    month = request.form.get('month', type=int)

    def _back(message, category='danger'):
        if year and month:
            return _bounce(message, 'cash_ledger', category,
                           store=entry_store, year=year, month=month)
        return _bounce(message, 'cash_store', category, store=entry_store)

    try:
        category_id = int(request.form['category_id'])
        amount = float(request.form['amount'])
        entry_date = request.form.get('entry_date', '').strip() or datetime.now().strftime('%Y-%m-%d')
        note = request.form.get('note', '')
    except (KeyError, ValueError):
        return _back('Please choose a category and enter a valid amount.')

    if amount <= 0:
        return _back('Amount must be greater than zero.')

    cat = next((c for c in db.get_recon_categories() if c['id'] == category_id), None)
    if cat is None:
        return _back('Please choose a valid category.')

    if not note.strip():
        return _back('Please enter a reason for this entry.')

    db.update_recon_entry(entry_id, entry_date, category_id, amount, note)
    return _back('Entry updated.', 'success')


@app.route('/cash/<store>/<int:year>/<int:month>/opening', methods=['POST'])
def cash_set_opening(store, year, month):
    # Admin-only: adjusting the float is not a store action (defense in depth —
    # the card is hidden from stores, but a store session is refused here too).
    blocked = _require_admin()
    if blocked:
        return blocked
    if store not in _stores():
        return _bounce('Unknown store.', 'cash_home')
    try:
        opening = float(request.form['opening'])
    except (KeyError, ValueError):
        return _bounce('Enter a valid opening balance.', 'cash_ledger',
                       store=store, year=year, month=month)
    db.set_recon_opening(store, year, month, opening)
    return _bounce('Opening balance updated.', 'cash_ledger', 'success',
                   store=store, year=year, month=month)


@app.route('/cash/entry/<int:entry_id>/delete', methods=['POST'])
def cash_delete_entry(entry_id):
    # Scope to the entry's REAL owner, never the posted store field — otherwise a
    # store could delete another store's entry by posting its own store name.
    entry_store = db.get_recon_entry_store(entry_id)
    if entry_store is None:
        return _bounce('Entry not found.', 'cash_home')
    blocked = _deny_other_store(entry_store)
    if blocked:
        return blocked
    year = request.form.get('year', type=int)
    month = request.form.get('month', type=int)
    db.delete_recon_entry(entry_id)
    flash('Entry removed.', 'info')
    if year and month:
        return redirect(url_for('cash_ledger', store=entry_store, year=year, month=month))
    return redirect(url_for('cash_store', store=entry_store))


# ── Xero MJ: admin setup (account codes + VAT + store tracking) ───────────────

@app.route('/cash/xero-setup')
def cash_xero_setup():
    blocked = _require_admin()
    if blocked:
        return blocked
    contra_code, narration_prefix = _cash_sales_settings()
    return render_template(
        'cash_xero_setup.html',
        categories=db.get_recon_categories_admin(),
        stores=db.get_store_xero_map(),
        vat_options=[('standard', cash_mj.TAX_RATES['standard']),
                     ('novat', cash_mj.TAX_RATES['novat'])],
        cash_sales_contra=contra_code,
        cash_sales_narration=narration_prefix)


@app.route('/cash/xero-setup/categories', methods=['POST'])
def cash_save_categories():
    blocked = _require_admin()
    if blocked:
        return blocked
    for cid in request.form.getlist('cat_id'):
        code = request.form.get(f'code_{cid}', '')
        vat = request.form.get(f'vat_{cid}', '')
        db.update_recon_category_xero(int(cid), code, vat)
        db.set_recon_category_active(int(cid), bool(request.form.get(f'active_{cid}')))
    return _bounce('Category Xero mappings saved.', 'cash_xero_setup', 'success')


@app.route('/cash/xero-setup/categories/add', methods=['POST'])
def cash_add_category():
    blocked = _require_admin()
    if blocked:
        return blocked
    ok, msg = db.add_recon_category(
        request.form.get('name', ''),
        request.form.get('code', ''),
        request.form.get('vat', ''))
    return _bounce(msg, 'cash_xero_setup', 'success' if ok else 'danger')


@app.route('/cash/xero-setup/stores', methods=['POST'])
def cash_save_stores():
    blocked = _require_admin()
    if blocked:
        return blocked
    for i, name in enumerate(request.form.getlist('store_name')):
        tracking = request.form.get(f'tracking_{i}', '')
        code = request.form.get(f'storecode_{i}', '')
        cash_label = request.form.get(f'cashlabel_{i}', '')
        db.set_store_xero(name, tracking, code, cash_label)
    db.invalidate_stores_cache()
    return _bounce('Store Xero mappings saved.', 'cash_xero_setup', 'success')


# ── Xero MJ: per-store expenses journal preview + CSV export ──────────────────

def _mj_context(store, year, month):
    """Shared preview data for a store-month's expenses MJ."""
    raw = db.get_recon_expense_lines(store, year, month)
    lines = []
    gross_total = net_total = 0
    for r in raw:
        vt = r['vat_type'] or cash_mj.DEFAULT_VAT
        net = cash_mj.net_cents(r['gross_cents'], vt)
        gross_total += r['gross_cents']
        net_total += net
        lines.append({
            **r, 'vat_type': vt,
            'description': cash_mj.build_description(r['category'], r['date'], r['note']),
            'net_cents': net,
            'unmapped': not (r['xero_code'] or '').strip(),
        })
    last_day = calendar.monthrange(year, month)[1]
    xero = db.get_store_xero(store)
    summary = cash_mj.journal_summary(raw)
    return {
        'store': store, 'year': year, 'month': month,
        'month_name': db.MONTH_FULL[month],
        'narration': f"{store} {db.MONTH_FULL[month]} {year}",
        'month_end': date(year, month, last_day).strftime('%d/%m/%Y'),
        'tracking_name': xero['tracking_name'],
        'dear_code': xero['store_code'],
        'rounding_code': cash_mj.ROUNDING_CODE,
        'lines': lines,
        'gross_total': gross_total, 'net_total': net_total,
        'vat_total': summary['vat_total'],
        'contra_cents': summary['contra_cents'],
        'rounding_cents': summary['rounding_cents'],
        'vat_options': [('standard', cash_mj.TAX_RATES['standard']),
                        ('novat', cash_mj.TAX_RATES['novat'])],
    }


def _dear_account_issues(store, store_code):
    """The one setup value EVERY cash journal needs: the store's POS account.

    Split out because the cash-sales journal needs only this — it carries no
    tracking at all, so running it through _expense_export_issues() could
    only ever produce a message about a column that journal does not have."""
    return [] if (store_code or '').strip() else [f'{store}: Cash In/Out (POS) account']


def _expense_export_issues(store, xero, lines):
    """Human-readable reasons an expense journal is not import-ready.

    The preview deliberately renders #### for incomplete setup so finance can see
    what needs attention. The download routes are stricter: a file presented as a
    Xero journal must have every account and tracking value resolved.
    """
    issues = []
    if not (xero.get('tracking_name') or '').strip():
        issues.append(f'{store}: Xero tracking option')
    issues.extend(_dear_account_issues(store, xero.get('store_code')))
    missing_codes = sorted({
        (ln.get('category') or ln.get('description') or 'expense line').strip()
        for ln in lines if not (ln.get('xero_code') or '').strip()
    })
    issues.extend(f'{store}: account for {label}' for label in missing_codes)
    return issues


# A money field holds plain ASCII decimal digits and nothing else. `\d` is not
# usable here: it matches fullwidth '１', which float() then happily reads as 1.
_PLAIN_DECIMAL = re.compile(r'^-?[0-9]+(\.[0-9]+)?$')


def _posted_gross_cents(raw):
    """A posted gross amount as integer cents, or None if it is not a usable one.

    `float()` was too generous in both directions: it accepts '1_0' and the
    fullwidth '１０' as ten, and it accepts magnitudes no store's cash expense
    can reach — a posted 1e300 exported a Xero CSV with a 303-digit *Amount.
    The ceilings are the Shopify parser's (`MAX_AMOUNT` / `MAX_EXPONENT`), not a
    second unrelated limit invented for the same kind of value.
    """
    text = (raw or '').strip()
    if not text:
        return 0                            # caller rejects it as "not > zero"
    if not _PLAIN_DECIMAL.match(text):
        return None
    number = Decimal(text)
    if abs(number.adjusted()) > shopify_csv.MAX_EXPONENT or abs(number) > shopify_csv.MAX_AMOUNT:
        return None
    return int((number * 100).quantize(Decimal('1'), rounding=ROUND_HALF_UP))


def _block_expense_export(issues, redirect_response):
    """Return a redirect with one fail-closed Xero setup message, or None."""
    if not issues:
        return None
    shown = issues[:6]
    suffix = f" (+{len(issues) - len(shown)} more)" if len(issues) > len(shown) else ''
    flash('Xero journal not exported. Complete: ' + '; '.join(shown) + suffix + '.', 'danger')
    return redirect_response


@app.route('/cash/<store>/<int:year>/<int:month>/mj')
def cash_mj_preview(store, year, month):
    blocked = _require_admin()
    if blocked:
        return blocked
    if store not in _stores():
        return _bounce('Unknown store.', 'cash_home')
    blocked = _deny_demo_journal(store)
    if blocked:
        return blocked
    if db.validate_month_year(year, month) is None:
        return redirect(url_for('cash_store', store=store))
    prev_year, prev_month = (year - 1, 12) if month == 1 else (year, month - 1)
    next_year, next_month = (year + 1, 1) if month == 12 else (year, month + 1)
    ctx = _mj_context(store, year, month)
    ctx.update(prev_year=prev_year, prev_month=prev_month,
               next_year=next_year, next_month=next_month)
    return render_template('cash_mj.html', **ctx)


@app.route('/cash/<store>/<int:year>/<int:month>/mj/export', methods=['POST'])
def cash_mj_export(store, year, month):
    blocked = _require_admin()
    if blocked:
        return blocked
    if store not in _stores():
        return _bounce('Unknown store.', 'cash_home')
    blocked = _deny_demo_journal(store)
    if blocked:
        return blocked
    if db.validate_month_year(year, month) is None:
        return redirect(url_for('cash_store', store=store))

    def _back(message, category='danger'):
        return _bounce(message, 'cash_mj_preview', category,
                       store=store, year=year, month=month)

    last_day = calendar.monthrange(year, month)[1]
    # DD/MM/YYYY — the date format the finance Xero org imports (proven by the
    # cash-split journal workbook). Both journals must agree; do not switch to ISO.
    month_end = date(year, month, last_day).strftime('%d/%m/%Y')
    narration = f"{store} {db.MONTH_FULL[month]} {year}"
    xero = db.get_store_xero(store)
    # The preview's descriptions are editable, so a posted line can no longer say
    # which category it came from. Take that from the ledger (keyed on the row's
    # entry id) so a missing account reads the same here as it does in the batch
    # ZIP — "account for Milk", not the whole edited description line.
    categories = {str(r['id']): r['category']
                  for r in db.get_recon_expense_lines(store, year, month)}

    lines = []
    for idx in request.form.getlist('row'):
        if not request.form.get(f'include_{idx}'):
            continue  # unticked -> excluded from the export
        gross_cents = _posted_gross_cents(request.form.get(f'gross_{idx}'))
        if gross_cents is None:
            return _back('Xero journal not exported. Every included gross amount must be valid.')
        if gross_cents <= 0:
            return _back('Xero journal not exported. '
                         'Every included gross amount must be greater than zero.')
        vat_type = request.form.get(f'vat_{idx}', cash_mj.DEFAULT_VAT)
        if vat_type not in cash_mj.TAX_RATES:
            return _back('Xero journal not exported. An included line has an invalid VAT rate.')
        lines.append({
            'category': categories.get(idx, ''),
            'description': request.form.get(f'desc_{idx}', '').strip(),
            'xero_code': request.form.get(f'code_{idx}', '').strip(),
            'vat_type': vat_type,
            'gross_cents': gross_cents,
        })

    if not lines:
        return _back('Nothing to export — every line was excluded.', 'warning')

    blocked = _block_expense_export(
        _expense_export_issues(store, xero, lines),
        redirect(url_for('cash_mj_preview', store=store, year=year, month=month)))
    if blocked:
        return blocked

    rows = cash_mj.build_balanced_rows(
        narration, month_end, xero['tracking_name'], store, xero['store_code'], lines)
    csv_text = cash_mj.to_csv(rows)
    filename = f"{store.replace(chr(34), '')} {db.MONTH_FULL[month]} {year}.csv"
    return Response(csv_text, mimetype='text/csv',
                    headers={'Content-Disposition': f'attachment; filename="{filename}"'})


# ── Consolidated cash-sales journal + Shopify reconciliation export ───────────

def _cash_sales_settings():
    """(contra_code, narration_prefix) for the cash-sales journal, with defaults."""
    contra = (db.get_setting(CASH_SALES_CONTRA_KEY) or '').strip() or cash_mj.DEFAULT_CASH_SALES_CONTRA
    narration = (db.get_setting(CASH_SALES_NARRATION_KEY) or '').strip() or DEFAULT_CASH_SALES_NARRATION
    return contra, narration


# Exclusions inherited from the newline-delimited era have no month of their own,
# so they apply to every month until they are restored somewhere — see _load_excluded_shopify.
_EXCLUDED_ALL_MONTHS = '*'


def _excluded_month_key(year, month):
    return f'{year:04d}-{month:02d}'


def _load_excluded_shopify():
    """The whole exclusion store as {'YYYY-MM': [location, …]}.

    Kept as one app_settings row (get_setting/set_setting, already used for the
    journal settings above) rather than a sentinel row in
    cash_shopify_store_mappings: that table's writer validates the store against
    get_stores(), so a sentinel would need a new database.py helper for what is
    a short list of names, not a mapping.

    JSON, and keyed by month, because the two obvious shortcuts were both wrong:
      • One name per line corrupts any name a CSV may legally contain — a quoted
        "Weird\\nLocation" saved as two names, so it could never match itself
        (blocked forever, no Restore) while silently excluding a real "Weird".
      • One flat list applied to every month, so excluding while viewing May
        rewrote June — turning a reconciled June store into an unexplained
        variance that blocked its workbook, with nothing on screen to say why.
    A value written by the old code is read back under the '*' bucket so nothing
    silently un-excludes itself; that bucket empties as each name is restored.
    """
    raw = (db.get_setting(CASH_SHOPIFY_EXCLUDED_KEY) or '').strip()
    if not raw:
        return {}
    try:
        data = json.loads(raw)
    except ValueError:
        data = None
    if not isinstance(data, dict):
        return {_EXCLUDED_ALL_MONTHS: [n.strip() for n in raw.splitlines() if n.strip()]}
    return {str(key): [str(n) for n in names if str(n).strip()]
            for key, names in data.items() if isinstance(names, list)}


def _excluded_shopify_locations(year, month):
    """POS location names finance excluded from THIS month's comparison, in order.

    The live export carries rows like "NORTHWIND Sample Sale" that will never be a
    store, and save_cash_shopify_mapping refuses anything that is not one — so
    without an explicit exclusion those locations block the recon workbook for
    that month forever."""
    data = _load_excluded_shopify()
    names, seen = [], set()
    for name in (data.get(_EXCLUDED_ALL_MONTHS, [])
                 + data.get(_excluded_month_key(year, month), [])):
        if name.casefold() not in seen:
            seen.add(name.casefold())
            names.append(name)
    return names


def _save_excluded_shopify_location(year, month, location, restore):
    """Exclude one location from (year, month)'s comparison, or put it back."""
    data = _load_excluded_shopify()
    key = _excluded_month_key(year, month)
    fold = location.casefold()
    kept = [n for n in data.get(key, []) if n.casefold() != fold]
    data[key] = kept if restore else kept + [location]
    if restore:
        # An inherited '*' name belongs to no month, so restoring it wherever it
        # is showing is the only way it can ever be undone.
        data[_EXCLUDED_ALL_MONTHS] = [n for n in data.get(_EXCLUDED_ALL_MONTHS, [])
                                      if n.casefold() != fold]
    data = {k: v for k, v in data.items() if v}
    db.set_setting(CASH_SHOPIFY_EXCLUDED_KEY,
                   json.dumps(data, ensure_ascii=False, sort_keys=True) if data else '')


def _cash_sales_context(year, month):
    """Per-store cash-sales figures + journal metadata for a month."""
    rows = db.get_cash_sales_journal_stores(year, month)
    contra_code, narration_prefix = _cash_sales_settings()
    last_day = calendar.monthrange(year, month)[1]
    total_cents = sum(r['sales_cents'] for r in rows)
    unmapped_count = sum(1 for r in rows if not (r['store_code'] or '').strip())

    # Group the individual Cash Sale entries per store for the drill-down.
    items_by_store = {}
    for it in db.get_cash_sales_line_items(year, month):
        items_by_store.setdefault(it['store'], []).append({
            'date': it['date'],
            'sale_no': (it['sale_no'] or '').strip(),
            'amount': money.to_rands(it['amount_cents'])})

    stores = db.get_stores()
    upload = db.get_cash_shopify_upload(year, month)
    shopify_rows = db.get_cash_shopify_summary(year, month) if upload else []
    excluded_names = _excluded_shopify_locations(year, month)
    excluded_folded = {n.casefold() for n in excluded_names}
    known_stores = set(stores)
    mapped = {}
    unresolved = []
    excluded = []
    for r in shopify_rows:
        if r['pos_location_name'].casefold() in excluded_folded:
            excluded.append(r)
        elif not r['store']:
            r['issue'] = 'Not linked to a Northwind store yet.'
            unresolved.append(r)
        elif r['store'] not in known_stores:
            # cash_shopify_store_mappings.store is free text with no FK, so
            # renaming a store leaves its mapping pointing at nothing. Trusting
            # it would invent a comparison row for a store that no longer exists
            # (all Shopify, no cash) AND leave the real store showing no Shopify
            # — two phantom differences that both satisfy the download gate.
            r['issue'] = (f'Mapped to “{r["store"]}”, which is no longer a Northwind store. '
                          'Re-map it to the store it became.')
            unresolved.append(r)
        else:
            dest = mapped.setdefault(r['store'], {
                'locations': [], 'transactions': 0, 'gross_cents': 0,
                'refunded_cents': 0, 'net_cents': 0})
            dest['locations'].append(r['pos_location_name'])
            for key in ('transactions', 'gross_cents', 'refunded_cents', 'net_cents'):
                dest[key] += r[key]
    # Anything stored has to be visible with a way to undo it. An exclusion whose
    # location is not in this month's upload (a name carried over from the
    # global-list era, or a location that vanished from a replaced CSV) would
    # otherwise sit in app_settings rendered nowhere, with no Restore button.
    present = {r['pos_location_name'].casefold() for r in shopify_rows}
    excluded.extend({'pos_location_name': name, 'store': None, 'transactions': 0,
                     'gross_cents': 0, 'refunded_cents': 0, 'net_cents': 0, 'absent': True}
                    for name in excluded_names if name.casefold() not in present)
    app_by_store = {r['store']: r for r in rows}
    reasons = db.get_cash_sales_variance_reasons(year, month)
    compare = []
    for store in sorted(set(app_by_store) | set(mapped)):
        app_row = app_by_store.get(store, {})
        shop = mapped.get(store, {})
        cash_cents = int(app_row.get('sales_cents', 0))
        net_cents = int(shop.get('net_cents', 0))
        comparison_row = {
            'store': store,
            'display_store': app_row.get('tracking_name') or store,
            'shopify_locations': ', '.join(shop.get('locations', [])),
            'transactions': int(shop.get('transactions', 0)),
            'gross_cents': int(shop.get('gross_cents', 0)),
            'refunded_cents': int(shop.get('refunded_cents', 0)),
            'net_cents': net_cents,
            'cash_cents': cash_cents,
            'diff_cents': cash_cents - net_cents,
            'reason': reasons.get(store, ''),
            'needs_reason': abs(cash_cents - net_cents) > 100,
        }
        for key in ('gross', 'refunded', 'net', 'cash', 'diff'):
            comparison_row[key] = money.to_rands(comparison_row[f'{key}_cents'])
        compare.append(comparison_row)

    for row in unresolved + excluded:
        for key in ('gross', 'refunded', 'net'):
            row[key] = money.to_rands(row[f'{key}_cents'])
    missing_reason_count = (
        sum(1 for r in compare if r['needs_reason'] and not r['reason']) if upload else 0)
    active_store_count = sum(1 for r in rows if r['sales_cents'])

    # Headline Shopify figures cover the WHOLE upload. Summing only the compared
    # stores understates Shopify whenever a location is still unresolved or has
    # been excluded, and a partial total presented as the month's Shopify net is
    # exactly the mistake a reviewer cannot see.
    upload_net_cents = sum(r['net_cents'] for r in shopify_rows)
    outside_cents = upload_net_cents - sum(r['net_cents'] for r in compare)

    return {
        'year': year, 'month': month, 'month_name': db.MONTH_FULL[month],
        'rows': [{'store': r['store'], 'tracking_name': r['tracking_name'],
                  'store_code': r['store_code'],
                  'sales': money.to_rands(r['sales_cents']),
                  'unmapped': not (r['store_code'] or '').strip(),
                  'entries': items_by_store.get(r['store'], [])} for r in rows],
        'total': money.to_rands(total_cents),
        'active_store_count': active_store_count,
        'zero_store_count': len(rows) - active_store_count,
        'contra_code': contra_code,
        'narration': f"{narration_prefix} - {db.MONTH_FULL[month]} {year}",
        'month_end': date(year, month, last_day).strftime('%d/%m/%Y'),
        'stores': stores,
        'unmapped_count': unmapped_count,
        # A #### account makes the CSV useful as a diagnostic, but not valid as
        # a Xero import. Keep the download fail-closed until every participating
        # store has its POS cash account mapped on the setup page.
        'xero_ready': total_cents != 0 and unmapped_count == 0,
        'shopify_upload': upload,
        'comparison': compare,
        'unmatched_shopify': unresolved,
        'excluded_shopify': excluded,
        'missing_reason_count': missing_reason_count,
        'comparison_ready': bool(upload) and not unresolved and missing_reason_count == 0,
        # Totals of the comparison table itself (the stores actually compared).
        'comparison_totals': {
            'transactions': sum(r['transactions'] for r in compare),
            **{key: money.to_rands(sum(r[f'{key}_cents'] for r in compare))
               for key in ('gross', 'refunded', 'net', 'cash', 'diff')},
        },
        # Totals of the whole Shopify upload, for the KPI strip.
        'shopify_totals': {
            'transactions': sum(r['transactions'] for r in shopify_rows),
            'net': money.to_rands(upload_net_cents),
            'diff': money.to_rands(total_cents - upload_net_cents),
            'outside': money.to_rands(outside_cents),
            # Only locations the upload actually carries sit "outside" the
            # comparison; a stored name absent from this month contributes zero.
            'outside_count': len(unresolved) + sum(1 for r in excluded if not r.get('absent')),
        },
        # The journal is run for a CLOSED month; warn if viewing the current one.
        'is_current_month': (year == datetime.now().year and month == datetime.now().month),
    }


@app.route('/cash/cash-sales/<int:year>/<int:month>')
def cash_sales_preview(year, month):
    blocked = _require_admin()
    if blocked:
        return blocked
    if db.validate_month_year(year, month) is None:
        return redirect(url_for('cash_home'))
    prev_year, prev_month = (year - 1, 12) if month == 1 else (year, month - 1)
    next_year, next_month = (year + 1, 1) if month == 12 else (year, month + 1)
    ctx = _cash_sales_context(year, month)
    ctx.update(prev_year=prev_year, prev_month=prev_month,
               next_year=next_year, next_month=next_month)
    return render_template('cash_sales.html', **ctx)


@app.route('/cash/cash-sales/<int:year>/<int:month>/mj.csv')
def cash_sales_mj_export(year, month):
    """The consolidated cash-sales manual journal as a Xero-import CSV."""
    blocked = _require_admin()
    if blocked:
        return blocked
    if db.validate_month_year(year, month) is None:
        return redirect(url_for('cash_home'))
    store_lines = db.get_cash_sales_journal_stores(year, month)
    if not store_lines or not sum(r['sales_cents'] for r in store_lines):
        return _bounce('No cash sales captured for that month — nothing to export.',
                       'cash_sales_preview', 'warning', year=year, month=month)
    # Same fail-closed preflight, and the same wording, as the expense journals:
    # a store with no POS account (including an orphan whose store row is gone,
    # which arrives with store_code NULL) blocks the download rather than
    # exporting the #### placeholder. Only the POS account — this journal has no
    # tracking column at all, so an "Xero tracking option" message could
    # only ever be misdirection.
    issues = []
    for r in store_lines:
        issues.extend(_dear_account_issues(r['store'], r['store_code']))
    blocked = _block_expense_export(
        issues, redirect(url_for('cash_sales_preview', year=year, month=month)))
    if blocked:
        return blocked
    contra_code, narration_prefix = _cash_sales_settings()
    narration = f"{narration_prefix} - {db.MONTH_FULL[month]} {year}"
    date_out = date(year, month, calendar.monthrange(year, month)[1]).strftime('%d/%m/%Y')
    rows = cash_mj.build_cash_sales_rows(narration, date_out, contra_code, store_lines)
    csv_text = cash_mj.cash_sales_to_csv(rows)
    filename = f"Retail cash sales journal - {db.MONTH_FULL[month]} {year}.csv"
    return Response(csv_text, mimetype='text/csv',
                    headers={'Content-Disposition': f'attachment; filename="{filename}"'})


@app.route('/cash/cash-sales/<int:year>/<int:month>/shopify', methods=['POST'])
def cash_sales_shopify_upload(year, month):
    blocked = _require_admin()
    if blocked:
        return blocked
    if db.validate_month_year(year, month) is None:
        return redirect(url_for('cash_home'))
    upload = request.files.get('shopify_csv')
    # Judge the extension on the name the browser sent, and only then sanitise
    # what we store: secure_filename('Отчёт.csv') is 'csv' — the dot is stripped,
    # so a perfectly good upload was bounced as "not a CSV".
    original = (upload.filename or '') if upload else ''
    if not original.lower().endswith('.csv'):
        return _bounce('Choose the monthly Shopify CSV file.',
                       'cash_sales_preview', year=year, month=month)
    filename = secure_filename(original)
    if not filename.lower().endswith('.csv'):
        # Nothing recognisable survived sanitising. The stored name is rendered
        # on the page, so show something honest rather than 'csv' or ''.
        filename = f'shopify-{year:04d}-{month:02d}.csv'
    if db.get_cash_shopify_upload(year, month) and request.form.get('confirm_replace') != '1':
        return _bounce('This month already has a Shopify upload. Tick Replace and upload again.',
                       'cash_sales_preview', 'warning', year=year, month=month)
    # The page advertises a 10 MB limit but MAX_CONTENT_LENGTH is 150 MB, so the
    # parser's check only fired after the whole body was already in memory.
    # Measure the part before reading it, against the parser's own constant.
    limit = shopify_csv.MAX_UPLOAD_BYTES
    size = upload.content_length
    if not size:
        upload.stream.seek(0, io.SEEK_END)
        size = upload.stream.tell()
        upload.stream.seek(0)
    if size > limit:
        return _bounce(f'The Shopify CSV is larger than {limit // (1024 * 1024)} MB.',
                       'cash_sales_preview', year=year, month=month)
    try:
        rows, digest = shopify_csv.parse_shopify_cash_csv(upload.read())
        actor = session.get('admin_display_name') or session.get('admin') or 'admin'
        db.save_cash_shopify_upload(year, month, filename, digest, rows, actor)
    except ValueError as exc:
        return _bounce(str(exc), 'cash_sales_preview', year=year, month=month)
    return _bounce(f'Shopify cash sales loaded for {db.MONTH_FULL[month]} {year}: '
                   f'{len(rows)} rows.', 'cash_sales_preview', 'success',
                   year=year, month=month)


@app.route('/cash/cash-sales/<int:year>/<int:month>/mapping', methods=['POST'])
def cash_sales_shopify_mapping(year, month):
    blocked = _require_admin()
    if blocked:
        return blocked
    try:
        db.save_cash_shopify_mapping(
            request.form.get('shopify_location'), request.form.get('store'))
    except ValueError as exc:
        return _bounce(str(exc), 'cash_sales_preview', year=year, month=month)
    return _bounce('Shopify location mapped.', 'cash_sales_preview', 'success',
                   year=year, month=month)


@app.route('/cash/cash-sales/<int:year>/<int:month>/exclude', methods=['POST'])
def cash_sales_shopify_exclude(year, month):
    """Exclude a Shopify location from the comparison, or put one back.

    Some POS locations (the live export's "NORTHWIND Sample Sale") are not a store and
    never will be, so they can never be mapped — and an unresolved location
    blocks the recon workbook. Excluding one keeps it on the page with its value
    and inside the headline Shopify net; it only stops blocking.

    The exclusion applies to the month being viewed, and only to a location that
    month's upload actually contains — a typo used to be stored forever, shown
    nowhere, with no Restore button, and would silently remove a real store's
    Shopify side the day that name appeared."""
    blocked = _require_admin()
    if blocked:
        return blocked
    if db.validate_month_year(year, month) is None:
        return redirect(url_for('cash_home'))
    location = (request.form.get('shopify_location') or '').strip()
    if not location:
        return _bounce('Choose a Shopify location to exclude.',
                       'cash_sales_preview', year=year, month=month)
    restore = request.form.get('restore') == '1'
    if restore:
        # Restore works off what is stored, not off the upload: a stored name the
        # upload no longer carries is exactly the one that must stay undoable.
        known = {n.casefold(): n for n in _excluded_shopify_locations(year, month)}
        message = '“{}” is back in the {} {} comparison.'
        not_found = (f'“{location}” is not excluded from {db.MONTH_FULL[month]} {year}.')
    else:
        known = {r['pos_location_name'].casefold(): r['pos_location_name']
                 for r in db.get_cash_shopify_summary(year, month)}
        message = ('“{}” is excluded from the {} {} comparison — its value still '
                   'counts towards the Shopify net.')
        not_found = (f'“{location}” is not a location in the {db.MONTH_FULL[month]} '
                     f'{year} Shopify upload.')
    # Store the upload's own spelling so the excluded panel always renders it.
    canonical = known.get(location.casefold())
    if canonical is None:
        return _bounce(not_found, 'cash_sales_preview', 'warning', year=year, month=month)
    _save_excluded_shopify_location(year, month, canonical, restore)
    return _bounce(message.format(canonical, db.MONTH_FULL[month], year),
                   'cash_sales_preview', 'success', year=year, month=month)


@app.route('/cash/cash-sales/<int:year>/<int:month>/reasons', methods=['POST'])
def cash_sales_save_reasons(year, month):
    blocked = _require_admin()
    if blocked:
        return blocked
    if db.validate_month_year(year, month) is None:
        return redirect(url_for('cash_home'))
    # The whitelist is the month's comparison, NOT the stores table. An orphan
    # (an entry whose `stores` row is gone) is deliberately kept in the
    # comparison and needs a variance reason, but is not in get_stores() — its
    # reason was being discarded while the page flashed "saved", so
    # missing_reason_count never cleared and recon.xlsx stayed blocked forever
    # with no way out of the UI. Matched case/whitespace-insensitively and
    # written back under the comparison's own spelling.
    comparison = _cash_sales_context(year, month)['comparison']
    valid_stores = {r['store'].strip().casefold(): r['store'] for r in comparison}
    reasons = {}
    for field, posted in request.form.items():
        # Driven by the posted rows themselves rather than by `row_count`, so a
        # stale count can neither truncate a real reason nor spin the loop.
        if not field.startswith('store_'):
            continue
        store = valid_stores.get(posted.strip().casefold())
        if store is not None:
            reasons[store] = request.form.get('reason_' + field[len('store_'):], '')
    actor = session.get('admin_display_name') or session.get('admin') or 'admin'
    db.save_cash_sales_variance_reasons(year, month, reasons, actor)
    return _bounce('Variance reasons saved.', 'cash_sales_preview', 'success',
                   year=year, month=month)


@app.route('/cash/cash-sales/<int:year>/<int:month>/recon.xlsx')
def cash_sales_recon_export(year, month):
    """Completed Shopify-vs-cash-sheet comparison for the selected month."""
    blocked = _require_admin()
    if blocked:
        return blocked
    if db.validate_month_year(year, month) is None:
        return redirect(url_for('cash_home'))
    ctx = _cash_sales_context(year, month)
    if not ctx['shopify_upload']:
        return _bounce('Upload the monthly Shopify CSV before downloading the comparison.',
                       'cash_sales_preview', 'warning', year=year, month=month)
    if ctx['unmatched_shopify']:
        return _bounce('Map or exclude every Shopify location before downloading '
                       'the comparison.', 'cash_sales_preview', year=year, month=month)
    if ctx['missing_reason_count']:
        return _bounce('Add a reason for every difference greater than R1.00 before '
                       'downloading.', 'cash_sales_preview', year=year, month=month)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Cash sales {db.MONTH_FULL[month][:3]} {year}"
    hdr_fill = PatternFill('solid', start_color='1F2937')
    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    total_fill = PatternFill('solid', start_color='E5E7EB')
    money_fmt = 'R#,##0.00;-R#,##0.00;"-"'
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ['POS location name', 'POS location name', 'Payment gateway',
               'Transactions', 'Gross payments', 'Refunded payments', 'Net payments',
               '', 'Cash Sheet', '', 'Diff', '', 'Reason']
    ws.append(headers)
    for c in ws[1]:
        c.fill, c.font = hdr_fill, hdr_font
        c.alignment = Alignment(horizontal='center')
        c.border = border

    first = 2
    for i, r in enumerate(ctx['comparison']):
        row = first + i
        values = [r['shopify_locations'], r['display_store'], 'cash', r['transactions'],
                  money.to_rands(r['gross_cents']), money.to_rands(r['refunded_cents']),
                  money.to_rands(r['net_cents']), None, money.to_rands(r['cash_cents']),
                  None, f'=I{row}-G{row}', None, r['reason']]
        for col, value in enumerate(values, start=1):
            ws.cell(row, col, value).border = border
        for col in (5, 6, 7, 9, 11):
            ws.cell(row, col).number_format = money_fmt
    last = first + len(ctx['comparison']) - 1

    def _total_row(row, label, sums):
        """One bold, shaded totals row: {column -> formula}."""
        ws.cell(row, 1, label).font = Font(bold=True)
        for col, formula in sums.items():
            cell = ws.cell(row, col, formula)
            if col != 4:
                cell.number_format = money_fmt
            cell.font = Font(bold=True)
        for col in range(1, 14):
            ws.cell(row, col).fill = total_fill
            ws.cell(row, col).border = border

    # Excluded locations are money the page counts in its headline Shopify net,
    # so the workbook has to carry them too — a sheet that quietly omits R4,321
    # while the screen shows it is the artefact finance keeps, and the honest
    # one is the page. They go BELOW the compared rows, in their own labelled
    # block, so the reference workbook's column order is untouched above.
    excluded = ctx['excluded_shopify']
    compared_total = last + 1 if ctx['comparison'] else 0
    if ctx['comparison']:
        _total_row(compared_total,
                   'TOTAL (compared stores)' if excluded else 'TOTAL',
                   {4: f"=SUM(D{first}:D{last})", 5: f"=SUM(E{first}:E{last})",
                    6: f"=SUM(F{first}:F{last})", 7: f"=SUM(G{first}:G{last})",
                    9: f"=SUM(I{first}:I{last})", 11: f"=SUM(K{first}:K{last})"})
    else:
        ws.append(['No cash sales captured for this month.'])

    if excluded:
        note = ws.cell(ws.max_row + 2, 1,
                       'Excluded from the store comparison — not a Northwind store')
        note.font = Font(bold=True, italic=True)
        xfirst = ws.max_row + 1
        for i, r in enumerate(excluded):
            row = xfirst + i
            values = [r['pos_location_name'], '(not a Northwind store)', 'cash',
                      r['transactions'], money.to_rands(r['gross_cents']),
                      money.to_rands(r['refunded_cents']), money.to_rands(r['net_cents']),
                      None, None, None, None, None,
                      'Not in this month’s Shopify upload' if r.get('absent')
                      else 'Excluded from the store comparison']
            for col, value in enumerate(values, start=1):
                ws.cell(row, col, value).border = border
            for col in (5, 6, 7):
                ws.cell(row, col).number_format = money_fmt
        xlast = xfirst + len(excluded) - 1
        excluded_total = xlast + 1
        _total_row(excluded_total, 'TOTAL (excluded locations)',
                   {4: f"=SUM(D{xfirst}:D{xlast})", 5: f"=SUM(E{xfirst}:E{xlast})",
                    6: f"=SUM(F{xfirst}:F{xlast})", 7: f"=SUM(G{xfirst}:G{xlast})"})
        # The grand row adds the two subtotal CELLS (never re-SUM a range that
        # holds subtotals — it double-counts), so it ties to the page's headline
        # "Shopify net" and, in column K, to its "Difference" KPI.
        parts = [str(excluded_total)] + ([str(compared_total)] if compared_total else [])
        grand = excluded_total + 1
        _total_row(grand, 'TOTAL (all Shopify cash)',
                   {col: '=' + '+'.join(f'{letter}{p}' for p in parts)
                    for col, letter in ((4, 'D'), (5, 'E'), (6, 'F'), (7, 'G'))})
        if compared_total:
            cash = ws.cell(grand, 9, f'=I{compared_total}')
            cash.number_format, cash.font = money_fmt, Font(bold=True)
            diff = ws.cell(grand, 11, f'=I{grand}-G{grand}')
            diff.number_format, diff.font = money_fmt, Font(bold=True)

    for col, width in zip('ABCDEFGHIJKLM',
                          (25, 20, 13, 12, 16, 17, 16, 3, 16, 3, 14, 3, 58)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = 'A2'

    # Second sheet — every individual cash sale (with its receipt number),
    # GROUPED per store: a store header row carrying that store's subtotal, then
    # its sales below it, then the next store, then a grand total. Subtotals +
    # grand total are literal (never =SUM() a column that also holds subtotal
    # rows — that double-counts; same rule as the monthly exports).
    items = db.get_cash_sales_line_items(year, month)
    groups = {}                       # store -> [items]; items are already store-ordered
    for it in items:
        groups.setdefault(it['store'], []).append(it)

    store_fill = PatternFill('solid', start_color='EEF2F7')
    store_font = Font(bold=True, size=11, color='1F2937')

    ws2 = wb.create_sheet(f"All sales {db.MONTH_FULL[month][:3]} {year}")
    ws2.append(['Store', 'Date', 'Sale / receipt no.', 'Amount'])
    for c in ws2[1]:
        c.fill, c.font = hdr_fill, hdr_font
        c.alignment = Alignment(horizontal='center')
        c.border = border

    r = 2
    grand_cents = 0
    for store, lst in groups.items():
        sub_cents = sum(it['amount_cents'] for it in lst)
        grand_cents += sub_cents
        # Store header row: name (col A) + that store's subtotal (col D).
        ws2.cell(r, 1, store).font = store_font
        subtotal = ws2.cell(r, 4, money.to_rands(sub_cents))
        subtotal.number_format = money_fmt
        subtotal.font = store_font
        for col in range(1, 5):
            cell = ws2.cell(r, col)
            cell.fill = store_fill
            cell.border = border
        r += 1
        # Its individual sales below (date / receipt no. / amount).
        for it in lst:
            ws2.cell(r, 2, it['date']).border = border
            ws2.cell(r, 3, (it['sale_no'] or '').strip()).border = border
            amt = ws2.cell(r, 4, money.to_rands(it['amount_cents']))
            amt.number_format = money_fmt
            amt.border = border
            ws2.cell(r, 1).border = border
            r += 1

    if groups:
        tc = ws2.cell(r, 1, 'GRAND TOTAL')
        tc.font = Font(bold=True)
        gt = ws2.cell(r, 4, money.to_rands(grand_cents))
        gt.number_format = money_fmt
        gt.font = Font(bold=True)
        for col in range(1, 5):
            ws2.cell(r, col).fill = total_fill
            ws2.cell(r, col).border = border
    else:
        ws2.append(['No cash sales captured for this month.'])
    for col, width in zip('ABCD', (26, 14, 26, 16)):
        ws2.column_dimensions[col].width = width
    ws2.freeze_panes = 'A2'

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(
        out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'Cash sales reconciliation - {db.MONTH_FULL[month]} {year}.xlsx')


@app.route('/cash/xero-setup/cash-sales', methods=['POST'])
def cash_save_cash_sales_settings():
    blocked = _require_admin()
    if blocked:
        return blocked
    db.set_setting(CASH_SALES_CONTRA_KEY, request.form.get('contra_code', '').strip())
    db.set_setting(CASH_SALES_NARRATION_KEY, request.form.get('narration', '').strip())
    return _bounce('Cash-sales journal settings saved.', 'cash_xero_setup', 'success')


# ── Reporting: overview grid + category×store matrix + batch expense MJs ───────

_MONEY_FMT = 'R#,##0.00;-R#,##0.00;"-"'


def _xlsx_header(ws, headers, border):
    fill = PatternFill('solid', start_color='1F2937')
    font = Font(bold=True, color='FFFFFF', size=11)
    ws.append(headers)
    for c in ws[1]:
        c.fill, c.font = fill, font
        c.alignment = Alignment(horizontal='center')
        c.border = border


def _xlsx_send(wb, download_name):
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(
        out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=download_name)


@app.route('/cash/overview.xlsx')
def cash_overview_export():
    """The all-stores overview grid (opening/in/expenses/banked/adjust/closing +
    entry count) for the current date range, as an Excel with a totals row."""
    blocked = _require_admin()
    if blocked:
        return blocked
    start_iso, end_iso, s, e = _range_from_request()
    rows = db.get_recon_overview_range(start_iso, end_iso)
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Cash overview'
    cols = [('store', 'Store', False), ('opening', 'Opening', True),
            ('total_in', 'In', True), ('total_expense', 'Expenses', True),
            ('total_banked', 'Banked', True), ('total_adjust', 'Adjust', True),
            ('closing', 'Closing', True), ('entry_count', '# entries', False)]
    _xlsx_header(ws, [c[1] for c in cols], border)
    tot = {k: 0 for k, _l, money_col in cols if money_col}
    ecount = 0
    for r in rows:
        vals = []
        for key, _label, money_col in cols:
            vals.append(r[key])
            if money_col:
                tot[key] += r[key]
        ecount += r['entry_count']
        ws.append(vals)
        for i, (key, _l, money_col) in enumerate(cols, start=1):
            cell = ws.cell(ws.max_row, i)
            cell.border = border
            if money_col:
                cell.number_format = _MONEY_FMT
    trow = ws.max_row + 1
    ws.cell(trow, 1, 'All stores').font = Font(bold=True)
    for i, (key, _l, money_col) in enumerate(cols, start=1):
        if money_col:
            c = ws.cell(trow, i, round(tot[key], 2))
            c.number_format = _MONEY_FMT
            c.font = Font(bold=True)
    ws.cell(trow, len(cols), ecount).font = Font(bold=True)
    for col, w in zip('ABCDEFGH', (26, 14, 14, 14, 14, 12, 14, 11)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A2'
    return _xlsx_send(wb, f'Cash overview {start_iso} to {end_iso}.xlsx')


@app.route('/cash/reports/category-matrix.xlsx')
def cash_category_matrix_export():
    """Expense spend as a category (rows) × store (columns) matrix for the range,
    with row + column totals — the classic month-end 'who spent what on what'."""
    blocked = _require_admin()
    if blocked:
        return blocked
    start_iso, end_iso, s, e = _range_from_request()
    stores = db.get_stores()
    cats = db.get_recon_category_store_breakdown(stores, start_iso, end_iso)
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Category x store'
    _xlsx_header(ws, ['Category', 'Xero code'] + stores + ['Total'], border)
    col_tot = {st: 0.0 for st in stores}
    grand = 0.0
    for cat in cats:
        by_store = {sv['store']: sv['total'] for sv in cat['stores']}
        row = [cat['name'], cat.get('xero_code') or ''] + [by_store.get(st, 0) for st in stores] + [cat['total']]
        ws.append(row)
        for st in stores:
            col_tot[st] += by_store.get(st, 0)
        grand += cat['total']
        for i in range(3, len(row) + 1):           # money columns start at col 3
            cell = ws.cell(ws.max_row, i)
            cell.number_format = _MONEY_FMT
            cell.border = border
        ws.cell(ws.max_row, 1).border = border
        ws.cell(ws.max_row, 2).border = border
    trow = ws.max_row + 1
    ws.cell(trow, 1, 'TOTAL').font = Font(bold=True)
    for j, st in enumerate(stores, start=3):
        c = ws.cell(trow, j, round(col_tot[st], 2))
        c.number_format = _MONEY_FMT
        c.font = Font(bold=True)
    gc = ws.cell(trow, len(stores) + 3, round(grand, 2))
    gc.number_format = _MONEY_FMT
    gc.font = Font(bold=True)
    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 11
    for idx in range(len(stores) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(3 + idx)].width = 15
    ws.freeze_panes = 'C2'
    return _xlsx_send(wb, f'Cash expenses by category x store {start_iso} to {end_iso}.xlsx')


@app.route('/cash/<int:year>/<int:month>/mj/all.zip')
def cash_mj_all_export(year, month):
    """Every store's balanced expenses manual journal for the month, zipped — one
    CSV per store (skips stores with no expenses). Built straight off the ledger
    (no per-line edits); mirrors the single-store cash_mj_export."""
    blocked = _require_admin()
    if blocked:
        return blocked
    if db.validate_month_year(year, month) is None:
        return redirect(url_for('cash_home'))
    month_end = date(year, month, calendar.monthrange(year, month)[1]).strftime('%d/%m/%Y')
    mem = io.BytesIO()
    count = 0
    journals = []
    issues = []
    # NOT get_stores(): a demo store's expenses would put seeded money in the ZIP,
    # and (with no POS code by design) its preflight issue blocks every OTHER
    # store's journal too — the whole batch download fails closed on fake data.
    for store in db.get_xero_export_stores():
        raw = db.get_recon_expense_lines(store, year, month)
        if not raw:
            continue
        xero = db.get_store_xero(store)
        lines = [{'category': r['category'],
                  'description': cash_mj.build_description(r['category'], r['date'], r['note']),
                  'xero_code': r['xero_code'], 'vat_type': r['vat_type'] or cash_mj.DEFAULT_VAT,
                  'gross_cents': r['gross_cents']} for r in raw]
        issues.extend(_expense_export_issues(store, xero, lines))
        journals.append((store, xero, lines))

    blocked = _block_expense_export(issues, redirect(url_for('cash_home')))
    if blocked:
        return blocked

    with zipfile.ZipFile(mem, 'w', zipfile.ZIP_DEFLATED) as z:
        for store, xero, lines in journals:
            rows = cash_mj.build_balanced_rows(
                f"{store} {db.MONTH_FULL[month]} {year}", month_end,
                xero['tracking_name'], store, xero['store_code'], lines)
            z.writestr(f"{store.replace(chr(34), '')} {db.MONTH_FULL[month]} {year}.csv",
                       cash_mj.to_csv(rows))
            count += 1
    if not count:
        return _bounce('No store expenses captured for that month — nothing to export.',
                       'cash_home', 'warning')
    mem.seek(0)
    return send_file(mem, mimetype='application/zip', as_attachment=True,
                     download_name=f'Expense MJs - {db.MONTH_FULL[month]} {year}.zip')
