"""Parse a Xero credit-card bank-reconciliation .xlsx export.

This module is intentionally free of Flask/DB dependencies so it can be unit
tested in isolation against representative workbook fixtures. It reads ONE workbook (one
cardholder) and returns a structured, classified snapshot of that card's
**unreconciled** statement lines.

The representative export format has three
sheets:

  1. "<Name> Credit Card Reconci..." — Summary. Row 4 carries the card name
     ("Primary Credit Card"); a "Plus Unreconciled Statement Lines" section
     lists the unreconciled lines.
  2. "Bank Statement" — every statement line with an explicit Reconciled
     (Yes/No) column. This is our PRIMARY source for "what is unreconciled",
     because the flag is unambiguous.
  3. "Statement Exceptions" — rows Xero itself deleted as duplicates
     ("Deleted (Duplicate)"). We surface these for audit only; the
     unreconciled list is already deduped by Xero.

Money is returned in **integer cents** to match the rest of the app (money is
stored in cents everywhere). Rands appear only at the edges.

Classification of each unreconciled line:
  - 'spend'    — a negative merchant charge → the cardholder owes a receipt.
  - 'transfer' — money in (positive) / a card funding or repayment → hidden
                 from the cardholder; kept for the admin reconciliation only.
  - 'fee'      — a bank/system line ("#DECLINED AUTH FEE", "#CARD_SERVICE_FEE")
                 → N/A, no receipt is possible.
"""

from __future__ import annotations

import datetime as _dt
import re
from dataclasses import dataclass, field, asdict

import openpyxl


# ── Classification ───────────────────────────────────────────────────────────

# Bank/system lines begin with '#' in the Reference (e.g. "#DECLINED AUTH FEE",
# "#CARD_SERVICE_FEE"). They can never have a receipt.
_FEE_PREFIX = "#"

# Transfers / money-in. On a credit card a positive amount reduces what is owed
# — a payment, funding transfer or refund. Cardholders never need to see these.
# We classify by SIGN (non-negative => transfer), which is the safe, robust
# rule: a genuine expense is always strictly negative, so it can never be
# mistaken for a transfer, and a zero-amount marker (R0.00 authorisation /
# reversal) is NOT treated as a spend that demands an impossible receipt.

# Internal money movements that carry a card's balance between statement periods
# or fund the card. These are NOT merchant spend even when they land as a
# negative amount (a balance carried out is negative), so match them by name and
# treat them as transfers regardless of sign — they must never live on the app.
_TRANSFER_NAME_RE = re.compile(
    r"\bbalance\s+(transferred|brought\s+forward|carried\s+forward)\b", re.I)


def classify(reference: str, amount_cents: int) -> str:
    """Return 'spend' | 'transfer' | 'fee' for one statement line."""
    ref = (reference or "").strip()
    if ref.startswith(_FEE_PREFIX):
        return "fee"
    if _TRANSFER_NAME_RE.search(ref):
        return "transfer"
    if amount_cents >= 0:
        return "transfer"
    return "spend"


def needs_receipt(category: str) -> bool:
    return category == "spend"


# ── Data shapes ──────────────────────────────────────────────────────────────

@dataclass
class StatementLine:
    line_date: _dt.date | None
    reference: str            # the merchant / description string (trimmed)
    amount_cents: int         # signed; negative = spend, positive = money in
    category: str             # 'spend' | 'transfer' | 'fee'
    reconciled: bool          # from the Bank Statement "Reconciled" column
    # Stable identity for idempotent re-import. Genuine repeats (e.g. two
    # identical -R6 fees on the same day) share a fingerprint and are
    # disambiguated by `occurrence` (0, 1, 2, ...) so we keep the right count.
    fingerprint: str = ""
    occurrence: int = 0

    @property
    def needs_receipt(self) -> bool:
        return needs_receipt(self.category) and not self.reconciled


@dataclass
class CardSnapshot:
    card_name: str                       # e.g. "Primary Credit Card" (identity)
    display_name: str                    # e.g. "Terrence"
    period_start: _dt.date | None
    period_end: _dt.date | None
    as_at: _dt.date | None
    statement_balance_cents: int | None
    lines: list[StatementLine] = field(default_factory=list)
    duplicates_removed_by_xero: int = 0  # count from Statement Exceptions sheet
    source_filename: str | None = None

    # convenience rollups
    @property
    def unreconciled_lines(self) -> list[StatementLine]:
        return [l for l in self.lines if not l.reconciled]

    @property
    def receipts_required(self) -> list[StatementLine]:
        return [l for l in self.lines if l.needs_receipt]


# ── Helpers ──────────────────────────────────────────────────────────────────

def _to_cents(value) -> int:
    """Rands (float/int) -> integer cents, rounded to avoid float drift."""
    if value is None:
        return 0
    return int(round(float(value) * 100))


def _as_date(value):
    if isinstance(value, _dt.datetime):
        return value.date()
    if isinstance(value, _dt.date):
        return value
    return None


def _clean(text) -> str:
    """Trim and collapse internal whitespace (Xero pads references heavily)."""
    if text is None:
        return ""
    return re.sub(r"\s+", " ", str(text)).strip()


def _find_summary_sheet(wb):
    for ws in wb.worksheets:
        title = (ws.title or "").lower()
        if "reconcili" in title or "credit card" in title:
            return ws
    return wb.worksheets[0]


def _derive_card_name(summary_ws) -> str:
    """Card name lives at A4 of the summary sheet (e.g. 'Primary Credit Card').

    Fall back to the sheet title with the trailing 'Reconciliation Summary…'
    stripped if A4 is empty for some reason.
    """
    a4 = _clean(summary_ws["A4"].value)
    if a4:
        return a4
    title = _clean(summary_ws.title)
    title = re.sub(r"\bReconcil\w*.*$", "", title, flags=re.I).strip()
    return title or "Unknown Card"


def _display_name(card_name: str) -> str:
    """'Primary Credit Card' -> 'Primary'; 'Operations Credit Card' -> 'Operations'."""
    name = re.sub(r"\bcredit card\b", "", card_name, flags=re.I)
    # Trim leftover separators/spaces (e.g. a dangling ' - ' before 'Credit Card').
    name = name.strip().strip("-–—").strip()
    return name or card_name


_PERIOD_RE = re.compile(
    r"period\s+(\d{1,2}\s+\w+\s+\d{4})\s+to\s+(\d{1,2}\s+\w+\s+\d{4})", re.I
)
_AS_AT_RE = re.compile(r"as at\s+(\d{1,2}\s+\w+\s+\d{4})", re.I)


def _parse_text_date(text: str):
    try:
        return _dt.datetime.strptime(text.strip(), "%d %B %Y").date()
    except (ValueError, AttributeError):
        return None


def _scan_meta(wb):
    """Pull period start/end and 'as at' date from the header text rows."""
    period_start = period_end = as_at = None
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            for cell in row:
                if not isinstance(cell, str):
                    continue
                if (m := _PERIOD_RE.search(cell)):
                    period_start = period_start or _parse_text_date(m.group(1))
                    period_end = period_end or _parse_text_date(m.group(2))
                if (m := _AS_AT_RE.search(cell)):
                    as_at = as_at or _parse_text_date(m.group(1))
    return period_start, period_end, as_at


# ── Bank Statement sheet ─────────────────────────────────────────────────────

def _header_index(ws):
    """Locate the header row and map column names -> 0-based index."""
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=12, values_only=True), 1):
        cells = [_clean(c).lower() for c in row]
        if "date" in cells and "amount" in cells:
            return r, {name: i for i, name in enumerate(cells) if name}
    return None, {}


def _parse_bank_statement(ws):
    """Yield (date, reference, amount_cents, reconciled_bool, import_batch) per line.

    `import_batch` is the raw "Date imported into Xero" value (or None). It lets
    the caller tell a genuine same-day duplicate charge (both in one import) from
    a Xero RE-IMPORT of the same line (the identical charge appearing under two
    different import dates — a common Xero artifact that must not be counted
    twice)."""
    header_row, idx = _header_index(ws)
    if header_row is None:
        return
    c_date = idx.get("date")
    c_ref = idx.get("reference")
    c_desc = idx.get("description")
    c_amt = idx.get("amount")
    c_rec = idx.get("reconciled")
    c_imp = idx.get("date imported into xero")
    # Without a Reconciled column we cannot tell settled lines from open ones,
    # and would silently mark EVERY historical line as outstanding (demanding
    # receipts for months of already-cleared spend). Fail loudly instead — the
    # upload handler surfaces this so the admin can check the export format.
    if c_rec is None:
        raise ValueError(
            "Bank Statement sheet has no 'Reconciled' column — cannot tell which "
            "lines are already reconciled. Is this the expected Xero credit-card "
            "reconciliation export?")
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        # The merchant string lives in EITHER the Reference or the Description
        # column depending on when Xero imported the line: older imports carry
        # it in Reference (Description empty), newer imports carry it in
        # Description (Reference blank whitespace). Prefer Reference, fall back
        # to Description — otherwise every "new-format" line is silently dropped
        # by the empty-reference guard below.
        ref = _clean(row[c_ref]) if c_ref is not None else ""
        if not ref and c_desc is not None:
            ref = _clean(row[c_desc])
        amt = row[c_amt] if c_amt is not None else None
        date = _as_date(row[c_date]) if c_date is not None else None
        # Skip section headers ("Statement Lines", "Opening Balance"...) and
        # zero-amount balance markers with no reference.
        if not ref and (amt in (None, 0)):
            continue
        if not ref:
            continue
        reconciled = False
        if c_rec is not None:
            reconciled = _clean(row[c_rec]).lower() == "yes"
        batch = _clean(row[c_imp]) if c_imp is not None else ""
        yield date, ref, _to_cents(amt), reconciled, batch


def _parse_summary_unreconciled(ws):
    """Yield (date, reference, amount_cents, reconciled=False) from the summary
    sheet's 'Plus Unreconciled Statement Lines' detail section.

    Some Xero recon exports leave the Bank Statement sheet with only opening/
    closing balance and itemise the chargeable lines here instead. Every line in
    this section is unreconciled by definition, so reconciled is always False.
    """
    in_section = False
    for row in ws.iter_rows(values_only=True):
        cells = list(row)
        first = _clean(cells[0]) if cells else ""
        low = first.lower()
        if not in_section:
            # The DETAIL header stands alone in the first column — distinct from
            # the one-line 'Plus unreconciled statement lines' TOTAL in the
            # Totals Summary block, whose label sits in a later column.
            if low == "plus unreconciled statement lines":
                in_section = True
            continue
        if low.startswith("total unreconciled statement lines"):
            break
        # A data row needs a date and a numeric amount; the reference is the
        # longest text cell that is neither the date nor the amount.
        date = None
        for c in cells:
            date = _as_date(c)
            if date:
                break
        amount = None
        for c in cells:
            if isinstance(c, (int, float)) and not isinstance(c, bool):
                amount = c
        ref = ""
        for c in cells:
            if isinstance(c, (int, float)) or _as_date(c) is not None:
                continue
            t = _clean(c)
            if len(t) > len(ref):
                ref = t
        if date is None or amount is None or not ref:
            continue
        yield date, ref, _to_cents(amount), False


def _count_exceptions(wb) -> int:
    for ws in wb.worksheets:
        if "exception" in (ws.title or "").lower():
            n = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(c is not None for c in row):
                    n += 1
            return n
    return 0


# ── Merge the two sheets ───────────────────────────────────────────────────

def _row_fp(date, ref, cents) -> str:
    iso = date.isoformat() if date else "?"
    return f"{iso}|{(ref or '').lower()}|{cents}"


def _union_rows(bank_rows, summ_rows):
    """Union Bank Statement rows with Summary unreconciled rows into the real,
    de-duplicated set of statement lines.

    Two things create phantom duplicates that must be collapsed:
      1. The same line appears on BOTH the Bank Statement and Summary sheets.
      2. Xero RE-IMPORTS a statement (e.g. later in a new column layout), so the
         identical charge is listed once per import — under different "Date
         imported into Xero" batches. That is one transaction, not several.

    So for the Bank Statement we count a fingerprint's occurrences as the MAX
    within any single import batch (a genuine same-day double charge shares one
    batch and is kept; a re-import repeats across batches and is collapsed), and
    take the Reconciled flags from the most recent such batch. The Summary
    (which has no batch column and simply relists both import layouts) only ever
    contributes fingerprints the Bank Statement never showed — carried-over
    lines from earlier months — at one occurrence each.

    Returns a flat list of (date, ref, amount_cents, reconciled) in stable order.
    """
    # Bank Statement: fingerprint -> {meta, batches: {batch: [reconciled, ...]}}
    bank: dict[str, dict] = {}
    order: list[str] = []
    for date, ref, cents, rec, batch in bank_rows:
        fp = _row_fp(date, ref, cents)
        e = bank.get(fp)
        if e is None:
            e = bank[fp] = {"date": date, "ref": ref, "cents": cents, "batches": {}}
            order.append(fp)
        e["batches"].setdefault(batch, []).append(bool(rec))

    merged: dict[str, dict] = {}
    for fp in order:
        e = bank[fp]
        # Occurrences = most a single import batch listed this line; flags from
        # the last batch achieving that count (the most recent import = truth).
        best_recs: list[bool] = []
        for recs in e["batches"].values():
            if len(recs) >= len(best_recs):
                best_recs = recs
        merged[fp] = {"date": e["date"], "ref": e["ref"], "cents": e["cents"],
                      "recs": list(best_recs)}

    for date, ref, cents, rec in summ_rows:
        fp = _row_fp(date, ref, cents)
        if fp in merged:
            continue  # already on the Bank Statement (authoritative) — or a
                      # Summary re-list of a line we've counted; skip either way.
        merged[fp] = {"date": date, "ref": ref, "cents": cents, "recs": [False]}
        order.append(fp)

    rows = []
    for fp in order:
        m = merged[fp]
        for rec in m["recs"]:
            rows.append((m["date"], m["ref"], m["cents"], rec))
    return rows


# ── Public entry point ───────────────────────────────────────────────────────

def parse_workbook(path_or_stream, source_filename: str | None = None) -> CardSnapshot:
    """Parse one Xero credit-card recon workbook into a CardSnapshot."""
    wb = openpyxl.load_workbook(path_or_stream, data_only=True, read_only=True)
    try:
        summary = _find_summary_sheet(wb)
        card_name = _derive_card_name(summary)
        period_start, period_end, as_at = _scan_meta(wb)

        bank_ws = None
        for ws in wb.worksheets:
            if "bank statement" in (ws.title or "").lower():
                bank_ws = ws
                break
        if bank_ws is None:
            bank_ws = summary

        # Two sources, UNIONED — never one-or-the-other. The Bank Statement
        # sheet carries the explicit Reconciled flag but, in an "as at"
        # reconciliation export, lists only the current period's lines; the
        # Summary's "Plus Unreconciled Statement Lines" section lists EVERY line
        # still outstanding (often carried over from earlier months). Reading
        # only the Bank Statement sheet silently drops those carried-over lines,
        # so we merge both and let no outstanding line fall through.
        try:
            bank_rows = list(_parse_bank_statement(bank_ws))
        except ValueError:
            bank_rows = []
        summ_rows = list(_parse_summary_unreconciled(summary))
        rows = _union_rows(bank_rows, summ_rows)

        lines: list[StatementLine] = []
        fp_counts: dict[str, int] = {}
        for date, ref, amount_cents, reconciled in rows:
            category = classify(ref, amount_cents)
            iso = date.isoformat() if date else "?"
            fp = f"{iso}|{ref.lower()}|{amount_cents}"
            occurrence = fp_counts.get(fp, 0)
            fp_counts[fp] = occurrence + 1
            lines.append(StatementLine(
                line_date=date,
                reference=ref,
                amount_cents=amount_cents,
                category=category,
                reconciled=reconciled,
                fingerprint=fp,
                occurrence=occurrence,
            ))

        return CardSnapshot(
            card_name=card_name,
            display_name=_display_name(card_name),
            period_start=period_start,
            period_end=period_end,
            as_at=as_at,
            statement_balance_cents=None,
            lines=lines,
            duplicates_removed_by_xero=_count_exceptions(wb),
            source_filename=source_filename,
        )
    finally:
        wb.close()
