"""Finance-friendly exports for the cross-card reconciliation queue.

These are reconciliation aids, not Xero import files.  Every spreadsheet row is
one ``cc_lines`` transaction and every receipt in a ZIP is copied into that
transaction's own folder so the evidence trail stays explicit.
"""
from collections import Counter
import datetime as dt
import io
import zipfile

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename

from northwind.data import database as db


_XLSX_MIME = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
_MONEY_FORMAT = 'R #,##0.00;[Red]-R #,##0.00'
_HEADER_FILL = PatternFill('solid', fgColor='242925')
_READY_FILL = PatternFill('solid', fgColor='E7F2EB')
_WAIT_FILL = PatternFill('solid', fgColor='FFF4DD')
_DONE_FILL = PatternFill('solid', fgColor='E8EEF0')
_THIN_BORDER = Border(bottom=Side(style='hair', color='D9DED9'))


def _safe(value):
    return db.xl_safe(value)


def _workflow_status(row):
    if row.get('xero_reconciled'):
        return 'Reconciled'
    if row.get('vat_invoice_required'):
        return 'VAT invoice requested'
    if row.get('ready_for_xero'):
        return 'Ready for Xero'
    if row.get('has_pending_suggestion'):
        return 'AI decision'
    return 'Needs cardholder'


def _evidence(row):
    if row.get('personal'):
        return 'Personal — employee to repay'
    count = int(row.get('receipt_count') or 0)
    if count:
        return f"{count} receipt{'s' if count != 1 else ''}"
    return 'Missing receipt'


def _account(row):
    if row.get('xero_account_code'):
        return (
            row.get('xero_account_code') or '',
            row.get('xero_account_name') or '',
            'Confirmed in NORTHWIND',
        )
    if row.get('ai_account_code'):
        return (
            row.get('ai_account_code') or '',
            row.get('ai_account_name') or '',
            'AI suggestion — check in Xero',
        )
    return '', '', 'Not supplied'


def _receipt_name(receipt):
    raw = (
        receipt.get('download_name')
        or receipt.get('original_filename')
        or receipt.get('file_path')
        or 'receipt'
    )
    return secure_filename(str(raw).rsplit('/', 1)[-1]) or 'receipt'


def _transaction_folder(row):
    month = f"{int(row['year']):04d}-{int(row['month']):02d}"
    card = secure_filename(
        str(row.get('display_name') or row.get('card_name') or 'card')
    )[:35] or 'card'
    merchant = secure_filename(str(row.get('reference') or 'transaction'))[:50]
    cents = abs(int(row.get('amount_cents') or 0))
    amount = f"R{cents // 100}-{cents % 100:02d}"
    return f"receipts/{month}/{int(row['id'])}_{card}_{merchant}_{amount}"


def _group_receipts(receipts):
    grouped = {}
    for receipt in receipts:
        grouped.setdefault(int(receipt['line_id']), []).append(dict(receipt))
    return grouped


def _scope_lines(scope):
    cards = scope.get('card_names') or []
    return [
        ('Cards', ', '.join(cards) if cards else 'No cards selected'),
        ('Statement month', scope.get('period') or 'All months'),
        ('Status', scope.get('status_label') or 'All unreconciled'),
        ('Search', scope.get('search') or 'None'),
    ]


def build_workbook(rows, receipts, scope, exported_at=None, receipt_issues=None):
    """Return an XLSX byte stream for already-scoped transaction rows."""
    exported_at = exported_at or dt.datetime.now(dt.timezone.utc)
    grouped = _group_receipts(receipts)
    issues = receipt_issues or {}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Transactions'
    headers = [
        'Transaction ID',
        'Statement month',
        'Transaction date',
        'Cardholder',
        'Xero card account',
        'Merchant / reference',
        'Statement amount (R)',
        'Workflow status',
        'Evidence',
        'Reason',
        'Location',
        'Submitted at',
        'Submitted by',
        'Personal',
        'VAT invoice requested',
        'VAT invoice requested at',
        'VAT invoice requested by',
        'Xero account code',
        'Xero account name',
        'Account source',
        'Reconciled in Xero',
        'Reconciled at',
        'Receipt count',
        'Receipt filenames',
        'Receipt folder in finance pack',
        'Receipt export note',
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 32

    for row in rows:
        row = dict(row)
        line_receipts = grouped.get(int(row['id']), [])
        code, account_name, account_source = _account(row)
        status = _workflow_status(row)
        receipt_names = ', '.join(_receipt_name(item) for item in line_receipts)
        issue_text = '; '.join(issues.get(int(row['id']), []))
        values = [
            int(row['id']),
            f"{int(row['year']):04d}-{int(row['month']):02d}",
            row.get('line_date') or '',
            row.get('display_name') or row.get('card_name') or '',
            row.get('card_name') or '',
            row.get('reference') or '',
            int(row.get('amount_cents') or 0) / 100,
            status,
            _evidence(row),
            row.get('reason') or '',
            row.get('location') or '',
            row.get('submitted_at') or '',
            row.get('submitted_by') or '',
            'Yes' if row.get('personal') else 'No',
            'Yes' if row.get('vat_invoice_required') else 'No',
            row.get('vat_invoice_requested_at') or '',
            row.get('vat_invoice_requested_by') or '',
            code,
            account_name,
            account_source,
            'Yes' if row.get('xero_reconciled') else 'No',
            row.get('xero_reconciled_at') or '',
            len(line_receipts),
            receipt_names,
            _transaction_folder(row) if line_receipts else '',
            issue_text,
        ]
        ws.append([_safe(value) for value in values])
        excel_row = ws.max_row
        fill = (
            _DONE_FILL if status == 'Reconciled'
            else _READY_FILL if status == 'Ready for Xero'
            else _WAIT_FILL
        )
        ws.cell(excel_row, 8).fill = fill
        ws.cell(excel_row, 7).number_format = _MONEY_FORMAT
        for cell in ws[excel_row]:
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    ws.freeze_panes = 'A2'
    # Follow the header list rather than a hardcoded last column, so adding a
    # column can't silently leave it outside the filter (and out of any sort).
    ws.auto_filter.ref = (
        f"A1:{get_column_letter(len(headers))}{max(1, ws.max_row)}")
    ws.sheet_view.showGridLines = False
    widths = {
        1: 14, 2: 16, 3: 18, 4: 24, 5: 28, 6: 34, 7: 20, 8: 20,
        9: 26, 10: 38, 11: 20, 12: 22, 13: 28, 14: 12, 15: 20,
        16: 24, 17: 28, 18: 18, 19: 30, 20: 28, 21: 18, 22: 22,
        23: 14, 24: 42, 25: 48, 26: 38,
    }
    for index, width in widths.items():
        ws.column_dimensions[get_column_letter(index)].width = width

    summary = wb.create_sheet('Pack summary')
    summary.sheet_view.showGridLines = False
    summary['A1'] = 'NORTHWIND credit-card reconciliation finance pack'
    summary['A1'].font = Font(size=16, bold=True, color='242925')
    summary.merge_cells('A1:D1')
    summary.append([
        'One row in Transactions equals one card transaction. This workbook is '
        'a reconciliation aid, not a Xero import file.'
    ])
    summary.merge_cells('A2:D2')
    summary['A2'].alignment = Alignment(wrap_text=True, vertical='top')
    summary.row_dimensions[2].height = 36
    summary.append([])
    summary.append(['Exported', exported_at.isoformat()])
    for label, value in _scope_lines(scope):
        summary.append([label, _safe(value)])
    summary.append([])
    status_header_row = summary.max_row + 1
    summary.append(['Workflow status', 'Transactions'])
    for cell in summary[status_header_row]:
        cell.fill = _HEADER_FILL
        cell.font = Font(color='FFFFFF', bold=True)
    counts = Counter(_workflow_status(dict(row)) for row in rows)
    for status in (
            'Ready for Xero', 'VAT invoice requested', 'Needs cardholder',
            'AI decision', 'Reconciled'):
        summary.append([status, counts.get(status, 0)])
    summary.append(['Total', len(rows)])
    summary.append([])
    summary.append(['Important', 'Xero account coding is advisory. Confirm it in Xero.'])
    summary.append([
        'Ready rule',
        'Receipt or personal, reason, location, submitted to finance, no '
        'receipt match awaiting a decision, and no open VAT invoice request.',
    ])
    summary.append([
        'Amounts',
        'Statement amounts keep their original sign; card spend is normally negative.',
    ])
    summary.column_dimensions['A'].width = 24
    summary.column_dimensions['B'].width = 90
    for row_cells in summary.iter_rows():
        for cell in row_cells:
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def build_finance_pack(rows, receipts, scope, read_receipt, exported_at=None):
    """Return ``(zip_stream, included_count, missing_count)``.

    Receipt blobs are read through the configured storage abstraction.  A
    missing blob is recorded in the workbook and README instead of making the
    entire export fail.
    """
    exported_at = exported_at or dt.datetime.now(dt.timezone.utc)
    grouped = _group_receipts(receipts)
    issues = {}
    missing = 0
    included = 0

    # Each blob is written into the archive as it is read, so only one receipt
    # is resident at a time instead of the whole pack. Entry order in a zip is
    # irrelevant, which lets the workbook and README — both of which need the
    # `issues` map this loop builds — be appended afterwards.
    output = io.BytesIO()
    with zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED) as archive:
        for row in rows:
            row = dict(row)
            line_id = int(row['id'])
            for receipt in grouped.get(line_id, []):
                name = _receipt_name(receipt)
                try:
                    data = read_receipt(receipt['file_path'])
                except FileNotFoundError:
                    issues.setdefault(line_id, []).append(
                        f'Missing stored file: {name}')
                    missing += 1
                    continue
                archive.writestr(
                    f"{_transaction_folder(row)}/"
                    f"{int(receipt['id'])}_{name}",
                    data,
                )
                included += 1

        workbook = build_workbook(
            rows,
            receipts,
            scope,
            exported_at=exported_at,
            receipt_issues=issues,
        )
        readme = [
            'NORTHWIND CREDIT-CARD RECONCILIATION FINANCE PACK',
            '',
            'This is a reconciliation aid, not a Xero import file.',
            'transactions.xlsx contains one row per transaction.',
            'Linked receipts are copied into that transaction’s own folder.',
            'If one receipt supports several transactions, a copy appears in each folder.',
            'Xero account coding is advisory and must be confirmed in Xero.',
            'Open VAT invoice requests are included in the workbook and block Ready for Xero.',
            'Negative statement amounts represent card spend.',
            '',
            f"Transactions: {len(rows)}",
            f"Receipt files included: {included}",
            f"Receipt files missing from storage: {missing}",
            '',
        ]
        for label, value in _scope_lines(scope):
            readme.append(f'{label}: {value}')
        if issues:
            readme.extend(['', 'MISSING RECEIPT FILES'])
            for line_id, messages in sorted(issues.items()):
                for message in messages:
                    readme.append(f'Transaction {line_id}: {message}')

        archive.writestr('transactions.xlsx', workbook.getvalue())
        archive.writestr('README.txt', '\n'.join(readme).encode('utf-8'))
    output.seek(0)
    return output, included, missing


def export_filename(scope, extension):
    period = scope.get('period') or 'all-months'
    status = secure_filename(scope.get('status') or 'unreconciled') or 'unreconciled'
    return f'NORTHWIND-card-finance-pack_{period}_{status}.{extension}'


XLSX_MIME = _XLSX_MIME
