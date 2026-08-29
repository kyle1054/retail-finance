from flask import (render_template, request, redirect, url_for, flash,
                   jsonify, send_file)
from datetime import datetime
import io
import os
import sqlite3
import tempfile
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from northwind.data import database as db
from northwind.core import app, BACKUP_DIR
from northwind.deductions import payroll_sync as sync_service


@app.route('/payroll/sync')
def payroll_sync():
    return render_template('payroll_sync.html')


# Roster parsing, resolution and application all live in
# northwind/deductions/payroll_sync.py so this route and the MCP connector's
# preview_payroll_sync / apply_payroll_sync run ONE implementation. These aliases keep
# the old private names working for anything that referenced them.
_valid_payroll_name = sync_service.valid_payroll_name
_parse_payroll_xlsx = sync_service.parse_xlsx
_parse_payroll_text = sync_service.parse_text


@app.route('/payroll/sync/preview', methods=['POST'])
def payroll_sync_preview():
    pasted = (request.form.get('pasted') or '').strip()

    if pasted:
        # Paste path — text copied straight out of the payroll spreadsheet.
        payroll_employees, rows_read, skipped = _parse_payroll_text(pasted)
        detected_sheet = f'Pasted text ({rows_read} rows read'
        detected_sheet += f', {skipped} skipped)' if skipped else ')'
        if not payroll_employees:
            flash('No valid rows found in the pasted text. Expected '
                  '"Surname, Firstname" then a tab, the store, a tab, and the job title.', 'danger')
            return redirect(url_for('payroll_sync'))
    else:
        # File path — uploaded payroll Excel export.
        if 'file' not in request.files or request.files['file'].filename == '':
            flash('Upload an Excel file or paste the roster text to analyse.', 'danger')
            return redirect(url_for('payroll_sync'))

        f = request.files['file']
        if not f.filename.lower().endswith(('.xlsx', '.xls')):
            flash('Only .xlsx / .xls files are supported.', 'danger')
            return redirect(url_for('payroll_sync'))

        try:
            payroll_employees, detected_sheet = _parse_payroll_xlsx(f.read())
        except ValueError as e:
            flash(str(e), 'danger')
            return redirect(url_for('payroll_sync'))

        if not payroll_employees:
            flash('No employee data found. Expected "Surname, Firstname" in column A.', 'danger')
            return redirect(url_for('payroll_sync'))

    # Resolution lives in the shared service (retail-scoped there, deliberately: an
    # uploaded retail roster must never match, move or terminate an HQ employee).
    resolved = sync_service.resolve(payroll_employees)

    return render_template('payroll_sync.html',
                           preview=True,
                           detected_sheet=detected_sheet,
                           payroll_count=resolved['payroll_count'],
                           store_changes=resolved['store_changes'],
                           not_in_payroll=resolved['not_in_payroll'],
                           new_in_payroll=resolved['new_in_payroll'],
                           fuzzy_matches=resolved['fuzzy_matches'],
                           ambiguous=resolved['ambiguous'],
                           duplicate_employees=resolved['duplicate_employees'],
                           new_stores=resolved['new_stores'])


@app.route('/payroll/sync/apply', methods=['POST'])
def payroll_sync_apply():
    move_ids = request.form.getlist('move[]')
    terminate_ids = request.form.getlist('terminate[]')
    add_store_names = request.form.getlist('add_store[]')
    fuzzy_link_ids = request.form.getlist('fuzzy_link[]')
    add_new_indexes = request.form.getlist('add_new[]')
    # Employees the user explicitly chose to terminate despite an outstanding
    # balance (P3 guard below otherwise keeps owing staff active).
    force_terminate = set(request.form.getlist('force_terminate[]'))

    # Effective date for store moves / terminations: the payroll month being
    # processed if supplied, otherwise now.
    effective = sync_service.effective_date(
        request.form.get('period_year', type=int),
        request.form.get('period_month', type=int))

    # Translate the form into the service's explicit selections. Nothing is inferred:
    # each move/termination/link/addition is something the user ticked.
    moves = [{'id': emp_id,
              'new_store': request.form.get(f'new_store_{emp_id}', '')}
             for emp_id in move_ids]
    fuzzy_links = [{'id': emp_id,
                    'full_name': request.form.get(f'fuzzy_name_{emp_id}', ''),
                    'store': request.form.get(f'fuzzy_store_{emp_id}', ''),
                    'job_title': request.form.get(f'fuzzy_title_{emp_id}', '')}
                   for emp_id in fuzzy_link_ids]
    additions = [{'full_name': request.form.get(f'add_new_name_{idx}', ''),
                  'store': request.form.get(f'add_new_store_{idx}', ''),
                  'job_title': request.form.get(f'add_new_title_{idx}', '')}
                 for idx in add_new_indexes]

    conn = db.get_db()
    try:
        # The writes live in the shared service so this route and the MCP connector's
        # apply_payroll_sync cannot diverge on the termination guard or store history.
        counts = sync_service.apply_decisions(
            conn, moves=moves, terminations=terminate_ids, fuzzy_links=fuzzy_links,
            additions=additions, new_stores=add_store_names,
            force_terminate=force_terminate, effective=effective)
        conn.commit()
    except Exception as e:
        conn.rollback()
        flash(f'An error occurred during database update: {e}', 'danger')
        return redirect(url_for('payroll_sync'))
    finally:
        conn.close()

    moved, terminated = counts['moved'], counts['terminated']
    linked, added, kept_owing = counts['linked'], counts['added'], counts['kept_owing']

    if add_store_names or added:
        db.invalidate_stores_cache()

    parts = []
    if moved:
        parts.append(f'{moved} employee{"s" if moved != 1 else ""} moved to new store')
    if terminated:
        parts.append(f'{terminated} employee{"s" if terminated != 1 else ""} terminated')
    if linked:
        parts.append(f'{linked} spelling/name correction{"s" if linked != 1 else ""} linked')
    if added:
        parts.append(f'{added} new employee{"s" if added != 1 else ""} bulk added')
    if add_store_names:
        parts.append(f'{len(add_store_names)} new store{"s" if len(add_store_names) != 1 else ""} added')

    flash('Payroll sync complete: ' + (', '.join(parts) if parts else 'no changes selected') + '.', 'success' if parts else 'info')
    if kept_owing:
        flash(f'{kept_owing} employee{"s" if kept_owing != 1 else ""} kept active because they still '
              'owe money — tick "terminate anyway" to override.', 'warning')
    return redirect(url_for('employees'))


@app.route('/payroll/sheet', methods=['GET', 'POST'])
def payroll_sheet():
    now = datetime.now()
    if request.method == 'GET':
        return render_template('payroll_sheet.html', now=now, MONTH_NAMES=db.MONTH_NAMES,
                               MONTH_FULL=db.MONTH_FULL)

    year  = int(request.form['year'])
    month = int(request.form['month'])
    paste = request.form.get('employee_list', '').strip()
    query_idx = year * 12 + month

    conn = db.get_db()
    # Retail-only: the payroll sheet builds the retail deductions run, so HQ
    # staff must never be matched into it by a same/similar name.
    all_emps = conn.execute(
        "SELECT id, full_name, current_store FROM employees "
        "WHERE status='active' AND sector='retail'"
    ).fetchall()
    emp_lookup = {e['full_name'].lower(): dict(e) for e in all_emps}

    export_rows = []
    for line in paste.splitlines():
        parts = line.strip().split('\t')
        if not parts or not parts[0].strip():
            continue
        pasted_name  = parts[0].strip()
        pasted_store = parts[1].strip() if len(parts) > 1 else ''

        emp = emp_lookup.get(pasted_name.lower())
        if not emp:
            # fallback: first token match (surname)
            surname = pasted_name.split(',')[0].strip().lower()
            for key, e in emp_lookup.items():
                if key.startswith(surname + ','):
                    emp = e
                    break

        if emp:
            eid = emp['id']
            # Fetch active uniform plans for this employee for this month
            u_rows = conn.execute('''
                SELECT monthly_amount, sale_number, description, total_amount, term_months, payments_made, start_month, start_year
                FROM uniform_deductions
                WHERE employee_id=? AND status='active'
                  AND (start_year*12+start_month)<=?
                  AND (start_year*12+start_month+term_months)>?
            ''', (eid, query_idx, query_idx)).fetchall()

            u_amt = 0.0
            u_refs = []
            u_descs = []
            for r in u_rows:
                inst_idx = query_idx - (r['start_year'] * 12 + r['start_month'])
                inst_amt = db.calc_installment_amount(r['total_amount'], r['monthly_amount'], r['term_months'], inst_idx)
                u_amt += inst_amt
                if r['sale_number']:
                    u_refs.append(r['sale_number'])
                if r['description']:
                    u_descs.append(r['description'])

            # Fetch active layby plans for this employee for this month
            l_rows = conn.execute('''
                SELECT monthly_amount, sale_number, description, total_amount, balance_remaining, term_months, start_month, start_year
                FROM layby_deductions
                WHERE employee_id=? AND status='active'
                  AND (start_year*12+start_month)<=?
                  AND (start_year*12+start_month+term_months)>?
            ''', (eid, query_idx, query_idx)).fetchall()

            l_amt = 0.0
            l_refs = []
            l_descs = []
            for r in l_rows:
                # Deduct the regular monthly, capped at what is still owed — never
                # balloon the final month to the whole remaining balance.
                if r['monthly_amount']:
                    bal = r['balance_remaining'] if r['balance_remaining'] is not None else r['total_amount']
                    inst_amt = round(max(0.0, min(r['monthly_amount'], bal or 0.0)), 2)
                else:
                    inst_idx = query_idx - (r['start_year'] * 12 + r['start_month'])
                    inst_amt = db.calc_installment_amount(r['total_amount'], r['monthly_amount'], r['term_months'], inst_idx)
                l_amt += inst_amt
                if r['sale_number']:
                    l_refs.append(r['sale_number'])
                if r['description']:
                    l_descs.append(r['description'])

            for plan in conn.execute(
                    "SELECT id FROM undercharges_cents WHERE employee_id=? "
                    "AND COALESCE(type,'undercharge')='undercharge'", (eid,)):
                db.ensure_undercharge_schedule(conn, plan['id'])
            uc_rows = conn.execute('''
                SELECT u.sale_number,u.reason,i.amount_cents / 100.0 amount
                FROM undercharge_schedule_items i
                JOIN undercharges u ON u.id=i.undercharge_id
                WHERE u.employee_id=? AND i.due_year=? AND i.due_month=?
                  AND i.state='scheduled'
                  AND (u.type IS NULL OR u.type='undercharge')
            ''', (eid, year, month)).fetchall()
            uc_amt = 0.0
            reimbursement_amt = 0.0
            uc_refs = []
            uc_reasons = []
            for r in uc_rows:
                uc_amt += r['amount']
                if r['amount'] < 0:
                    reimbursement_amt += -r['amount']
                sale = r['sale_number'] or ''
                if sale:
                    uc_refs.append(sale)
                # Reason cell carries the reason text AND this cash-miss's own
                # sale/invoice number, so each undercharge is self-describing.
                reason_txt = ''
                if r['reason']:
                    reason_txt = (f"{r['reason']} (Reimbursement)"
                                  if r['amount'] < 0 else r['reason'])
                if reason_txt and sale:
                    uc_reasons.append(f"{reason_txt} ({sale})")
                elif reason_txt or sale:
                    uc_reasons.append(reason_txt or sale)

            # Ref # and Reason are undercharge-only: a cash-miss is the only
            # deduction that needs an explanation + its sale/invoice number on
            # the payroll sheet. Uniform/lay-by refs & descriptions are omitted.
            combined_ref = ' / '.join(filter(None, dict.fromkeys(uc_refs)))
            combined_reason = ' / '.join(filter(None, dict.fromkeys(uc_reasons)))

            emp_name = emp['full_name']
            if reimbursement_amt > 0:
                if reimbursement_amt == int(reimbursement_amt):
                    emp_name = f"{emp_name} (Pay {int(reimbursement_amt)})"
                else:
                    emp_name = f"{emp_name} (Pay {reimbursement_amt:.2f})"

            export_rows.append({
                'name': emp_name, 'store': emp['current_store'],
                'uniform': round(u_amt, 2), 'layby': round(l_amt, 2),
                'uc': round(uc_amt, 2), 'uc_ref': combined_ref, 'uc_reason': combined_reason,
                'matched': True,
            })
        else:
            export_rows.append({
                'name': pasted_name, 'store': pasted_store,
                'uniform': 0, 'layby': 0, 'uc': 0, 'uc_ref': '', 'uc_reason': '',
                'matched': False,
            })

    conn.close()

    # ── Build Excel ──────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'{db.MONTH_FULL[month]} {year}'

    hdr_fill    = PatternFill('solid', start_color='1F3864')
    hdr_font    = Font(bold=True, color='FFFFFF', size=11)
    alt_fill    = PatternFill('solid', start_color='EBF3FB')
    total_fill  = PatternFill('solid', start_color='D6E4F0')
    unmatched_fill = PatternFill('solid', start_color='FEE2E2')
    unmatched_font = Font(color='991B1B', italic=True)
    zero_font   = Font(color='94A3B8')
    currency_fmt = 'R#,##0.00;-R#,##0.00;"-"'
    thin = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),  bottom=Side(style='thin', color='CCCCCC'),
    )

    headers    = ['#', 'Employee', 'Store', 'Staff Lay-by (R)', 'Staff Uniform (R)',
                  'Undercharge (R)', 'Ref #', 'Reason', 'Total (R)']
    col_widths = [4, 35, 22, 16, 18, 16, 16, 38, 14]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    total_u = total_l = total_uc = 0
    for idx, r in enumerate(export_rows, 1):
        row_num = idx + 1
        total   = r['uniform'] + r['layby'] + r['uc']
        total_u  += r['uniform']
        total_l  += r['layby']
        total_uc += r['uc']

        fill = unmatched_fill if not r['matched'] else (alt_fill if idx % 2 == 0 else None)
        font = unmatched_font if not r['matched'] else None

        values = [idx, r['name'], r['store'],
                  r['layby'] or None, r['uniform'] or None, r['uc'] or None,
                  r['uc_ref'] or None, r['uc_reason'] or None,
                  total or None]

        for col, val in enumerate(map(db.xl_safe, values), 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.border = thin
            cell.alignment = Alignment(vertical='center', wrap_text=(col == 8))
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
            if col in (4, 5, 6, 9) and isinstance(val, (int, float)):
                cell.number_format = currency_fmt
            elif col in (4, 5, 6, 9) and val is None:
                cell.font = zero_font
                cell.value = '-'

    # Totals row
    tr = len(export_rows) + 2
    ws.cell(row=tr, column=1, value='TOTAL').font = Font(bold=True, size=11)
    ws.cell(row=tr, column=2, value=f'{len(export_rows)} employees').font = Font(bold=True, size=11)
    for col, val in [(4, total_l), (5, total_u), (6, total_uc), (9, total_u+total_l+total_uc)]:
        cell = ws.cell(row=tr, column=col, value=val)
        cell.number_format = currency_fmt
        cell.font = Font(bold=True, size=11)
        cell.fill = total_fill
        cell.border = thin
    for col in range(1, 10):
        ws.cell(row=tr, column=col).fill = total_fill
        ws.cell(row=tr, column=col).border = thin

    ws.auto_filter.ref = f'A1:{get_column_letter(9)}1'

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f'PayrollSheet_{db.MONTH_NAMES[month]}{year}.xlsx'
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)


# Hard cap so a fat range can't build a runaway workbook.
MAX_PERIOD_MONTHS = 24


@app.route('/payroll/period', methods=['GET', 'POST'])
def payroll_period():
    """Export every retail deduction across a chosen start→end month range,
    laid out as one month-by-month block per month (the deductions actually
    falling in each month, grouped by store)."""
    now = datetime.now()
    if request.method == 'GET':
        return render_template('payroll_period.html', now=now, MONTH_FULL=db.MONTH_FULL)

    try:
        sm, sy = int(request.form['start_month']), int(request.form['start_year'])
        em, ey = int(request.form['end_month']), int(request.form['end_year'])
    except (KeyError, ValueError):
        flash('Please choose a valid start and end period.', 'danger')
        return redirect(url_for('payroll_period'))

    start_idx, end_idx = sy * 12 + sm, ey * 12 + em
    if not (1 <= sm <= 12 and 1 <= em <= 12):
        flash('Please choose valid months.', 'danger')
        return redirect(url_for('payroll_period'))
    if end_idx < start_idx:
        flash('The end month is before the start month.', 'danger')
        return redirect(url_for('payroll_period'))
    if end_idx - start_idx + 1 > MAX_PERIOD_MONTHS:
        flash(f'Please choose a range of {MAX_PERIOD_MONTHS} months or fewer.', 'danger')
        return redirect(url_for('payroll_period'))

    months = []
    for idx in range(start_idx, end_idx + 1):
        y, m = idx // 12, idx % 12
        if m == 0:
            y, m = y - 1, 12
        months.append((y, m))

    # ── Excel ────────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Deductions'

    banner_fill = PatternFill('solid', start_color='1C1C1C')
    banner_font = Font(bold=True, color='FFFFFF', size=12)
    hdr_fill = PatternFill('solid', start_color='1F3864')
    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    total_fill = PatternFill('solid', start_color='D6E4F0')
    grand_fill = PatternFill('solid', start_color='1F3864')
    grand_font = Font(bold=True, color='FFFFFF', size=12)
    bold = Font(bold=True, size=11)
    muted = Font(color='94A3B8', italic=True)
    center = Alignment(horizontal='center', vertical='center')
    currency_fmt = 'R#,##0.00;-R#,##0.00;"-"'
    thin = Border(*(Side(style='thin', color='CCCCCC'),) * 4)

    headers = ['#', 'Employee', 'Store', 'Uniform (R)', 'Lay-by (R)',
               'Undercharge (R)', 'Ref #', 'Total (R)']
    col_widths = [4, 32, 22, 16, 16, 16, 26, 14]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ncols = len(headers)

    r = 1
    grand_total = 0.0
    for (y, m) in months:
        data = db.get_monthly_data(y, m, 'retail')
        rows = [d for d in data
                if round(d['uniform_total'] + d['layby_total'] + d['undercharge_total'], 2) != 0]
        rows.sort(key=lambda d: ((d['employee']['current_store'] or '').lower(),
                                 (d['employee']['full_name'] or '').lower()))

        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        banner = ws.cell(row=r, column=1, value=f'{db.MONTH_FULL[m]} {y}')
        banner.font = banner_font
        banner.fill = banner_fill
        banner.alignment = center
        ws.row_dimensions[r].height = 20
        r += 1

        if not rows:
            ws.cell(row=r, column=2, value='No deductions this month').font = muted
            r += 2
            continue

        for col, h in enumerate(headers, 1):
            c = ws.cell(row=r, column=col, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = center
            c.border = thin
        r += 1

        month_total = 0.0
        for i, d in enumerate(rows, 1):
            emp = d['employee']
            refs = [p['sale_number'] for p in (d['uniform_plans'] + d['layby_plans'])
                    if p.get('sale_number')]
            refs += [u['sale_number'] for u in d['undercharge_rows'] if u.get('sale_number')]
            combined_ref = ' / '.join(dict.fromkeys(refs))
            total = round(d['uniform_total'] + d['layby_total'] + d['undercharge_total'], 2)
            month_total += total
            values = [i, emp['full_name'], emp['current_store'],
                      d['uniform_total'] or None, d['layby_total'] or None,
                      d['undercharge_total'] or None, combined_ref or None, total or None]
            for col, val in enumerate(map(db.xl_safe, values), 1):
                c = ws.cell(row=r, column=col, value=val)
                c.border = thin
                if col in (4, 5, 6, 8) and isinstance(val, (int, float)):
                    c.number_format = currency_fmt
            r += 1

        sub_label = ws.cell(row=r, column=2, value=f'{len(rows)} staff · month total')
        sub_label.font = bold
        sub_val = ws.cell(row=r, column=ncols, value=round(month_total, 2))
        sub_val.number_format = currency_fmt
        sub_val.font = bold
        for col in range(1, ncols + 1):
            ws.cell(row=r, column=col).fill = total_fill
            ws.cell(row=r, column=col).border = thin
        grand_total += month_total
        r += 2

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols - 1)
    g_label = ws.cell(row=r, column=1,
                      value=f'GRAND TOTAL — {db.MONTH_NAMES[sm]} {sy} to {db.MONTH_NAMES[em]} {ey}')
    g_label.font = grand_font
    g_label.fill = grand_fill
    g_label.alignment = Alignment(horizontal='right', vertical='center')
    g_val = ws.cell(row=r, column=ncols, value=round(grand_total, 2))
    g_val.number_format = currency_fmt
    g_val.font = grand_font
    g_val.fill = grand_fill
    ws.row_dimensions[r].height = 22

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f'Deductions_{db.MONTH_NAMES[sm]}{sy}-{db.MONTH_NAMES[em]}{ey}.xlsx'
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)


# ── Phase 4 Premium Single-Operator Productivity Endpoints ───────────────────


def _snapshot_to(dest_path, source_path=None):
    """Write a consistent copy of a SQLite DB using the online backup API
    (safe alongside live WAL writes, unlike copying the raw file)."""
    src = sqlite3.connect(source_path or db.DB_PATH)
    dst = sqlite3.connect(dest_path)
    try:
        with dst:
            src.backup(dst)
    finally:
        dst.close()
        src.close()


@app.route('/download/backup')
def download_backup():
    try:
        if not os.path.exists(db.DB_PATH):
            flash("Database file not found.", "danger")
            return redirect(url_for('dashboard'))

        fd, tmp_path = tempfile.mkstemp(suffix='.db', prefix='nw_download_')
        os.close(fd)
        try:
            _snapshot_to(tmp_path)
            with open(tmp_path, 'rb') as f:
                payload = io.BytesIO(f.read())
        finally:
            os.remove(tmp_path)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return send_file(
            payload,
            as_attachment=True,
            download_name=f"deductions_backup_{timestamp}.db",
            mimetype='application/x-sqlite3'
        )
    except Exception as e:
        flash(f"Error creating database backup: {str(e)}", "danger")
        return redirect(url_for('dashboard'))


@app.route('/admin/restore-db', methods=['POST'])
def restore_backup():
    """Replace the live database with an uploaded .db backup (admin only).

    The upload is integrity-checked before anything is touched, and the current
    DB is snapshotted to BACKUP_DIR first so a mistaken restore can be undone.
    Main use: seeding/refreshing a hosted instance from a local backup.
    """
    file = request.files.get('backup_file')
    if not file or not file.filename or not file.filename.lower().endswith('.db'):
        flash('Choose a .db backup file to restore.', 'danger')
        return redirect(url_for('import_center'))
    if request.form.get('confirm') != 'yes':
        flash('Tick the confirmation box to restore a backup.', 'danger')
        return redirect(url_for('import_center'))

    fd, tmp_path = tempfile.mkstemp(suffix='.db', prefix='nw_restore_')
    os.close(fd)
    try:
        file.save(tmp_path)

        # Validate before touching the live DB: must be a healthy SQLite file
        # that actually looks like a deductions database.
        try:
            check = sqlite3.connect(tmp_path)
            try:
                ok = check.execute('PRAGMA integrity_check').fetchone()[0]
                names = {r[0] for r in check.execute(
                    "SELECT name FROM sqlite_master WHERE type IN ('table','view')")}
            finally:
                check.close()
        except sqlite3.DatabaseError:
            ok, names = 'corrupt', set()
        if ok != 'ok' or not {'employees', 'stores'} <= names:
            flash('That file is not a valid deductions database backup.', 'danger')
            return redirect(url_for('import_center'))

        # Safety snapshot of what is live right now (never auto-pruned).
        pre_name = None
        if os.path.exists(db.DB_PATH):
            os.makedirs(BACKUP_DIR, exist_ok=True)
            stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            pre_name = f'manual_pre_restore_{stamp}.db'
            _snapshot_to(os.path.join(BACKUP_DIR, pre_name))

        # Copy the uploaded data over the live DB, then bring an older backup
        # up to the current schema.
        _snapshot_to(db.DB_PATH, source_path=tmp_path)
        import migrations
        applied = migrations.run_migrations()
        db.invalidate_stores_cache()

        msg = f'Database restored from {file.filename}.'
        if applied:
            msg += f' Applied {len(applied)} migration(s).'
        if pre_name:
            msg += f' Previous data saved as {pre_name}.'
        flash(msg, 'success')
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
    return redirect(url_for('dashboard'))


@app.route('/payroll/reconcile')
def payroll_reconcile():
    now = datetime.now()
    year = request.args.get('year', now.year, type=int)
    month = request.args.get('month', now.month, type=int)
    
    data = db.get_monthly_data(year, month)
    
    store_map = {}
    for d in data:
        s_name = d['employee']['current_store'] or 'No Store'
        if s_name not in store_map:
            store_map[s_name] = {'name': s_name, 'pending_count': 0, 'pending_amount': 0.0}
        
        active_count = len(d['uniform_plans']) + len(d['layby_plans']) + len(d['undercharge_rows'])
        if active_count > 0:
            store_map[s_name]['pending_count'] += active_count
            store_map[s_name]['pending_amount'] += d['total']
            
    summaries = sorted(store_map.values(), key=lambda x: x['name'])
    
    return render_template(
        'payroll_reconcile.html',
        year=year,
        month=month,
        store_summaries=summaries,
        now=now
    )


@app.route('/api/payroll/reconcile/store/<store_name>')
def api_reconcile_store(store_name):
    try:
        now = datetime.now()
        year = request.args.get('year', now.year, type=int)
        month = request.args.get('month', now.month, type=int)
        
        data = db.get_monthly_data(year, month)
        
        store_data = []
        for d in data:
            if d['employee']['current_store'] == store_name and d['total'] > 0:
                store_data.append({
                    'employee': d['employee'],
                    'uniform_total': d['uniform_total'],
                    'uniform_plans': d['uniform_plans'],
                    'layby_total': d['layby_total'],
                    'layby_plans': d['layby_plans'],
                    'undercharge_total': d['undercharge_total'],
                    'undercharge_rows': d['undercharge_rows'],
                    'total': d['total']
                })
        
        return jsonify({
            'success': True,
            'store': store_name,
            'year': year,
            'month': month,
            'employees': store_data
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/payroll/reconcile/store/<store_name>/bulk-tick', methods=['POST'])
def api_reconcile_bulk_tick(store_name):
    # NOTE: this handler READS plan state from the *_deductions views (rands) but
    # WRITES updates to the *_cents base tables. That split is intentional and
    # safe: the views are read-only projections of the cents tables, so reading
    # the rands view and writing cents back to the base table stay consistent.
    conn = None
    try:
        now = datetime.now()
        year = request.form.get('year', now.year, type=int)
        month = request.form.get('month', now.month, type=int)

        # Respect the payroll lock — the per-employee tick handlers all guard
        # this, and the bulk 'tick all' must too, or it silently mutates a
        # finalised period.
        if db.is_period_locked(year, month):
            return jsonify({'success': False, 'message': 'This payroll period is locked and cannot be modified.'}), 409

        conn = db.get_db()
        emp_rows = conn.execute(
            "SELECT id FROM employees WHERE current_store=? AND status='active'", 
            (store_name,)
        ).fetchall()
        emp_ids = [r['id'] for r in emp_rows]
        
        ticked_count = 0
        for emp_id in emp_ids:
            # All three tick helpers live in database.py and are shared with the
            # monthly per-employee tick, so the reconcile 'tick all' and the
            # monthly path always behave identically (installments, balances,
            # completion, reimbursements).
            ticked_count += db.tick_uniform_due(conn, emp_id, year, month)
            ticked_count += db.tick_layby_due(conn, emp_id, year, month)
            ticked_count += db.tick_undercharges_due(conn, emp_id, year, month)

        conn.commit()
        return jsonify({
            'success': True,
            'store': store_name,
            'ticked_count': ticked_count,
            'message': f"Successfully ticked {ticked_count} active plan(s) for {store_name}."
        })
    except Exception as e:
        if conn is not None:
            conn.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn is not None:
            conn.close()


@app.route('/api/payroll/reconcile/stores-summary')
def api_reconcile_stores_summary():
    try:
        now = datetime.now()
        year = request.args.get('year', now.year, type=int)
        month = request.args.get('month', now.month, type=int)
        
        data = db.get_monthly_data(year, month)
        
        store_map = {}
        for d in data:
            s_name = d['employee']['current_store'] or 'No Store'
            if s_name not in store_map:
                store_map[s_name] = {'name': s_name, 'pending_count': 0, 'pending_amount': 0.0}
            
            active_count = len(d['uniform_plans']) + len(d['layby_plans']) + len(d['undercharge_rows'])
            if active_count > 0:
                store_map[s_name]['pending_count'] += active_count
                store_map[s_name]['pending_amount'] += d['total']
                
        summaries = sorted(store_map.values(), key=lambda x: x['name'])
        return jsonify({
            'success': True,
            'summaries': summaries
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ── Deductions Import Center ───────────────────────────────────────────────
