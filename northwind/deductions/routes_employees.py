from flask import (render_template, request, redirect, url_for, flash,
                   jsonify, send_file)
from datetime import datetime
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from northwind.data import database as db
from northwind.core import app, JOB_TITLES, employee_detail_url
from northwind.deductions.pagination import paginate
from northwind.deductions import requests as staff_requests


@app.route('/employees')
def employees():
    conn = db.get_db()
    store_filter = request.args.get('store', '')
    search = request.args.get('q', '').strip()
    status_filter = request.args.get('status', 'active')

    query = "SELECT * FROM employees WHERE sector='retail'"
    params = []
    if store_filter:
        query += " AND current_store = ?"
        params.append(store_filter)
    if search:
        query += " AND full_name LIKE ?"
        params.append(f'%{search}%')
    if status_filter and status_filter != 'all':
        query += " AND status = ?"
        params.append(status_filter)
    # Default order matches the first visible column (ID) so the list doesn't
    # look shuffled; store/name grouping is one header-click away.
    query += " ORDER BY id"

    emps = conn.execute(query, params).fetchall()
    stores = [s[0] for s in conn.execute("SELECT DISTINCT current_store FROM employees WHERE sector='retail' ORDER BY current_store").fetchall()]
    # Possible duplicates: names that collapse to the same canonical key
    # (exact, or swapped first/last, or a missing comma). Retail scope to match
    # this list. Surfaced as a banner so a mistakenly-added twin is easy to spot.
    duplicate_groups = []
    for g in db.find_duplicate_employees(conn):
        if not any(e['sector'] == 'retail' for e in g):
            continue
        for m in g:
            m['links'] = db.employee_link_count(m['id'], conn)
        # Show the record that actually holds the most data first.
        g.sort(key=lambda m: (-m['links'], m['status'] != 'active', m['full_name']))
        duplicate_groups.append(g)
    conn.close()
    outstanding = db.get_all_outstanding_totals()
    # The count in the card header must stay the size of the whole filtered
    # result, not of the window — otherwise "156 employees" silently becomes
    # "50 employees" and looks like staff went missing.
    total_employees = len(emps)
    emps, pager = paginate(emps, noun='employees')
    return render_template('employees.html', employees=emps, stores=stores,
                           store_filter=store_filter, search=search,
                           status_filter=status_filter,
                           job_titles=JOB_TITLES, outstanding=outstanding,
                           duplicate_groups=duplicate_groups,
                           total_employees=total_employees, pager=pager)


@app.route('/employees/merge', methods=['POST'])
def merge_employee_records():
    """Merge duplicate employee records: pick one to keep, fold the rest into it
    (moving all their plans/transactions), then delete the duplicates."""
    keep_id = request.form.get('keep_id', '').strip()
    member_ids = [m.strip() for m in request.form.getlist('member_ids') if m.strip()]
    remove_ids = [m for m in member_ids if m != keep_id]
    if not keep_id or not remove_ids:
        flash('Select which record to keep before merging.', 'warning')
        return redirect(url_for('employees'))
    try:
        conn = db.get_db()
        kept = conn.execute("SELECT full_name FROM employees WHERE id=?", (keep_id,)).fetchone()
        conn.close()
        moved = db.merge_employees(keep_id, remove_ids)
    except Exception as e:
        flash(f'Merge failed — nothing changed: {e}', 'danger')
        return redirect(url_for('employees'))
    total = sum(moved.values())
    kept_name = kept['full_name'] if kept else keep_id
    flash(f'Merged {len(remove_ids)} duplicate record(s) into {kept_name} ({keep_id}); '
          f'moved {total} linked record(s).', 'success')
    return redirect(url_for('employees'))


@app.route('/employees/add', methods=['POST'])
def add_employee():
    conn = db.get_db()
    emp_id = db.next_employee_id(conn)
    full_name = request.form['full_name'].strip()
    store = request.form['store'].strip()
    job_title = request.form['job_title'].strip()
    conn.execute(
        "INSERT INTO employees (id, full_name, current_store, job_title) VALUES (?, ?, ?, ?)",
        (emp_id, full_name, store, job_title)
    )
    conn.execute(
        "INSERT INTO store_history (employee_id, store, from_date) VALUES (?, ?, ?)",
        (emp_id, store, datetime.now().isoformat())
    )
    conn.commit()
    conn.close()
    flash(f'Employee {full_name} added — ID: {emp_id}', 'success')
    return redirect(url_for('employees'))


@app.route('/employees/<emp_id>')
def employee_detail(emp_id):
    conn = db.get_db()
    emp = conn.execute("SELECT * FROM employees WHERE id=?", (emp_id,)).fetchone()
    if not emp:
        flash('Employee not found.', 'danger')
        return redirect(url_for('employees'))
    # Keep sections separate: an HQ employee opened on a retail URL goes to HQ.
    if (emp['sector'] or 'retail') == 'hq':
        conn.close()
        return redirect(url_for('hq_employee_detail', emp_id=emp_id))

    uniforms = conn.execute(
        "SELECT * FROM uniform_deductions WHERE employee_id=? ORDER BY start_year, start_month, id", (emp_id,)
    ).fetchall()
    laybys = conn.execute(
        "SELECT * FROM layby_deductions WHERE employee_id=? ORDER BY start_year, start_month, id", (emp_id,)
    ).fetchall()
    undercharge_rows = conn.execute(
        "SELECT * FROM undercharges WHERE employee_id=? ORDER BY incident_year DESC, incident_month DESC, id", (emp_id,)
    ).fetchall()
    undercharges = []
    for row in undercharge_rows:
        item = dict(row)
        if (row['type'] or 'undercharge') == 'undercharge':
            account = db.get_undercharge_account(row['id'], conn)
            future = conn.execute(
                "SELECT amount_cents,due_year,due_month FROM undercharge_schedule_items "
                "WHERE undercharge_id=? AND state='scheduled' AND transaction_id IS NULL "
                "AND amount_cents>0 ORDER BY due_year,due_month,id",
                (row['id'],)).fetchall()
            item.update(
                paid_amount=db.to_rands(account['net_employee_paid_cents']),
                remaining_amount=db.to_rands(account['remaining_cents']),
                payments_made=account['payment_count'],
                display_status=account['status'],
                future_count=len(future),
                future_monthly=(db.to_rands(future[0]['amount_cents']) if future else 0))
        undercharges.append(item)
    # Anything this person has asked for that is still open, so an admin on the
    # profile sees the ask without going to /requests first.
    open_requests = [r for r in staff_requests.list_requests(
        conn, status='open', employee_id=emp_id, limit=10)]
    history = conn.execute(
        "SELECT * FROM store_history WHERE employee_id=? ORDER BY from_date DESC", (emp_id,)
    ).fetchall()
    transactions = conn.execute(
        "SELECT * FROM deduction_transactions WHERE employee_id=? ORDER BY created_at DESC LIMIT 20", (emp_id,)
    ).fetchall()
    conn.close()

    outstanding = db.get_outstanding_summary(emp_id)
    cat_totals  = db.get_category_totals(emp_id)
    schedule    = db.get_employee_schedule(emp_id)
    return render_template('employee.html', emp=emp, uniforms=uniforms, laybys=laybys,
                           undercharges=undercharges, history=history, outstanding=outstanding,
                           cat_totals=cat_totals, schedule=schedule,
                           stores=db.get_stores(), job_titles=JOB_TITLES,
                           MONTH_NAMES=db.MONTH_NAMES, MONTH_FULL=db.MONTH_FULL,
                           transactions=transactions, open_requests=open_requests)


@app.route('/employees/<emp_id>/change-store', methods=['POST'])
def change_store(emp_id):
    conn = db.get_db()
    new_store = request.form['store']
    emp = conn.execute("SELECT current_store FROM employees WHERE id=?", (emp_id,)).fetchone()
    if emp and emp['current_store'] == new_store:
        conn.close()
        flash('Employee is already assigned to that store.', 'info')
        return redirect(employee_detail_url(emp_id))
    now = datetime.now().isoformat()
    conn.execute("UPDATE store_history SET to_date=? WHERE employee_id=? AND to_date IS NULL", (now, emp_id))
    conn.execute("INSERT INTO store_history (employee_id, store, from_date) VALUES (?, ?, ?)", (emp_id, new_store, now))
    conn.execute("UPDATE employees SET current_store=? WHERE id=?", (new_store, emp_id))
    conn.commit()
    conn.close()
    flash(f'Store updated to {new_store}.', 'success')
    return redirect(employee_detail_url(emp_id))


@app.route('/employees/<emp_id>/terminate', methods=['POST'])
def terminate_employee(emp_id):
    conn = db.get_db()
    now = datetime.now().isoformat()
    conn.execute("UPDATE employees SET status='terminated', terminated_at=? WHERE id=?", (now, emp_id))
    conn.execute("UPDATE store_history SET to_date=? WHERE employee_id=? AND to_date IS NULL", (now, emp_id))
    conn.commit()
    conn.close()
    flash('Employee marked as terminated.', 'warning')
    return redirect(employee_detail_url(emp_id))


@app.route('/employees/<emp_id>/edit', methods=['POST'])
def edit_employee(emp_id):
    conn = db.get_db()
    full_name = request.form['full_name'].strip()
    job_title = request.form['job_title'].strip()
    notes = request.form.get('notes', '').strip()
    conn.execute(
        "UPDATE employees SET full_name=?, job_title=?, notes=? WHERE id=?",
        (full_name, job_title, notes, emp_id)
    )
    conn.commit()
    conn.close()
    flash('Employee details updated.', 'success')
    return redirect(employee_detail_url(emp_id))


@app.route('/employees/<emp_id>/reactivate', methods=['POST'])
def reactivate_employee(emp_id):
    conn = db.get_db()
    emp = conn.execute("SELECT current_store FROM employees WHERE id=?", (emp_id,)).fetchone()
    now = datetime.now().isoformat()
    conn.execute("UPDATE employees SET status='active', terminated_at=NULL WHERE id=?", (emp_id,))
    conn.execute("INSERT INTO store_history (employee_id, store, from_date) VALUES (?, ?, ?)",
                 (emp_id, emp['current_store'], now))
    conn.commit()
    conn.close()
    flash('Employee reactivated.', 'success')
    return redirect(employee_detail_url(emp_id))


@app.route('/api/employees/search')
def search_employees_api():
    q = request.args.get('q', '').strip()
    exclude = request.args.get('exclude', '')
    if not q:
        return jsonify([])
    conn = db.get_db()
    rows = conn.execute(
        "SELECT id, full_name, current_store FROM employees "
        "WHERE status='active' AND (full_name LIKE ? OR id LIKE ?) AND id != ? "
        "ORDER BY full_name LIMIT 10",
        (f'%{q}%', f'%{q}%', exclude)
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/uniform/<int:plan_id>/transfer', methods=['POST'])
def transfer_uniform(plan_id):
    target = request.form['target_emp_id']
    conn = db.get_db()
    plan = conn.execute("SELECT employee_id FROM uniform_deductions WHERE id=?", (plan_id,)).fetchone()
    source = plan['employee_id']
    conn.execute("UPDATE uniform_deductions_cents SET employee_id=? WHERE id=?", (target, plan_id))
    conn.commit()
    conn.close()
    flash('Uniform plan transferred successfully.', 'success')
    return redirect(employee_detail_url(source))


@app.route('/layby/<int:plan_id>/transfer', methods=['POST'])
def transfer_layby(plan_id):
    target = request.form['target_emp_id']
    conn = db.get_db()
    plan = conn.execute("SELECT employee_id FROM layby_deductions WHERE id=?", (plan_id,)).fetchone()
    source = plan['employee_id']
    conn.execute("UPDATE layby_deductions_cents SET employee_id=? WHERE id=?", (target, plan_id))
    conn.commit()
    conn.close()
    flash('Lay-by plan transferred successfully.', 'success')
    return redirect(employee_detail_url(source))


@app.route('/undercharge/<int:uc_id>/transfer', methods=['POST'])
def transfer_undercharge(uc_id):
    target = request.form['target_emp_id']
    conn = db.get_db()
    uc = conn.execute("SELECT employee_id FROM undercharges WHERE id=?", (uc_id,)).fetchone()
    source = uc['employee_id']
    has_history = conn.execute(
        "SELECT 1 FROM deduction_transactions_cents WHERE plan_type='undercharge' "
        "AND plan_id=? AND COALESCE(voided,0)=0 LIMIT 1", (uc_id,)).fetchone()
    has_events = conn.execute(
        "SELECT 1 FROM undercharge_events WHERE undercharge_id=? LIMIT 1",
        (uc_id,)).fetchone()
    if has_history or has_events:
        conn.close()
        flash('This undercharge already has financial history and cannot be '
              'transferred without rewriting who paid it.', 'danger')
        return redirect(employee_detail_url(source))
    conn.execute("UPDATE undercharges_cents SET employee_id=? WHERE id=?", (target, uc_id))
    conn.commit()
    conn.close()
    flash('Undercharge transferred successfully.', 'success')
    return redirect(employee_detail_url(source))


@app.route('/api/employees/<emp_id>/notes', methods=['POST'])
def update_employee_notes_api(emp_id):
    conn = db.get_db()
    if request.is_json:
        data = request.get_json()
        notes = data.get('notes', '').strip()
    else:
        notes = request.form.get('notes', '').strip()
    conn.execute("UPDATE employees SET notes=? WHERE id=?", (notes, emp_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'notes': notes})


@app.route('/employees/<emp_id>/export')
def export_employee(emp_id):
    conn = db.get_db()
    emp = conn.execute("SELECT * FROM employees WHERE id=?", (emp_id,)).fetchone()
    uniforms = conn.execute(
        "SELECT * FROM uniform_deductions WHERE employee_id=? ORDER BY start_year,start_month", (emp_id,)
    ).fetchall()
    laybys = conn.execute(
        "SELECT * FROM layby_deductions WHERE employee_id=? ORDER BY start_year,start_month", (emp_id,)
    ).fetchall()
    undercharge_rows = conn.execute(
        "SELECT * FROM undercharges WHERE employee_id=? ORDER BY incident_year DESC,incident_month DESC", (emp_id,)
    ).fetchall()
    undercharges = []
    for row in undercharge_rows:
        item = dict(row)
        if (row['type'] or 'undercharge') == 'undercharge':
            account = db.get_undercharge_account(row['id'], conn)
            item.update(
                paid_amount=db.to_rands(account['net_employee_paid_cents']),
                remaining_amount=db.to_rands(account['remaining_cents']),
                payments_made=account['payment_count'],
                display_status=account['status'])
        undercharges.append(item)
    conn.close()

    wb = openpyxl.Workbook()
    hdr_fill   = PatternFill('solid', start_color='1F3864')
    hdr_font   = Font(bold=True, color='FFFFFF', size=11)
    sec_fill   = PatternFill('solid', start_color='EFF6FF')
    thin       = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
                        top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))
    cur_fmt    = 'R#,##0.00;-R#,##0.00;"-"'

    def make_header(ws, title, headers, widths):
        ws.title = title
        for col, (h, w) in enumerate(zip(headers, widths), 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = hdr_font; c.fill = hdr_fill
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = thin
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[1].height = 20
        ws.freeze_panes = 'A2'

    # Sheet 1: Uniform
    ws1 = wb.active
    make_header(ws1, 'Uniform', ['Description', 'SKU', 'Sale #', 'Start', 'Term', 'Monthly (R)', 'Total (R)', 'Paid (R)', 'Remaining (R)', 'Status'],
                [32, 14, 14, 10, 7, 14, 14, 14, 16, 12])
    for i, u in enumerate(uniforms, 2):
        total_amt = u['total_amount'] if u['total_amount'] is not None else round(u['term_months'] * u['monthly_amount'], 2)
        pmt_made = u['payments_made']
        monthly_rounded = round(u['monthly_amount'], 2)
        if pmt_made >= u['term_months']:
            paid_val = total_amt
            rem_val = 0.0
        else:
            paid_val = pmt_made * monthly_rounded
            rem_val = total_amt - paid_val

        row = [u['description'] or '', u['sku'] or '', u['sale_number'] or '',
               f"{db.MONTH_NAMES[u['start_month']]} {u['start_year']}", u['term_months'],
               u['monthly_amount'], total_amt,
               paid_val, rem_val, u['status']]
        for col, val in enumerate(map(db.xl_safe, row), 1):
            c = ws1.cell(row=i, column=col, value=val)
            c.border = thin
            if col in (6, 7, 8, 9):
                c.number_format = cur_fmt; c.alignment = Alignment(horizontal='right')
        if i % 2 == 0:
            for col in range(1, 11):
                ws1.cell(row=i, column=col).fill = sec_fill

    # Sheet 2: Lay-by
    ws2 = wb.create_sheet('Lay-by')
    make_header(ws2, 'Lay-by', ['Description', 'Sale #', 'Start', 'Term', 'Basket (R)', 'Disc %', 'Total (R)', 'Monthly (R)', 'Balance (R)', 'Status'],
                [32, 14, 10, 7, 14, 8, 14, 14, 14, 12])
    for i, l in enumerate(laybys, 2):
        row = [l['description'] or '', l['sale_number'] or '',
               f"{db.MONTH_NAMES[l['start_month']]} {l['start_year']}", l['term_months'],
               l['basket_total'] or 0, l['discount_pct'] or 40,
               l['total_amount'] or 0, l['monthly_amount'],
               l['balance_remaining'] or 0, l['status']]
        for col, val in enumerate(map(db.xl_safe, row), 1):
            c = ws2.cell(row=i, column=col, value=val)
            c.border = thin
            if col in (5, 7, 8, 9):
                c.number_format = cur_fmt; c.alignment = Alignment(horizontal='right')
            if col == 6:
                c.number_format = '0.0"%"'; c.alignment = Alignment(horizontal='right')
        if i % 2 == 0:
            for col in range(1, 11):
                ws2.cell(row=i, column=col).fill = sec_fill

    # Sheet 3: Undercharges
    ws3 = wb.create_sheet('Undercharges')
    make_header(ws3, 'Undercharges', ['Reason', 'Sale #', 'Date', 'Amount (R)', 'Recovery', 'Payments Made', 'Remaining (R)', 'Status'],
                [28, 14, 10, 14, 20, 14, 16, 12])
    for i, uc in enumerate(undercharges, 2):
        rem = uc.get('remaining_amount', 0)
        recovery_str = ('On record only' if (uc['type'] or 'undercharge') == 'overcharge'
                        else 'Versioned schedule')
        row = [uc['reason'] or '', uc['sale_number'] or '',
               f"{db.MONTH_NAMES[uc['incident_month']]} {uc['incident_year']}",
               uc['total_amount'], recovery_str, uc['payments_made'], round(rem, 2),
               uc.get('display_status', uc['status'])]
        for col, val in enumerate(map(db.xl_safe, row), 1):
            c = ws3.cell(row=i, column=col, value=val)
            c.border = thin
            if col in (4, 7):
                c.number_format = cur_fmt; c.alignment = Alignment(horizontal='right')
        if i % 2 == 0:
            for col in range(1, 9):
                ws3.cell(row=i, column=col).fill = sec_fill

    out = io.BytesIO()
    wb.save(out); out.seek(0)
    safe_name = emp['full_name'].replace(', ', '_').replace(' ', '_')
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'Employee_{safe_name}_{emp_id}.xlsx')


@app.route('/employees/<emp_id>/termination-summary')
def termination_summary(emp_id):
    conn = db.get_db()
    emp = conn.execute("SELECT full_name FROM employees WHERE id=?", (emp_id,)).fetchone()

    uniform_plans = conn.execute(
        "SELECT description, monthly_amount, term_months, payments_made, total_amount, balance_remaining FROM uniform_deductions WHERE employee_id=? AND status='active'",
        (emp_id,)
    ).fetchall()

    layby_plans = conn.execute(
        "SELECT description, monthly_amount, term_months, payments_made FROM layby_deductions WHERE employee_id=? AND status='active'",
        (emp_id,)
    ).fetchall()

    uc_rows = conn.execute(
        "SELECT id,reason FROM undercharges WHERE employee_id=? "
        "AND (type IS NULL OR type='undercharge')", (emp_id,)).fetchall()
    uc_items = []
    for row in uc_rows:
        remaining = db.to_rands(db.get_undercharge_account(row['id'], conn)['remaining_cents'])
        if remaining > 0:
            uc_items.append({'reason': row['reason'] or 'Undercharge',
                             'amount': remaining})
    conn.close()

    uniform_total = 0.0
    for r in uniform_plans:
        if r['balance_remaining'] is not None:
            uniform_total += r['balance_remaining']
        else:
            total_amt = r['total_amount'] if r['total_amount'] is not None else round(r['term_months'] * r['monthly_amount'], 2)
            pmt_made = r['payments_made']
            monthly_rounded = round(r['monthly_amount'], 2)
            if pmt_made < r['term_months']:
                uniform_total += (total_amt - (pmt_made * monthly_rounded))

    layby_total = sum((r['term_months'] - r['payments_made']) * r['monthly_amount'] for r in layby_plans)
    uc_total = sum(r['amount'] for r in uc_items)

    return jsonify({
        'name': emp['full_name'],
        'uniform': round(uniform_total, 2),
        'uniform_plans': [{'desc': r['description'] or 'Uniform', 'remaining': max(0, r['term_months'] - r['payments_made']),
                           'monthly': r['monthly_amount']} for r in uniform_plans],
        'layby': round(layby_total, 2),
        'layby_plans': [{'desc': r['description'] or 'Lay-by', 'remaining': max(0, r['term_months'] - r['payments_made']),
                         'monthly': r['monthly_amount']} for r in layby_plans],
        'undercharges': round(uc_total, 2),
        'uc_items': [{'reason': r['reason'], 'amount': round(r['amount'], 2)}
                     for r in uc_items],
        'total': round(uniform_total + layby_total + uc_total, 2)
    })


# ── Uniform Deductions ──────────────────────────────────────────────────────
