from flask import render_template, request, redirect, url_for, flash
from datetime import datetime
import openpyxl
from northwind.data import database as db
from northwind.core import app


@app.route('/import-center')
def import_center():
    return render_template('import_center.html', results=None)


@app.route('/import/uniforms', methods=['POST'])
def import_uniforms_action():
    if 'file' not in request.files:
        flash('No file part', 'danger')
        return redirect(url_for('import_center'))
    
    file = request.files['file']
    if file.filename == '':
        flash('No selected file', 'danger')
        return redirect(url_for('import_center'))
    
    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        flash('Only Excel files are allowed.', 'danger')
        return redirect(url_for('import_center'))
        
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        conn = db.get_db()
        
        # ── Build employee lookup ────────────────────────────────────────────
        db_employees = conn.execute('SELECT id, full_name FROM employees').fetchall()
        db_lookup = {}   # (surname_lower, first_lower) -> emp_id
        for emp in db_employees:
            parts = emp['full_name'].split(',', 1)
            if len(parts) == 2:
                surname = parts[0].strip().lower()
                first = parts[1].strip().lower()
                db_lookup[(surname, first)] = emp['id']
                
        # ── Handle Name Overrides and Skips ──────────────────────────────────
        # NOTE: This dictionary resolves historical manual spelling mismatches
        # between payroll Excel imports and existing employees in the database.
        # This mapping is safe to retain for backwards compatibility of one-time imports.
        NAME_OVERRIDES = {
            # Maps a payroll-sheet spelling to a staff number when the two
            # disagree. Populate per deployment; empty by default.
        }
        
        SKIP_NAMES = {
            # Payroll-sheet rows that are not employees and must be ignored.
        }
        
        MONTH_MAP = {
            'jan': 1, 'feb': 2, 'mar': 3, 'march': 3,
            'apr': 4, 'april': 4, 'may': 5, 'jun': 6, 'june': 6,
            'jul': 7, 'july': 7, 'aug': 8, 'sep': 9, 'sept': 9,
            'oct': 10, 'nov': 11, 'dec': 12
        }
        
        CYCLE_START_MONTH = 4
        CYCLE_START_YEAR = 2026
        GREEN_RGB = 'FF92D050'
        PAYMENT_COL_START = 10
        PAYMENT_COL_END = 14
        SKIP_SHEETS = {'Master', 'J', 'Payroll', 'TEMPLATE'}
        
        def is_green(cell):
            try:
                if cell.fill and cell.fill.start_color:
                    return cell.fill.start_color.rgb == GREEN_RGB
            except Exception:
                pass
            return False

        def count_green_payments(row_cells):
            count = 0
            for i in range(PAYMENT_COL_START, min(PAYMENT_COL_END + 1, len(row_cells))):
                if is_green(row_cells[i]):
                    count += 1
                else:
                    break
            return count

        def match_employee(first_name, surname, db_lookup):
            first_lower = first_name.lower().strip()
            surname_lower = surname.lower().strip()

            if (first_lower, surname_lower) in SKIP_NAMES:
                return '__SKIP__'

            if (first_lower, surname_lower) in NAME_OVERRIDES:
                return NAME_OVERRIDES[(first_lower, surname_lower)]

            candidates = []
            for (db_sn, db_fn), emp_id in db_lookup.items():
                if db_sn == surname_lower and (first_lower in db_fn or db_fn in first_lower):
                    candidates.append(emp_id)
                elif (surname_lower in db_sn or db_sn in surname_lower) and \
                     (first_lower in db_fn or db_fn in first_lower):
                    candidates.append(emp_id)

            if len(candidates) == 1:
                return candidates[0]
            elif len(candidates) > 1:
                exact = [c for (db_sn, db_fn), c in db_lookup.items()
                         if db_sn == surname_lower and (first_lower in db_fn or db_fn in first_lower)]
                if len(exact) == 1:
                    return exact[0]
                return candidates[0]

            return None

        def parse_sheet(ws, db_lookup):
            items = {}
            unmatched = []
            current_first = None
            current_surname = None

            for row in ws.iter_rows(min_row=2, values_only=False):
                if len(row) < 11:
                    continue

                name_val = row[0].value
                surname_val = row[1].value
                if name_val and str(name_val).strip():
                    current_first = str(name_val).strip()
                if surname_val and str(surname_val).strip():
                    current_surname = str(surname_val).strip()

                desc = row[3].value
                if not desc or not str(desc).strip():
                    continue

                term_val = row[5].value
                if not term_val:
                    continue
                try:
                    term = int(float(term_val))
                except (ValueError, TypeError):
                    continue
                if term <= 0:
                    continue

                monthly_val = row[4].value
                if not monthly_val:
                    continue
                try:
                    monthly = float(monthly_val)
                except (ValueError, TypeError):
                    continue
                if monthly <= 0:
                    continue

                total_val = row[6].value
                try:
                    total = float(total_val) if total_val and float(total_val) > 0 else round(monthly * term, 2)
                except (ValueError, TypeError):
                    total = round(monthly * term, 2)

                sku_val = row[2].value
                if isinstance(sku_val, (int, float)) and sku_val > 0:
                    sku = str(int(sku_val))
                elif sku_val:
                    sku = str(sku_val).strip()
                    if sku == '0':
                        sku = ''
                else:
                    sku = ''

                start_str = row[8].value
                start_month = CYCLE_START_MONTH
                if start_str:
                    clean = str(start_str).strip().lower()
                    start_month = MONTH_MAP.get(clean, CYCLE_START_MONTH)

                sale_val = row[7].value
                sale_number = str(sale_val).strip() if sale_val else ''
                if isinstance(sale_val, float) and sale_val.is_integer():
                    sale_number = str(int(sale_val))
                elif isinstance(sale_val, (int, float)):
                    sale_number = str(sale_val)

                green_count = count_green_payments(row)

                if not current_first or not current_surname:
                    continue

                emp_id = match_employee(current_first, current_surname, db_lookup)
                if emp_id == '__SKIP__':
                    continue
                if emp_id is None:
                    key = (current_first, current_surname)
                    if key not in [(u[0], u[1]) for u in unmatched]:
                        unmatched.append((current_first, current_surname, ws.title, row[0].row))
                    continue

                if emp_id not in items:
                    items[emp_id] = []

                items[emp_id].append({
                    'description': str(desc).strip(),
                    'sku': sku,
                    'sale_number': sale_number,
                    'monthly': monthly,
                    'total': total,
                    'term': term,
                    'start_month': start_month,
                    'green_payments': green_count,
                    'source_name': f"{current_first} {current_surname}",
                })

            return items, unmatched

        # Employees are created by the import when a payroll row has
        # no match, not seeded here.
        created_employees = 0


        # ── Find already-imported employees ──────────────────────────────────
        existing = set()
        for r in conn.execute(
            'SELECT DISTINCT employee_id FROM uniform_deductions '
            'WHERE start_month = ? AND start_year = ?',
            (CYCLE_START_MONTH, CYCLE_START_YEAR)
        ).fetchall():
            existing.add(r['employee_id'])

        # ── Parse every store sheet ──────────────────────────────────────────
        all_items = {}
        all_unmatched = []
        sheets_processed = 0
        skipped_existing = set()
        import_logs = []

        import_logs.append("Starting Uniform Import...")
        for sheet_name in wb.sheetnames:
            if sheet_name in SKIP_SHEETS:
                continue

            ws = wb[sheet_name]
            sheet_items, unmatched = parse_sheet(ws, db_lookup)
            sheets_processed += 1
            import_logs.append(f"Parsed sheet '{sheet_name}': found {len(sheet_items)} employees.")

            for emp_id, items in sheet_items.items():
                if emp_id in existing:
                    skipped_existing.add(emp_id)
                    continue
                if emp_id not in all_items:
                    all_items[emp_id] = []
                all_items[emp_id].extend(items)

            all_unmatched.extend(unmatched)

        # ── Insert consolidated plans ────────────────────────────────────────
        inserted = 0
        emp_info = {}
        for r in conn.execute("SELECT id, full_name, current_store FROM employees").fetchall():
            emp_info[r['id']] = {'name': r['full_name'], 'store': r['current_store']}

        for emp_id in sorted(all_items.keys()):
            items = all_items[emp_id]
            info = emp_info.get(emp_id, {'name': '???', 'store': '???'})

            descriptions = [item['description'] for item in items]
            skus = [item['sku'] for item in items if item['sku']]
            sale_numbers = [item['sale_number'] for item in items if item['sale_number']]
            total_amount = sum(item['total'] for item in items)
            term = max(item['term'] for item in items)
            start_month = items[0]['start_month']

            # Recalculate monthly from total/term for precision
            monthly_amount = total_amount / term
            payments_made = min(item['green_payments'] for item in items)

            combined_desc = ' / '.join(descriptions)
            combined_skus = ', '.join(skus)
            combined_sales = ', '.join(dict.fromkeys(sale_numbers))

            # Calculate initial balance_remaining
            monthly_rounded = round(monthly_amount, 2)
            if payments_made >= term:
                balance = 0.0
                status = 'complete'
            else:
                balance = round(total_amount - (payments_made * monthly_rounded), 2)
                status = 'active'

            conn.execute('''
                INSERT INTO uniform_deductions_cents
                    (employee_id, sku, description, sale_number, total_amount_cents, monthly_amount_cents,
                     balance_remaining_cents, term_months, start_month, start_year, payments_made, status)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                emp_id, combined_skus, combined_desc, combined_sales, db.to_cents(total_amount), db.to_cents(monthly_amount),
                db.to_cents(balance), term, start_month, CYCLE_START_YEAR, payments_made, status
            ))
            
            inserted += 1
            log_msg = (f"✓ Consolidated Uniform: {info['name']} | {info['store']} | "
                       f"R{monthly_rounded:.2f}/mo × {term} | Paid: {payments_made}/{term} | Bal: R{balance:.2f}")
            import_logs.append(log_msg)

        conn.commit()
        conn.close()

        flash(f"Successfully processed {sheets_processed} sheets. Imported {inserted} uniform plans.", "success")
        
        results = {
            'inserted': inserted,
            'skipped': len(skipped_existing),
            'unmatched': [{'sheet': u[2], 'row_num': u[3], 'name': f"{u[0]} {u[1]}"} for u in all_unmatched],
            'created_employees': created_employees,
            'logs': import_logs
        }
        return render_template('import_center.html', results=results)

    except Exception as e:
        flash(f"Error processing workbook: {str(e)}", "danger")
        return redirect(url_for('import_center'))


@app.route('/import/laybys', methods=['POST'])
def import_laybys_action():
    if 'file' not in request.files:
        flash('No file part', 'danger')
        return redirect(url_for('import_center'))
    
    file = request.files['file']
    if file.filename == '':
        flash('No selected file', 'danger')
        return redirect(url_for('import_center'))
        
    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        flash('Only Excel files are allowed.', 'danger')
        return redirect(url_for('import_center'))
        
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        conn = db.get_db()
        
        # ── Build employees lookups ──────────────────────────────────────────
        db_employees = conn.execute("SELECT id, full_name, current_store FROM employees").fetchall()
        db_lookup_by_store = {}
        db_lookup_global = {} # (store, surname, first) -> emp_id
        
        for r in db_employees:
            store = (r['current_store'] or '').lower().strip()
            name = r['full_name'].strip()
            eid = r['id']
            db_lookup_by_store[(store, name.lower())] = eid
            
            parts = []
            if ',' in name:
                parts = [p.strip() for p in name.split(',', 1)]
                surname = parts[0].lower()
                first = parts[1].lower()
            else:
                p_split = name.split()
                if len(p_split) >= 2:
                    first = p_split[0].lower()
                    surname = p_split[-1].lower()
                else:
                    first = name.lower()
                    surname = ""
            db_lookup_global[(store, surname, first)] = eid

        MONTH_MAP = {
            'jan': 1, 'feb': 2, 'mar': 3, 'march': 3,
            'apr': 4, 'april': 4, 'may': 5, 'jun': 6, 'june': 6,
            'jul': 7, 'july': 7, 'aug': 8, 'sep': 9, 'sept': 9,
            'oct': 10, 'nov': 11, 'dec': 12
        }
        
        def match_layby_employee(name_str, store_str):
            name_clean = name_str.strip().lower()
            store_clean = store_str.strip().lower()
            
            if (store_clean, name_clean) in db_lookup_by_store:
                return db_lookup_by_store[(store_clean, name_clean)]
                
            parts = []
            if ',' in name_clean:
                parts = [p.strip() for p in name_clean.split(',', 1)]
                first = parts[1]
                last = parts[0]
            else:
                parts = name_clean.split()
                if len(parts) >= 2:
                    first = parts[0]
                    last = parts[-1]
                elif len(parts) == 1:
                    first = parts[0]
                    last = ""
                else:
                    first = ""
                    last = ""
                    
            # Try fuzzy match in store
            candidates_store = []
            for (db_store, db_sn, db_fn), emp_id in db_lookup_global.items():
                if db_store == store_clean:
                    if last and db_sn == last and (first in db_fn or db_fn in first):
                        candidates_store.append(emp_id)
                    elif last and (last in db_sn or db_sn in last) and (first in db_fn or db_fn in first):
                        candidates_store.append(emp_id)
                        
            if len(candidates_store) == 1:
                return candidates_store[0]
            elif len(candidates_store) > 1:
                return candidates_store[0]
                
            # Try fuzzy match company-wide (transferred employee)
            candidates_global = []
            for (db_store, db_sn, db_fn), emp_id in db_lookup_global.items():
                if last and db_sn == last and (first in db_fn or db_fn in first):
                    candidates_global.append(emp_id)
                elif last and (last in db_sn or db_sn in last) and (first in db_fn or db_fn in first):
                    candidates_global.append(emp_id)
                    
            if len(candidates_global) == 1:
                return candidates_global[0]
            elif len(candidates_global) > 1:
                return candidates_global[0]
                
            if len(parts) == 1 and name_clean:
                for (db_store, db_sn, db_fn), emp_id in db_lookup_global.items():
                    if db_sn == name_clean or db_fn == name_clean:
                        return emp_id
                        
            return None

        # ── Parse Sheets ─────────────────────────────────────────────────────
        inserted = 0
        skipped = 0
        unmatched = []
        import_logs = []
        
        import_logs.append("Starting Lay-by Import...")
        sheets_processed = 0
        
        for sheet_name in wb.sheetnames:
            if sheet_name in {'Master', 'J', 'Payroll', 'TEMPLATE'}:
                continue
                
            ws = wb[sheet_name]
            sheets_processed += 1
            import_logs.append(f"Parsing sheet '{sheet_name}'...")
            
            for row in ws.iter_rows(min_row=2, values_only=False):
                if len(row) < 6:
                    continue
                    
                store_val = row[0].value
                emp_val = row[1].value
                
                # Check for empty essential names
                if not store_val or not emp_val:
                    continue
                    
                store_str = str(store_val).strip()
                emp_str = str(emp_val).strip()
                
                # Description
                desc_val = row[3].value
                description = str(desc_val).strip() if desc_val else 'Lay-by Item'
                
                # Basket total
                basket_val = row[4].value
                if not basket_val:
                    continue
                try:
                    basket_total = float(basket_val)
                except (ValueError, TypeError):
                    continue
                if basket_total <= 0:
                    continue
                    
                # Sale Number (SO)
                sale_val = row[2].value
                sale_number = str(sale_val).strip() if sale_val else ''
                if isinstance(sale_val, float) and sale_val.is_integer():
                    sale_number = str(int(sale_val))
                elif isinstance(sale_val, (int, float)):
                    sale_number = str(sale_val)
                
                # Term months
                term_val = row[5].value
                term = 6
                if term_val:
                    try:
                        term = int(float(term_val))
                    except (ValueError, TypeError):
                        pass
                if term <= 0:
                    term = 6
                    
                # Optional discount_pct
                discount_pct = 40.0
                if len(row) >= 7 and row[6].value is not None:
                    try:
                        discount_pct = float(row[6].value)
                    except (ValueError, TypeError):
                        pass
                        
                # Optional payments_made
                payments_made = 0
                if len(row) >= 8 and row[7].value is not None:
                    try:
                        payments_made = int(float(row[7].value))
                    except (ValueError, TypeError):
                        pass
                        
                # Optional start_month / start_year
                start_month = 5  # Default to May
                start_year = 2026
                if len(row) >= 9 and row[8].value is not None:
                    raw_m = str(row[8].value).strip().lower()
                    if raw_m.isdigit():
                        start_month = int(raw_m)
                    else:
                        start_month = MONTH_MAP.get(raw_m, 5)
                        
                if len(row) >= 10 and row[9].value is not None:
                    try:
                        start_year = int(float(row[9].value))
                    except (ValueError, TypeError):
                        pass
                        
                # Optional notes
                notes = ''
                if len(row) >= 11 and row[10].value is not None:
                    notes = str(row[10].value).strip()
                    
                # Match employee
                emp_id = match_layby_employee(emp_str, store_str)
                if not emp_id:
                    key = (emp_str, store_str)
                    if key not in [(u[0], u[1]) for u in unmatched]:
                        unmatched.append((emp_str, store_str, ws.title, row[0].row))
                    continue
                    
                # Check for duplicates to prevent double-inserting
                existing_check = None
                if sale_number:
                    existing_check = conn.execute(
                        "SELECT id FROM layby_deductions WHERE employee_id=? AND sale_number=? AND sale_number != ''",
                        (emp_id, sale_number)
                    ).fetchone()
                else:
                    existing_check = conn.execute(
                        "SELECT id FROM layby_deductions WHERE employee_id=? AND description=? AND basket_total=? AND start_month=? AND start_year=?",
                        (emp_id, description, basket_total, start_month, start_year)
                    ).fetchone()
                    
                if existing_check:
                    skipped += 1
                    continue
                    
                # Math
                total_amount = round(basket_total * (1 - discount_pct / 100), 2)
                monthly_amount = round(total_amount / term, 2)
                balance_remaining = round(total_amount - (payments_made * monthly_amount), 2)
                if payments_made >= term or balance_remaining <= 0.01:
                    balance_remaining = 0
                    status = 'complete'
                else:
                    status = 'active'
                    
                # Insert layby plan
                cur = conn.execute('''
                    INSERT INTO layby_deductions_cents
                        (employee_id, sale_number, description, basket_total_cents, discount_pct,
                         total_amount_cents, monthly_amount_cents, balance_remaining_cents,
                         term_months, start_month, start_year, payments_made, status, notes)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    emp_id, sale_number, description, db.to_cents(basket_total), discount_pct,
                    db.to_cents(total_amount), db.to_cents(monthly_amount), db.to_cents(max(0, balance_remaining)),
                    term, start_month, start_year, payments_made, status, notes
                ))
                layby_id = cur.lastrowid
                
                # Insert standard item link in layby_items
                conn.execute('''
                    INSERT INTO layby_items_cents (layby_id, description, unit_price_cents, quantity, line_total_cents)
                    VALUES (?, ?, ?, 1, ?)
                ''', (layby_id, description, db.to_cents(basket_total), db.to_cents(basket_total)))
                
                inserted += 1
                log_msg = (f"✓ Lay-by Plan: {emp_str} ({store_str}) | Basket R{basket_total:.2f} ({discount_pct}% off) | "
                           f"Total: R{total_amount:.2f} | R{monthly_amount:.2f}/mo × {term} | Bal: R{balance_remaining:.2f}")
                import_logs.append(log_msg)
                
        conn.commit()
        conn.close()
        
        flash(f"Successfully processed {sheets_processed} sheets. Imported {inserted} lay-bys.", "success")
        
        results = {
            'inserted': inserted,
            'skipped': skipped,
            'unmatched': [{'sheet': u[2], 'row_num': u[3], 'name': f"{u[0]} ({u[1]})"} for u in unmatched],
            'created_employees': 0,
            'logs': import_logs
        }
        return render_template('import_center.html', results=results)
        
    except Exception as e:
        flash(f"Error processing workbook: {str(e)}", "danger")
        return redirect(url_for('import_center'))
