from flask import (render_template, request, redirect, url_for, flash,
                   jsonify, send_file, session)
from datetime import datetime
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from northwind.data import database as db
from northwind.core import app
from northwind.deductions import plans
from northwind.deductions.pagination import paginate


@app.route('/uniform/add', methods=['POST'])
def add_uniform():
    """Parse the form, then delegate to plans.create_uniform_plan.

    The validation / lock check / insert live in northwind/deductions/plans.py so the MCP
    connector creates plans through exactly the same code path.
    """
    emp_id = request.form['employee_id']
    try:
        total_form = request.form.get('total_amount')
        conn = db.get_db()
        try:
            with conn:
                plans.create_uniform_plan(
                    conn, emp_id,
                    monthly=request.form['monthly_amount'],
                    term=request.form['term_months'],
                    start_year=request.form['start_year'],
                    start_month=request.form['start_month'],
                    total=total_form if total_form else None,
                    sku=request.form.get('sku', ''),
                    description=request.form.get('description', ''),
                    sale_number=request.form.get('sale_number', ''),
                    notes=request.form.get('notes', ''),
                    actor=session.get('admin') or 'admin')
        finally:
            conn.close()
    except KeyError:
        flash('Please enter valid numbers for amount, term and start period.', 'danger')
        return redirect(url_for('employee_detail', emp_id=emp_id))
    except ValueError as exc:
        flash(str(exc), 'danger')
        return redirect(url_for('employee_detail', emp_id=emp_id))

    flash('Uniform deduction added.', 'success')
    return redirect(url_for('employee_detail', emp_id=emp_id))


@app.route('/uniform/<int:plan_id>/tick', methods=['POST'])
def tick_uniform(plan_id):
    conn = db.get_db()
    plan = conn.execute("SELECT * FROM uniform_deductions WHERE id=?", (plan_id,)).fetchone()
    if not plan:
        conn.close()
        if request.headers.get('Accept') == 'application/json':
            return jsonify({'success': False, 'message': 'Plan not found.'})
        flash('Plan not found.', 'danger')
        return redirect(request.referrer or url_for('employees'))

    if plan['payments_made'] >= plan['term_months'] or (plan['balance_remaining'] is not None and plan['balance_remaining'] <= 0.01):
        conn.close()
        if request.headers.get('Accept') == 'application/json':
            return jsonify({'success': False, 'message': 'This plan is already fully paid.'})
        flash('Plan is already fully paid.', 'warning')
        return redirect(request.referrer or url_for('employee_detail', emp_id=plan['employee_id']))

    # Calculate target period
    start_m = plan['start_month']
    start_y = plan['start_year']
    payments = plan['payments_made']
    target_m = start_m + payments
    target_y = start_y + (target_m - 1) // 12
    target_m = ((target_m - 1) % 12) + 1

    if db.is_period_locked(target_y, target_m):
        conn.close()
        if request.headers.get('Accept') == 'application/json':
            return jsonify({'success': False, 'message': 'The targeted payroll period is locked.'})
        flash('Cannot process payment. The targeted payroll period is locked.', 'danger')
        return redirect(request.referrer or url_for('employee_detail', emp_id=plan['employee_id']))

    # Prevent duplicate payments for the target month
    exists = conn.execute('''
        SELECT 1 FROM deduction_transactions 
        WHERE plan_type = 'uniform' AND plan_id = ? AND year = ? AND month = ? AND COALESCE(voided, 0) = 0
    ''', (plan_id, target_y, target_m)).fetchone()
    if exists:
        conn.close()
        if request.headers.get('Accept') == 'application/json':
            return jsonify({'success': False, 'message': f'A payment has already been allocated to {db.MONTH_FULL[target_m]} {target_y}.'})
        flash(f'A payment has already been allocated to {db.MONTH_FULL[target_m]} {target_y}.', 'warning')
        return redirect(request.referrer or url_for('employee_detail', emp_id=plan['employee_id']))

    # Calculate installment amount
    total_amt = plan['total_amount'] if plan['total_amount'] is not None else round(plan['term_months'] * plan['monthly_amount'], 2)
    monthly_rounded = round(plan['monthly_amount'], 2)
    term = plan['term_months']
    current_bal = plan['balance_remaining'] if plan['balance_remaining'] is not None else (total_amt - (payments * monthly_rounded))
    
    if payments >= term - 1:
        installment = current_bal
    else:
        installment = min(plan['monthly_amount'], current_bal)
    installment = round(max(0, installment), 2)

    new_balance = round(max(0, current_bal - installment), 2)
    new_count = payments + 1
    status = 'complete' if (new_count >= term or new_balance <= 0.01) else 'active'

    conn.execute("UPDATE uniform_deductions_cents SET payments_made=?, balance_remaining_cents=?, status=? WHERE id=?", (new_count, db.to_cents(max(0, new_balance)), status, plan_id))
    conn.execute('''
        INSERT INTO deduction_transactions_cents (plan_type, plan_id, employee_id, amount_cents, year, month)
        VALUES ('uniform', ?, ?, ?, ?, ?)
    ''', (plan_id, plan['employee_id'], db.to_cents(installment), target_y, target_m))
    conn.commit()
    conn.close()

    if request.headers.get('Accept') == 'application/json':
        outstanding = db.get_outstanding_summary(plan['employee_id'])
        cat_totals = db.get_category_totals(plan['employee_id'])
        
        plan_paid = round(total_amt - new_balance, 2)
        plan_remaining = new_balance

        return jsonify({
            'success': True,
            'new_payments_made': new_count,
            'status': status,
            'plan_remaining': round(plan_remaining, 2),
            'plan_paid': round(plan_paid, 2),
            'outstanding': outstanding,
            'cat_totals': cat_totals
        })
    return redirect(request.referrer or url_for('employee_detail', emp_id=plan['employee_id']))


@app.route('/uniform/<int:plan_id>/adjust', methods=['POST'])
def adjust_uniform(plan_id):
    """Record an ad-hoc payment and recalculate remaining monthly amount."""
    conn = db.get_db()
    try:
        with conn:
            result = plans.adjust_plan(
                conn, 'uniform', plan_id,
                amount=request.form['amount'],
                note=request.form.get('note', ''),
                new_term=request.form.get('new_term', '').strip() or None,
                actor=session.get('admin') or 'admin')
    except ValueError as exc:
        flash(str(exc), 'danger')
        return redirect(request.referrer or url_for('employees'))
    finally:
        conn.close()

    flash(f"Ad-hoc payment of R{result['amount']:.2f} recorded. "
          f"New monthly: R{result['new_monthly']:.2f}.", 'success')
    return redirect(url_for('employee_detail', emp_id=result['employee_id']))


@app.route('/uniform/<int:plan_id>/edit', methods=['POST'])
def edit_uniform(plan_id):
    """Parse the edit form, then delegate to plans.edit_plan.

    The validation, status derivation and UPDATE live in northwind/deductions/plans.py,
    shared with the MCP connector's edit_plan tool.
    """
    conn = db.get_db()
    try:
        with conn:
            result = plans.edit_plan(
                conn, 'uniform', plan_id,
                actor=session.get('admin_username') or 'admin',
                monthly=request.form.get('monthly_amount'),
                term_months=request.form.get('term_months'),
                total=request.form.get('total_amount'),
                payments_made=request.form.get('payments_made'),
                balance_remaining=request.form.get('balance_remaining'),
                start_month=request.form.get('start_month'),
                start_year=request.form.get('start_year'),
                sku=request.form.get('sku', ''),
                sale_number=request.form.get('sale_number', ''),
                description=request.form.get('description', ''),
                notes=request.form.get('notes', ''))
    except ValueError as exc:
        flash(str(exc), 'danger')
        return redirect(request.referrer or url_for('employees'))
    finally:
        conn.close()

    flash('Uniform plan updated.', 'success')
    return redirect(url_for('employee_detail', emp_id=result['employee_id']))


@app.route('/uniform/<int:plan_id>/write-off', methods=['POST'])
def write_off_uniform(plan_id):
    conn = db.get_db()
    try:
        with conn:
            result = plans.write_off_plan(
                conn, 'uniform', plan_id,
                reason=request.form.get('reason'),
                actor=session.get('admin') or 'admin')
    except ValueError as exc:
        flash(str(exc), 'danger')
        return redirect(request.referrer or url_for('employees'))
    finally:
        conn.close()

    flash('Uniform plan written off.', 'warning')
    return redirect(url_for('employee_detail', emp_id=result['employee_id']))


# ── Lay-by Deductions ───────────────────────────────────────────────────────


@app.route('/uniforms')
def uniforms_list():
    conn = db.get_db()
    store_filter = request.args.get('store', '')
    status_filter = request.args.get('status', 'active')
    search = request.args.get('q', '').strip()
    if status_filter not in ('active', 'complete', 'written_off', 'all'):
        status_filter = 'active'
    query = '''
        SELECT u.*, e.full_name, e.current_store, e.job_title
        FROM uniform_deductions u
        JOIN employees e ON e.id = u.employee_id
        WHERE 1=1
    '''
    params = []
    if status_filter != 'all':
        query += ' AND u.status = ?'
        params.append(status_filter)
    if store_filter:
        query += ' AND e.current_store = ?'
        params.append(store_filter)
    if search:
        # The name box used to filter only the rows already in the DOM. Now the
        # list is windowed, a match on page 4 has to be found by the server or
        # the search would report "no results" for a plan that exists.
        query += ' AND e.full_name LIKE ?'
        params.append(f'%{search}%')
    query += ' ORDER BY e.current_store, e.full_name'
    rows = conn.execute(query, params).fetchall()
    stores = [r[0] for r in conn.execute(
        "SELECT DISTINCT e.current_store FROM uniform_deductions u "
        "JOIN employees e ON e.id=u.employee_id WHERE u.status='active' ORDER BY e.current_store"
    ).fetchall()]

    # Group plans by employee
    grouped = {}
    total_active_plans = 0
    for r in rows:
        emp_id = r['employee_id']
        total_amt = r['total_amount'] if r['total_amount'] is not None else round(r['term_months'] * r['monthly_amount'], 2)
        pmt_made = r['payments_made']
        monthly_rounded = round(r['monthly_amount'], 2)
        
        if pmt_made >= r['term_months']:
            balance = 0.0
        else:
            balance = round(total_amt - (pmt_made * monthly_rounded), 2)
            
        if emp_id not in grouped:
            grouped[emp_id] = {
                'employee_id': emp_id,
                'full_name': r['full_name'],
                'current_store': r['current_store'],
                'job_title': r['job_title'],
                'monthly_amount': 0.0,
                'term_months': 0,
                'payments_made': 0,
                'payments_left': 0,
                'balance': 0.0,
                'plan_count': 0,
            }
        
        grouped[emp_id]['monthly_amount'] += r['monthly_amount']
        grouped[emp_id]['term_months'] += r['term_months']
        grouped[emp_id]['payments_made'] += r['payments_made']
        grouped[emp_id]['payments_left'] += (r['term_months'] - r['payments_made'])
        grouped[emp_id]['balance'] += balance
        grouped[emp_id]['plan_count'] += 1
        total_active_plans += 1

    plans = []
    for emp_id, data in grouped.items():
        data['monthly_amount'] = round(data['monthly_amount'], 2)
        data['balance'] = round(data['balance'], 2)
        label = 'active' if status_filter == 'active' else status_filter.replace('_', ' ') if status_filter != 'all' else ''
        data['description'] = f"{data['plan_count']} {label} plan{'s' if data['plan_count'] != 1 else ''}".replace('  ', ' ')
        plans.append(data)

    plans.sort(key=lambda x: (x['current_store'], x['full_name']))

    now = datetime.now()
    all_employees = [dict(r) for r in conn.execute(
        "SELECT id, full_name, current_store FROM employees WHERE status='active' AND sector='retail' ORDER BY current_store, full_name"
    ).fetchall()]
    conn.close()
    # Every figure the page shows outside the table body is a whole-filtered-set
    # figure, computed before the window is taken. A page-only "Monthly Total"
    # would understate what payroll is about to deduct.
    totals = {
        'employees': len(plans),
        'monthly': sum(p['monthly_amount'] for p in plans),
        'outstanding': sum(p['balance'] for p in plans),
    }
    plans, pager = paginate(plans, noun='employees')
    return render_template('uniforms.html', plans=plans, stores=stores,
                           store_filter=store_filter, status_filter=status_filter,
                           search=search,
                           total_active_plans=total_active_plans,
                           total_outstanding=totals['outstanding'],
                           totals=totals, pager=pager,
                           all_employees=all_employees,
                           now=now, MONTH_NAMES=db.MONTH_NAMES)


# ── Uniform Plans Export ────────────────────────────────────────────────────


@app.route('/uniforms/export')
def export_uniforms():
    conn = db.get_db()
    store_filter = request.args.get('store', '').strip()
    from_month   = request.args.get('from_month', type=int)
    from_year    = request.args.get('from_year',  type=int)
    to_month     = request.args.get('to_month',   type=int)
    to_year      = request.args.get('to_year',    type=int)

    query = '''
        SELECT u.*, e.full_name, e.current_store, e.job_title, e.id as emp_id
        FROM uniform_deductions u
        JOIN employees e ON e.id = u.employee_id
        WHERE u.status = 'active'
          AND e.status = 'active'
          AND u.payments_made < u.term_months
          AND (u.balance_remaining IS NULL OR u.balance_remaining > 0.01)
    '''
    params = []
    if store_filter:
        query += ' AND e.current_store = ?'
        params.append(store_filter)
    if from_month and from_year:
        query += ' AND (u.start_year * 12 + u.start_month) >= ?'
        params.append(from_year * 12 + from_month)
    if to_month and to_year:
        query += ' AND (u.start_year * 12 + u.start_month) <= ?'
        params.append(to_year * 12 + to_month)
    query += ' ORDER BY e.current_store, e.full_name, u.id'
    rows = conn.execute(query, params).fetchall()
    conn.close()

    # ── Styles ──────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    navy    = PatternFill('solid', start_color='1A2B4A')
    amber   = PatternFill('solid', start_color='B45309')
    alt_row = PatternFill('solid', start_color='F0F4FF')
    store_hdr_fill = PatternFill('solid', start_color='1E3A5F')
    white_bold = Font(bold=True, color='FFFFFF', size=11)
    store_font = Font(bold=True, color='BDD7FF', size=10, italic=True)
    thin = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )
    cur = 'R#,##0.00;-R#,##0.00;"-"'

    def col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 1: Uniform Plans ───────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Uniform Plans'
    headers = [
        'Employee ID', 'Employee Name', 'Store', 'Job Title',
        'SKU', 'Sale Number', 'Description',
        'Start', 'Term (mo)', 'Regular Monthly (R)', 'Total (R)',
        'Paid (R)', 'Remaining (R)', 'Payments Made', 'Status',
    ]
    widths = [12, 28, 20, 22, 14, 14, 38, 12, 10, 18, 14, 14, 16, 14, 12]
    col_widths(ws, widths)
    ws.freeze_panes = 'A3'

    # Title row
    ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
    title_cell = ws['A1']
    from_label = f'{db.MONTH_NAMES[from_month]} {from_year}' if from_month and from_year else ''
    to_label   = f'{db.MONTH_NAMES[to_month]} {to_year}' if to_month and to_year else ''
    date_range = f' · {from_label} – {to_label}' if from_label else ''
    store_label = f' · {store_filter}' if store_filter else ' · All Stores'
    title_cell.value = f'NORTHWIND Active Staff Uniform Plans{store_label}{date_range}'
    title_cell.font = Font(bold=True, color='FFFFFF', size=13)
    title_cell.fill = navy
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24

    # Header row
    for col, (h, _) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = white_bold
        c.fill = amber
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = thin
    ws.row_dimensions[2].height = 18

    data_row = 3
    last_store = None
    alt = False
    store_totals = {}  # store -> {monthly, total, remaining, count}

    for r in rows:
        store = r['current_store']
        total_amt = r['total_amount'] if r['total_amount'] is not None else round(r['term_months'] * r['monthly_amount'], 2)
        pmt = r['payments_made']
        monthly_r = round(r['monthly_amount'], 2)
        if pmt >= r['term_months']:
            paid_val = total_amt
            rem_val  = 0.0
        else:
            paid_val = pmt * monthly_r
            rem_val  = total_amt - paid_val

        # Store group header
        if store != last_store:
            last_store = store
            alt = False
            ws.merge_cells(f'A{data_row}:{get_column_letter(len(headers))}{data_row}')
            sc = ws.cell(row=data_row, column=1, value=f'  🏪  {store}')
            sc.font = store_font
            sc.fill = store_hdr_fill
            sc.alignment = Alignment(horizontal='left', vertical='center')
            ws.row_dimensions[data_row].height = 16
            data_row += 1
            # init store totals
            if store not in store_totals:
                store_totals[store] = {'monthly': 0, 'total': 0, 'remaining': 0, 'count': 0}

        fill = alt_row if alt else None
        alt = not alt

        row_vals = [
            r['emp_id'], r['full_name'], store, r['job_title'] or '',
            r['sku'] or '', r['sale_number'] or '', r['description'] or '',
            f"{db.MONTH_NAMES[r['start_month']]} {r['start_year']}",
            r['term_months'], monthly_r, total_amt,
            round(paid_val, 2), round(rem_val, 2), pmt, r['status'],
        ]
        for col, val in enumerate(map(db.xl_safe, row_vals), 1):
            c = ws.cell(row=data_row, column=col, value=val)
            c.border = thin
            c.alignment = Alignment(vertical='center')
            if fill:
                c.fill = fill
            if col == 10:   # Monthly
                c.number_format = cur; c.alignment = Alignment(horizontal='right', vertical='center')
            elif col == 11:  # Total
                c.number_format = cur; c.alignment = Alignment(horizontal='right', vertical='center')
            elif col == 12:  # Paid
                c.number_format = cur; c.alignment = Alignment(horizontal='right', vertical='center')
            elif col == 13:  # Remaining
                c.number_format = cur; c.alignment = Alignment(horizontal='right', vertical='center')
                c.font = Font(bold=True, color='C84B11' if rem_val > 0 else '059669')

        store_totals[store]['monthly']   += monthly_r
        store_totals[store]['total']     += total_amt
        store_totals[store]['remaining'] += rem_val
        store_totals[store]['count']     += 1
        data_row += 1

    # Grand total row
    data_row += 1
    total_monthly   = sum(s['monthly']   for s in store_totals.values())
    total_total     = sum(s['total']     for s in store_totals.values())
    total_remaining = sum(s['remaining'] for s in store_totals.values())
    total_count     = sum(s['count']     for s in store_totals.values())
    gt_vals = ['', f'GRAND TOTAL  ({total_count} plans)', '', '', '', '', '', '', '',
               total_monthly, total_total, '', total_remaining, '', '']
    for col, val in enumerate(map(db.xl_safe, gt_vals), 1):
        c = ws.cell(row=data_row, column=col, value=val)
        c.font = Font(bold=True, color='FFFFFF', size=11)
        c.fill = navy
        c.border = thin
        if col == 10:
            c.number_format = cur; c.alignment = Alignment(horizontal='right', vertical='center')
        elif col == 11:
            c.number_format = cur; c.alignment = Alignment(horizontal='right', vertical='center')
        elif col == 13:
            c.number_format = cur; c.alignment = Alignment(horizontal='right', vertical='center')

    # ── Sheet 2: Summary by Store ────────────────────────────────────────────
    ws2 = wb.create_sheet('Summary by Store')
    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 10
    ws2.column_dimensions['C'].width = 22
    ws2.column_dimensions['D'].width = 16
    ws2.column_dimensions['E'].width = 16

    ws2.merge_cells('A1:E1')
    t2 = ws2['A1']
    t2.value = f'Uniform Plans — Summary by Store{store_label}{date_range}'
    t2.font = Font(bold=True, color='FFFFFF', size=12)
    t2.fill = navy
    t2.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 22

    s2_hdrs = ['Store', '# Plans', 'Regular Monthly (R)', 'Gross Total (R)', 'Remaining (R)']
    for col, h in enumerate(s2_hdrs, 1):
        c = ws2.cell(row=2, column=col, value=h)
        c.font = white_bold; c.fill = amber
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin
    ws2.row_dimensions[2].height = 18

    for i, (store, s) in enumerate(sorted(store_totals.items()), 3):
        row_data = [store, s['count'], round(s['monthly'], 2), round(s['total'], 2), round(s['remaining'], 2)]
        fill = alt_row if i % 2 == 1 else None
        for col, val in enumerate(map(db.xl_safe, row_data), 1):
            c = ws2.cell(row=i, column=col, value=val)
            c.border = thin
            c.alignment = Alignment(vertical='center')
            if fill:
                c.fill = fill
            if col in (3, 4, 5):
                c.number_format = cur; c.alignment = Alignment(horizontal='right', vertical='center')

    last_summary_row = 2 + len(store_totals) + 1
    gt_s = [f'TOTAL ({len(store_totals)} stores)', total_count, round(total_monthly, 2), round(total_total, 2), round(total_remaining, 2)]
    for col, val in enumerate(map(db.xl_safe, gt_s), 1):
        c = ws2.cell(row=last_summary_row, column=col, value=val)
        c.font = Font(bold=True, color='FFFFFF', size=11)
        c.fill = navy; c.border = thin
        if col in (3, 4, 5):
            c.number_format = cur; c.alignment = Alignment(horizontal='right', vertical='center')

    ws2.freeze_panes = 'A3'

    out = io.BytesIO()
    wb.save(out); out.seek(0)
    now = datetime.now()
    store_slug = f'_{store_filter.replace(" ", "_")}' if store_filter else '_All_Stores'
    date_slug  = f'_{from_year}{from_month:02d}-{to_year}{to_month:02d}' if from_month and from_year else ''
    filename   = f'Uniform_Plans{store_slug}{date_slug}_{now.strftime("%Y%m%d")}.xlsx'
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


# ── Lay-by Plans Export ─────────────────────────────────────────────────────
