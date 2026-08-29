from flask import (render_template, request, redirect, url_for, flash,
                   jsonify, send_file, session)
from datetime import datetime
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from northwind.data import database as db
from northwind.core import app
from northwind.deductions import plans
from northwind.deductions.pagination import paginate


@app.route('/undercharge/add', methods=['POST'])
def add_undercharge():
    """Parse the form, then delegate to plans.create_undercharge.

    Validation, the overcharge special-casing, the lock check and the recovery
    schedule live in northwind/deductions/plans.py, shared with the MCP connector.
    """
    emp_id = request.form['employee_id']
    uc_type = request.form.get('type', 'undercharge')
    conn = db.get_db()
    try:
        with conn:
            plans.create_undercharge(
                conn, emp_id,
                total=request.form['total_amount'],
                incident_year=request.form['incident_year'],
                incident_month=request.form['incident_month'],
                recovery=request.form.get('recovery_method', 'full'),
                split_months=request.form.get('split_months', 1),
                start_year=request.form.get('start_year'),
                start_month=request.form.get('start_month'),
                uc_type=uc_type,
                reason=request.form.get('reason', ''),
                sale_number=request.form.get('sale_number', ''),
                notes=request.form.get('notes', ''),
                actor=request.form.get('actor') or session.get('admin') or 'admin')
    except KeyError:
        flash('Please enter valid numbers for amount, split and dates.', 'danger')
        return redirect(url_for('employee_detail', emp_id=emp_id))
    except ValueError as exc:
        flash(str(exc), 'danger')
        return redirect(url_for('employee_detail', emp_id=emp_id))
    finally:
        conn.close()

    label = 'Overcharge' if uc_type == 'overcharge' else 'Undercharge'
    flash(f'{label} added.', 'success')
    return redirect(url_for('employee_detail', emp_id=emp_id))


@app.route('/undercharge/<int:uc_id>/tick', methods=['POST'])
def tick_undercharge(uc_id):
    conn = db.get_db()
    uc = conn.execute("SELECT * FROM undercharges WHERE id=?", (uc_id,)).fetchone()
    if not uc:
        conn.close()
        if request.headers.get('Accept') == 'application/json':
            return jsonify({'success': False, 'message': 'Deduction not found.'})
        flash('Deduction not found.', 'danger')
        return redirect(request.referrer or url_for('employees'))

    if uc['status'] in ('written_off', 'accounted_for'):
        conn.close()
        if request.headers.get('Accept') == 'application/json':
            return jsonify({'success': False, 'message': 'This deduction is already complete.'})
        flash('Deduction is already complete.', 'warning')
        return redirect(request.referrer or url_for('employee_detail', emp_id=uc['employee_id']))

    if uc['type'] == 'overcharge':
        target_y = uc['incident_year']
        target_m = uc['incident_month']
    else:
        db.ensure_undercharge_schedule(conn, uc_id)
        item = conn.execute(
            "SELECT due_year,due_month,amount_cents FROM undercharge_schedule_items "
            "WHERE undercharge_id=? AND state='scheduled' AND transaction_id IS NULL "
            "ORDER BY due_year,due_month,id LIMIT 1", (uc_id,)).fetchone()
        if not item:
            conn.close()
            message = 'There is no outstanding scheduled installment.'
            if request.headers.get('Accept') == 'application/json':
                return jsonify({'success': False, 'message': message})
            flash(message, 'warning')
            return redirect(request.referrer or url_for('employee_detail', emp_id=uc['employee_id']))
        target_y, target_m = item['due_year'], item['due_month']

    if db.is_period_locked(target_y, target_m):
        conn.close()
        if request.headers.get('Accept') == 'application/json':
            return jsonify({'success': False, 'message': 'The targeted payroll period is locked.'})
        flash('Cannot process payment. The targeted payroll period is locked.', 'danger')
        return redirect(request.referrer or url_for('employee_detail', emp_id=uc['employee_id']))

    if uc['type'] == 'overcharge':
        with conn:
            conn.execute("UPDATE undercharges_cents SET status='accounted_for', payments_made=1 WHERE id=?", (uc_id,))
        status = 'accounted_for'
    else:
        with conn:
            db.tick_undercharges_due(conn, uc['employee_id'], target_y, target_m)
        account = db.get_undercharge_account(uc_id)
        status = account['status']
    conn.close()

    if request.headers.get('Accept') == 'application/json':
        outstanding = db.get_outstanding_summary(uc['employee_id'])
        cat_totals = db.get_category_totals(uc['employee_id'])
        if uc['type'] == 'overcharge':
            plan_remaining = uc['total_amount'] if status == 'pending' else 0
            plan_paid = 0
        else:
            plan_remaining = db.to_rands(account['remaining_cents'])
            plan_paid = db.to_rands(account['net_employee_paid_cents'])
        return jsonify({
            'success': True,
            'new_payments_made': 1 if uc['type'] == 'overcharge' else account['payment_count'],
            'status': status,
            'plan_remaining': round(plan_remaining, 2),
            'plan_paid': round(plan_paid, 2),
            'outstanding': outstanding,
            'cat_totals': cat_totals
        })
    return redirect(request.referrer or url_for('employee_detail', emp_id=uc['employee_id']))


@app.route('/undercharge/<int:uc_id>/write-off', methods=['POST'])
def write_off_undercharge(uc_id):
    conn = db.get_db()
    try:
        with conn:
            result = plans.write_off_plan(
                conn, 'undercharge', uc_id,
                reason=request.form.get('reason'),
                actor=session.get('admin') or 'admin')
    except ValueError as exc:
        flash(str(exc), 'danger')
        return redirect(request.referrer or url_for('employees'))
    finally:
        conn.close()

    flash('Undercharge written off.', 'warning')
    return redirect(url_for('employee_detail', emp_id=result['employee_id']))


@app.route('/undercharge/<int:uc_id>/customer-paid', methods=['POST'])
def customer_paid_undercharge(uc_id):
    def respond(success, message=None, extra=None, emp_id=None):
        if request.headers.get('Accept') == 'application/json':
            payload = {'success': success}
            if message:
                payload['message'] = message
            if extra:
                payload.update(extra)
            return jsonify(payload)
        flash(message or ('Saved.' if success else 'Something went wrong.'),
              'success' if success else 'danger')
        return redirect(request.referrer or
                        (url_for('employee_detail', emp_id=emp_id) if emp_id else url_for('employees')))

    conn = db.get_db()
    try:
        uc = conn.execute("SELECT * FROM undercharges WHERE id=?", (uc_id,)).fetchone()
        if not uc:
            return respond(False, 'Undercharge not found.')
        if (uc['type'] or 'undercharge') != 'undercharge':
            return respond(False, 'Customer settlements apply only to undercharges.',
                           emp_id=uc['employee_id'])
        emp_id = uc['employee_id']
        employee = conn.execute(
            "SELECT status FROM employees WHERE id=?", (emp_id,)).fetchone()
        actor = session.get('admin') or 'admin'
        db.ensure_undercharge_schedule(conn, uc_id)
        before = db.get_undercharge_account(uc_id, conn)

        try:
            customer_amount_cents = db.to_cents(
                request.form.get('customer_amount') or
                db.to_rands(before['adjusted_total_cents']))
        except (TypeError, ValueError):
            return respond(False, 'Enter a valid customer payment amount.', emp_id=emp_id)
        if customer_amount_cents <= 0:
            return respond(False, 'Customer payment must be greater than zero.', emp_id=emp_id)

        # Legacy count-only rows are surfaced as an assumed paid amount for the
        # existing workflow. Real rows always use the immutable transaction sum.
        deducted_cents = before['payroll_deducted_cents']
        if deducted_cents <= 0 and (uc['payments_made'] or 0) > 0:
            deducted_cents = round(
                uc['payments_made'] * db.to_cents(uc['total_amount'])
                / max(uc['split_months'] or 1, 1))

        recovery_target = max(
            before['adjusted_total_cents'] - before['written_off_cents'], 0)
        potential_refund = max(
            customer_amount_cents + deducted_cents - recovery_target, 0)
        refund = (request.form.get('refund') or '').strip().lower()
        if refund not in ('yes', 'no'):
            refund = 'yes' if potential_refund > 0 else 'no'
        if potential_refund <= 0:
            refund = 'no'
        refund_method = (request.form.get('refund_method') or 'payroll').strip().lower()
        if refund_method not in ('payroll', 'external'):
            refund_method = 'payroll'
        valid_refund_month = None
        if refund == 'yes' and refund_method == 'payroll':
            if employee and employee['status'] != 'active':
                return respond(False, 'This employee is not active on payroll. '
                               'Record the reimbursement as an external payment.',
                               emp_id=emp_id)
            valid_refund_month = db.validate_month_year(
                request.form.get('reimburse_year'),
                request.form.get('reimburse_month'))
            if valid_refund_month is None:
                return respond(False, 'A valid reimbursement month is required — '
                               'money was already deducted and must be paid back.',
                               emp_id=emp_id)
            ry, rm = valid_refund_month
            if db.is_period_locked(ry, rm, db.get_employee_sector(emp_id, conn)):
                return respond(False, f'{db.MONTH_FULL[rm]} {ry} payroll is locked — '
                               'choose a different reimbursement month.', emp_id=emp_id)
            clash = conn.execute(
                "SELECT 1 FROM deduction_transactions_cents WHERE plan_type='undercharge' "
                "AND plan_id=? AND year=? AND month=? AND COALESCE(voided,0)=0",
                (uc_id, ry, rm)).fetchone()
            if clash:
                return respond(False, f'A deduction for this undercharge already sits in '
                               f'{db.MONTH_FULL[rm]} {ry}. Pick a later month.',
                               emp_id=emp_id)

        future_before = conn.execute(
            "SELECT due_year,due_month FROM undercharge_schedule_items "
            "WHERE undercharge_id=? AND amount_cents>0 AND state='scheduled' "
            "AND transaction_id IS NULL ORDER BY due_year,due_month,id",
            (uc_id,)).fetchall()
        with conn:
            db.record_undercharge_event(
                conn, uc_id, 'customer_payment', customer_amount_cents,
                note=request.form.get('customer_payment_note') or
                     'Customer payment recorded',
                actor=actor)
            # A customer settlement always invalidates the old future amounts.
            conn.execute(
                "UPDATE undercharge_schedule_items SET state='cancelled',"
                "state_reason='Customer payment changed employee liability',"
                "state_changed_at=datetime('now') WHERE undercharge_id=? "
                "AND state='scheduled' AND transaction_id IS NULL AND amount_cents>0",
                (uc_id,))
            after_payment = db.get_undercharge_account(uc_id, conn)

            reimburse_year = reimburse_month = None
            refund_due = max(
                after_payment['refund_due_cents'],
                customer_amount_cents + deducted_cents
                - max(after_payment['adjusted_total_cents']
                      - after_payment['written_off_cents'], 0))
            if refund == 'yes' and refund_due > 0:
                if refund_method == 'external':
                    db.record_undercharge_event(
                        conn, uc_id, 'external_refund', refund_due,
                        note=request.form.get('external_refund_reference') or
                             'Employee reimbursed outside payroll',
                        actor=actor)
                else:
                    reimburse_year, reimburse_month = valid_refund_month
                    db.create_undercharge_schedule(
                        conn, uc_id, reimburse_year, reimburse_month, 1, refund_due,
                        kind='refund', reason='Refund after customer payment',
                        actor=actor)
            elif refund == 'no' and refund_due > 0:
                db.record_undercharge_event(
                    conn, uc_id, 'refund_waiver', refund_due,
                    note=request.form.get('refund_waiver_reason') or
                         'Employee refund explicitly declined',
                    actor=actor)

            # A partial customer payment leaves employee liability. Rebuild the
            # remaining schedule using the requested values or the old future
            # schedule shape.
            current = db.get_undercharge_account(uc_id, conn)
            if current['remaining_cents'] > 0:
                old_future = future_before
                start_year = int(request.form.get('start_year') or
                                 (old_future[0]['due_year'] if old_future else
                                  (uc['start_year'] or uc['incident_year'])))
                start_month = int(request.form.get('start_month') or
                                  (old_future[0]['due_month'] if old_future else
                                   (uc['start_month'] or uc['incident_month'])))
                months = int(request.form.get('split_months') or
                             max(len(old_future), 1))
                db.create_undercharge_schedule(
                    conn, uc_id, start_year, start_month, months,
                    current['remaining_cents'], kind='deduction',
                    reason='Remaining balance after customer part-payment',
                    actor=actor)

            db.sync_undercharge_state(conn, uc_id)
            final = db.get_undercharge_account(uc_id, conn)
            legacy_status = (
                'partial' if final['remaining_cents'] > 0 else
                'paid_by_customer')
            conn.execute(
                "UPDATE undercharges_cents SET status=?,reimburse_month=?,"
                "reimburse_year=? WHERE id=?",
                (legacy_status, reimburse_month, reimburse_year, uc_id))
    except ValueError as exc:
        conn.rollback()
        return respond(False, str(exc), emp_id=uc['employee_id'] if uc else None)
    finally:
        conn.close()

    outstanding = db.get_outstanding_summary(emp_id)
    cat_totals = db.get_category_totals(emp_id)
    if final['remaining_cents'] > 0:
        msg = (f'Customer payment recorded. R '
               f'{db.to_rands(final["remaining_cents"]):,.2f} remains scheduled.')
    elif refund == 'yes' and refund_method == 'external':
        msg = 'Customer payment and external employee reimbursement recorded.'
    elif final['scheduled_refunds_cents'] > 0:
        msg = f'Reimbursement scheduled for {db.MONTH_FULL[reimburse_month]} {reimburse_year}.'
    elif deducted_cents == 0:
        msg = 'Customer payment recorded — no employee reimbursement is needed.'
    else:
        msg = 'Customer payment recorded and the employee refund decision was saved.'
    return respond(True, msg, emp_id=emp_id, extra={
        'status': final['status'],
        'plan_remaining': db.to_rands(final['remaining_cents']),
        'plan_paid': db.to_rands(final['net_employee_paid_cents']),
        'outstanding': outstanding,
        'cat_totals': cat_totals,
        'reimburse_month': reimburse_month,
        'reimburse_year': reimburse_year,
        'no_reimbursement': final['scheduled_refunds_cents'] == 0,
        'nothing_deducted': deducted_cents == 0,
    })


@app.route('/undercharge/<int:uc_id>/revert', methods=['POST'])
def revert_undercharge(uc_id):
    conn = db.get_db()
    try:
        uc = conn.execute("SELECT * FROM undercharges WHERE id=?", (uc_id,)).fetchone()
        if not uc:
            flash('Undercharge not found.', 'danger')
            return redirect(url_for('employees'))

        # A 'reimbursed' row carries a negative payback transaction in the
        # ledger; reverting the status must void it too, otherwise the ledger
        # says the employee was paid back while the plan says they weren't.
        reimb_txns = conn.execute(
            "SELECT id, year, month FROM deduction_transactions "
            "WHERE plan_type='undercharge' AND plan_id=? AND amount < 0 AND COALESCE(voided,0)=0",
            (uc_id,)).fetchall()
        sector = db.get_employee_sector(uc['employee_id'], conn)
        for t in reimb_txns:
            if db.is_period_locked(t['year'], t['month'], sector):
                flash(f'Cannot revert: the reimbursement sits in {db.MONTH_FULL[t["month"]]} '
                      f'{t["year"]}, which is locked.', 'danger')
                return redirect(url_for('employee_detail', emp_id=uc['employee_id']))

        # Revert paid_by_customer or reimbursed to pending/partial based on payments_made
        if uc['payments_made'] == 0:
            new_status = 'pending'
        elif uc['payments_made'] < uc['split_months']:
            new_status = 'partial'
        else:
            new_status = 'recovered'

        with conn:
            for t in reimb_txns:
                conn.execute("UPDATE deduction_transactions_cents SET voided=1 WHERE id=?", (t['id'],))
                conn.execute(
                    "UPDATE undercharge_schedule_items SET transaction_id=NULL,"
                    "state='cancelled',state_reason='Customer settlement reverted',"
                    "state_changed_at=datetime('now') WHERE transaction_id=?",
                    (t['id'],))
            customer_event = conn.execute(
                "SELECT e.* FROM undercharge_events e "
                "WHERE e.undercharge_id=? AND e.event_type='customer_payment' "
                "AND NOT EXISTS (SELECT 1 FROM undercharge_events r "
                "WHERE r.reverses_event_id=e.id) ORDER BY e.id DESC LIMIT 1",
                (uc_id,)).fetchone()
            if customer_event:
                db.record_undercharge_event(
                    conn, uc_id, 'customer_payment_reversal',
                    customer_event['amount_cents'],
                    note='Customer settlement reverted',
                    actor=session.get('admin') or 'admin',
                    reverses_event_id=customer_event['id'])
                account = db.get_undercharge_account(uc_id, conn)
                if account['remaining_cents'] > 0:
                    old = conn.execute(
                        "SELECT due_year,due_month FROM undercharge_schedule_items "
                        "WHERE undercharge_id=? AND amount_cents>0 "
                        "ORDER BY due_year,due_month,id LIMIT 1", (uc_id,)).fetchone()
                    sy = old['due_year'] if old else (uc['start_year'] or uc['incident_year'])
                    sm = old['due_month'] if old else (uc['start_month'] or uc['incident_month'])
                    sector = db.get_employee_sector(uc['employee_id'], conn)
                    while db.is_period_locked(sy, sm, sector):
                        sy, sm = db._uc_add_month(sy, sm, 1)
                    db.create_undercharge_schedule(
                        conn, uc_id, sy, sm, 1, account['remaining_cents'],
                        kind='deduction', reason='Customer settlement reverted',
                        actor=session.get('admin') or 'admin')
                db.sync_undercharge_state(conn, uc_id)
            else:
                conn.execute("UPDATE undercharges_cents SET status=? WHERE id=?", (new_status, uc_id))

        flash('Customer payment / reimbursement reverted successfully.', 'success')
        return redirect(url_for('employee_detail', emp_id=uc['employee_id']))
    finally:
        conn.close()


@app.route('/undercharge/<int:uc_id>/reschedule', methods=['POST'])
def reschedule_undercharge(uc_id):
    """Replace only unpaid future installments with an exact new schedule."""
    conn = db.get_db()
    try:
        with conn:
            result = plans.reschedule_undercharge(
                conn, uc_id,
                start_year=request.form.get('start_year'),
                start_month=request.form.get('start_month'),
                months=request.form.get('months'),
                reason=request.form.get('reason'),
                actor=session.get('admin') or 'admin')
        formatted = ', '.join(f"R {i['amount']:,.2f}" for i in result['schedule'])
        flash(f"Remaining R {result['rescheduled']:,.2f} rescheduled over "
              f"{result['months']} month(s): {formatted}.", 'success')
    except ValueError as exc:
        conn.rollback()
        flash(str(exc), 'danger')
    finally:
        conn.close()
    return redirect(request.referrer or url_for('undercharges_list'))


@app.route('/undercharge/<int:uc_id>/timeline')
def undercharge_timeline(uc_id):
    conn = db.get_db()
    try:
        uc = conn.execute(
            "SELECT u.*,e.full_name FROM undercharges u "
            "JOIN employees e ON e.id=u.employee_id WHERE u.id=?",
            (uc_id,)).fetchone()
        if not uc:
            flash('Undercharge not found.', 'danger')
            return redirect(url_for('undercharges_list'))
        account = db.get_undercharge_account(uc_id, conn)
        timeline = db.get_undercharge_timeline(uc_id, conn)
        return render_template(
            'undercharge_timeline.html', uc=dict(uc), account=account,
            timeline=timeline, MONTH_NAMES=db.MONTH_NAMES)
    finally:
        conn.close()


@app.route('/undercharge/<int:uc_id>/edit', methods=['POST'])
def edit_undercharge(uc_id):
    conn = db.get_db()
    uc = conn.execute("SELECT * FROM undercharges WHERE id=?", (uc_id,)).fetchone()
    if not uc:
        conn.close()
        flash('Undercharge/Overcharge not found.', 'danger')
        return redirect(url_for('employees'))

    uc_type = request.form.get('type', 'undercharge')
    reason = request.form.get('reason', '').strip()
    sale_number = request.form.get('sale_number', '').strip()
    try:
        total_amount = float(request.form.get('total_amount', 0))
        incident_month = int(request.form.get('incident_month', 1))
        incident_year = int(request.form.get('incident_year', datetime.now().year))
        payments_made_form = int(request.form.get('payments_made', uc['payments_made'] or 0))
    except (ValueError, TypeError):
        conn.close()
        flash('Please enter valid numbers for amount, dates and payments.', 'danger')
        return redirect(request.referrer or url_for('employee_detail', emp_id=uc['employee_id']))
    if total_amount <= 0:
        conn.close()
        flash('Amount must be greater than zero.', 'danger')
        return redirect(request.referrer or url_for('employee_detail', emp_id=uc['employee_id']))
    notes = request.form.get('notes', '').strip()
    status = request.form.get('status', uc['status'] or 'pending')

    if uc_type == 'overcharge':
        recovery_method = 'full'
        split_months = 1
        start_month = incident_month
        start_year = incident_year
        reimburse_month = None
        reimburse_year = None
    else:
        recovery_method = request.form.get('recovery_method', 'full')
        split_months = int(request.form.get('split_months', 1)) if recovery_method == 'split' else 1
        start_month = int(request.form.get('start_month', incident_month))
        start_year = int(request.form.get('start_year', incident_year))
        
        # Check starting period lock if period is modified
        if start_month != uc['start_month'] or start_year != uc['start_year']:
            if db.is_period_locked(start_year, start_month):
                conn.close()
                flash('Cannot change to a locked payroll start period.', 'danger')
                return redirect(request.referrer or url_for('employee_detail', emp_id=uc['employee_id']))

        reimburse_m_val = request.form.get('reimburse_month')
        reimburse_y_val = request.form.get('reimburse_year')

        # Whether the employee is being paid back. 'reimbursed' always implies a
        # payback; 'paid_by_customer' honours the explicit Yes/No refund choice.
        # Safety net: when NO explicit choice is submitted (a non-UI / legacy
        # post) and money was already deducted, we still require a refund month
        # so a payback can't silently vanish — only an explicit "No refund"
        # from the modal skips it.
        refund = (request.form.get('refund') or '').strip().lower()
        if status == 'reimbursed':
            wants_refund = True
        elif status == 'paid_by_customer':
            wants_refund = (refund == 'yes') if refund in ('yes', 'no') \
                else payments_made_form > 0
        else:
            wants_refund = False

        if wants_refund:
            valid = db.validate_month_year(reimburse_y_val, reimburse_m_val)
            if valid is None:
                conn.close()
                flash('A refund month is required when refunding the employee.', 'danger')
                return redirect(request.referrer or url_for('employee_detail', emp_id=uc['employee_id']))
            reimburse_year, reimburse_month = valid
            if reimburse_month != uc['reimburse_month'] or reimburse_year != uc['reimburse_year']:
                if db.is_period_locked(reimburse_year, reimburse_month):
                    conn.close()
                    flash('Cannot assign reimbursement to a locked payroll period.', 'danger')
                    return redirect(request.referrer or url_for('employee_detail', emp_id=uc['employee_id']))
        else:
            # No refund (incl. "Customer Paid" with the No-refund choice) — any
            # amount already deducted simply stays on the employee's record.
            reimburse_month = None
            reimburse_year = None

    payments_made = payments_made_form

    has_ledger = conn.execute(
        "SELECT 1 FROM deduction_transactions_cents WHERE plan_type='undercharge' "
        "AND plan_id=? AND COALESCE(voided,0)=0 LIMIT 1", (uc_id,)).fetchone()
    has_events = conn.execute(
        "SELECT 1 FROM undercharge_events WHERE undercharge_id=? LIMIT 1",
        (uc_id,)).fetchone()
    has_legacy = conn.execute(
        "SELECT 1 FROM undercharges_cents WHERE id=? "
        "AND (legacy_paid_cents>0 OR legacy_payments_count>0)",
        (uc_id,)).fetchone()
    financial_change = (
        uc_type != (uc['type'] or 'undercharge')
        or db.to_cents(total_amount) != db.to_cents(uc['total_amount'])
        or recovery_method != uc['recovery_method']
        or split_months != (uc['split_months'] or 1)
        or start_month != (uc['start_month'] or uc['incident_month'])
        or start_year != (uc['start_year'] or uc['incident_year'])
        or payments_made != (uc['payments_made'] or 0)
    )
    if (has_ledger or has_events or has_legacy) and financial_change:
        conn.close()
        flash('Completed financial history cannot be rewritten. Use “Reschedule '
              'remaining balance”, record a customer payment, or void an unlocked '
              'transaction.', 'danger')
        return redirect(request.referrer or
                        url_for('employee_detail', emp_id=uc['employee_id']))

    with conn:
        conn.execute('''
            UPDATE undercharges_cents
            SET type=?, reason=?, sale_number=?, total_amount_cents=?, recovery_method=?, split_months=?,
                payments_made=?, status=?, incident_month=?, incident_year=?, start_month=?, start_year=?,
                reimburse_month=?, reimburse_year=?, notes=?
            WHERE id=?
        ''', (uc_type, reason, sale_number, db.to_cents(total_amount), recovery_method, split_months,
              payments_made, status, incident_month, incident_year, start_month, start_year,
              reimburse_month, reimburse_year, notes, uc_id))
        if uc_type == 'undercharge' and not has_ledger and not has_events and not has_legacy:
            db.create_undercharge_schedule(
                conn, uc_id, start_year, start_month, split_months,
                db.to_cents(total_amount), kind='deduction',
                reason='Schedule updated before first payment',
                actor=session.get('admin') or 'admin')
    conn.close()
    
    flash('Deduction entry updated successfully.', 'success')
    return redirect(request.referrer or url_for('employee_detail', emp_id=uc['employee_id']))



# ── Monthly Payroll ─────────────────────────────────────────────────────────


@app.route('/undercharges')
def undercharges_list():
    conn = db.get_db()
    store_filter  = request.args.get('store', '')
    status_filter = request.args.get('status', 'outstanding')
    type_filter   = request.args.get('type', 'all')   # all | undercharge | overcharge
    search        = request.args.get('q', '').strip()

    query = '''
        SELECT uc.*, e.full_name, e.current_store, e.job_title
        FROM undercharges uc
        JOIN employees e ON e.id = uc.employee_id
        WHERE 1=1
    '''
    params = []
    if status_filter == 'outstanding':
        query += " AND uc.status IN ('pending','partial')"
    elif status_filter == 'recovered':
        query += " AND uc.status IN ('recovered','accounted_for','paid_by_customer')"
    elif status_filter == 'written_off':
        query += " AND uc.status = 'written_off'"
    if type_filter in ('undercharge', 'overcharge'):
        query += " AND uc.type = ?"
        params.append(type_filter)
    if store_filter:
        query += ' AND e.current_store = ?'
        params.append(store_filter)
    if search:
        # The name/reason box used to filter only the rows already in the DOM.
        # Now the list is windowed, a match on page 4 has to be found by the
        # server or the search would report "no results" for a live undercharge.
        query += (' AND (e.full_name LIKE ? OR IFNULL(uc.reason, "") LIKE ?'
                  ' OR IFNULL(uc.sale_number, "") LIKE ?)')
        params.extend([f'%{search}%'] * 3)
    query += ' ORDER BY uc.type DESC, e.current_store, e.full_name, uc.incident_year DESC, uc.incident_month DESC'
    rows = conn.execute(query, params).fetchall()
    stores = [r[0] for r in conn.execute(
        "SELECT DISTINCT e.current_store FROM undercharges uc "
        "JOIN employees e ON e.id=uc.employee_id WHERE uc.status IN ('pending','partial') ORDER BY e.current_store"
    ).fetchall()]
    # Priced and scheduled in two passes instead of per row: the per-row account
    # is five statements and the schedule lookup a sixth, so this page issued 141
    # for 22 rows and grew linearly with the table. Overcharges are skipped —
    # they are on-record only and never consult an account.
    priced_ids = [r['id'] for r in rows if (r['type'] or 'undercharge') != 'overcharge']
    accounts = db.get_undercharge_accounts(priced_ids, conn)
    future_by_uc = {}
    if priced_ids:
        marks = ','.join('?' * len(priced_ids))
        for row in conn.execute(
                "SELECT undercharge_id,due_year,due_month,amount_cents "
                "FROM undercharge_schedule_items "
                f"WHERE undercharge_id IN ({marks}) AND state='scheduled' "
                "AND transaction_id IS NULL ORDER BY due_year,due_month,id",
                priced_ids):
            # Only the three columns the per-row query selected — undercharge_id
            # is the grouping key here, and would otherwise leak into the dicts
            # handed to the template.
            future_by_uc.setdefault(row['undercharge_id'], []).append({
                'due_year': row['due_year'], 'due_month': row['due_month'],
                'amount_cents': row['amount_cents']})

    items = []
    for r in rows:
        uc_type = r['type'] or 'undercharge'
        if uc_type == 'overcharge':
            # Overcharges are never "remaining" — they're just on record
            remaining = r['total_amount'] if r['status'] == 'pending' else 0
            paid_amount = 0
            payment_count = r['payments_made']
            display_status = r['status']
            future_items = []
            refund_due = 0
            scheduled_refund = 0
        else:
            # One branch, not two: the terminal-status case (recovered /
            # written_off / accounted_for / paid_by_customer) and the live case
            # were byte-identical, because get_undercharge_account already
            # derives the display status from the ledger for both. Two copies
            # were only ever a chance for one to drift.
            account = accounts[r['id']]
            remaining = db.to_rands(account['remaining_cents'])
            paid_amount = db.to_rands(account['net_employee_paid_cents'])
            payment_count = account['payment_count']
            display_status = account['status']
            refund_due = db.to_rands(account['refund_due_cents'])
            scheduled_refund = db.to_rands(account['scheduled_refunds_cents'])
            future_items = future_by_uc.get(r['id'], [])
        deduction_future = [dict(x) for x in future_items if x['amount_cents'] > 0]
        refund_future = [dict(x) for x in future_items if x['amount_cents'] < 0]
        items.append({
            'id': r['id'], 'employee_id': r['employee_id'],
            'full_name': r['full_name'], 'current_store': r['current_store'],
            'job_title': r['job_title'], 'reason': r['reason'] or ('Overcharge' if uc_type == 'overcharge' else 'Undercharge'),
            'sale_number': r['sale_number'] or '', 'total_amount': r['total_amount'],
            'recovery_method': r['recovery_method'], 'split_months': r['split_months'],
            'payments_made': payment_count, 'remaining': round(remaining, 2),
            'paid_amount': round(paid_amount, 2),
            'future_items': deduction_future, 'refund_items': refund_future,
            'future_count': len(deduction_future),
            'refund_due': refund_due,
            'scheduled_refund': scheduled_refund,
            'future_monthly': (db.to_rands(deduction_future[0]['amount_cents'])
                               if deduction_future else 0),
            'incident_month': r['incident_month'], 'incident_year': r['incident_year'],
            'start_month': r['start_month'] if r['start_month'] is not None else r['incident_month'],
            'start_year': r['start_year'] if r['start_year'] is not None else r['incident_year'],
            'status': display_status, 'stored_status': r['status'],
            'reimburse_month': r['reimburse_month'],
            'reimburse_year': r['reimburse_year'],
            'notes': r['notes'] or '',
            'type': uc_type,
        })
    now = datetime.now()
    all_employees = [dict(r) for r in conn.execute(
        "SELECT id, full_name, current_store FROM employees WHERE status='active' AND sector='retail' ORDER BY current_store, full_name"
    ).fetchall()]
    conn.close()
    # Only count undercharge outstanding in the financial total
    uc_outstanding = sum(i['remaining'] for i in items if i['type'] == 'undercharge')
    # Every stat card, the reimbursement banner and the TOTAL row describe the
    # whole filtered set. They are computed here, before the display window, so
    # paging can never make the amount owed look smaller than it is. The
    # template used to derive these with |selectattr/|sum over `items`.
    uc_items = [i for i in items if i['type'] == 'undercharge']
    ov_items = [i for i in items if i['type'] == 'overcharge']
    totals = {
        'items': len(items),
        'undercharges': len(uc_items),
        'overcharges': len(ov_items),
        'employees': len({i['employee_id'] for i in items}),
        'undercharge_amount': sum(i['total_amount'] for i in uc_items),
        'overcharge_amount': sum(i['total_amount'] for i in ov_items),
        'reimbursements': sum(i['refund_due'] or 0 for i in items),
    }
    items, pager = paginate(items, noun='entries')
    return render_template('undercharges.html', items=items, stores=stores,
                           store_filter=store_filter, status_filter=status_filter,
                           type_filter=type_filter, all_employees=all_employees,
                           search=search, totals=totals, pager=pager,
                           total_outstanding=uc_outstanding,
                           now=now, MONTH_NAMES=db.MONTH_NAMES)


# ── Undercharges Export ──────────────────────────────────────────────────────


def _uc_remaining(r, conn):
    """Outstanding Rands for an undercharge/overcharge row — same rule as the
    list view (see undercharges_list) so the export matches the screen."""
    uc_type = r['type'] or 'undercharge'
    if uc_type == 'overcharge':
        return r['total_amount'] if r['status'] == 'pending' else 0
    return db.to_rands(db.get_undercharge_account(r['id'], conn)['remaining_cents'])


@app.route('/undercharges/export')
def export_undercharges():
    """Excel export of the undercharges list, honouring the page's store/status/
    type filters. Mirrors the uniforms export styling; grouped by store with a
    grand total. Money is read as Rands (the views expose cents/100)."""
    conn = db.get_db()
    store_filter  = request.args.get('store', '')
    status_filter = request.args.get('status', 'outstanding')
    type_filter   = request.args.get('type', 'all')
    from_month    = request.args.get('from_month', type=int)
    from_year     = request.args.get('from_year',  type=int)
    to_month      = request.args.get('to_month',   type=int)
    to_year       = request.args.get('to_year',    type=int)

    query = '''
        SELECT uc.*, e.full_name, e.current_store, e.job_title
        FROM undercharges uc
        JOIN employees e ON e.id = uc.employee_id
        WHERE e.status = 'active'
    '''
    params = []
    if status_filter == 'outstanding':
        query += " AND uc.status IN ('pending','partial')"
    elif status_filter == 'recovered':
        query += " AND uc.status IN ('recovered','accounted_for','paid_by_customer')"
    elif status_filter == 'written_off':
        query += " AND uc.status = 'written_off'"
    if type_filter in ('undercharge', 'overcharge'):
        query += " AND uc.type = ?"
        params.append(type_filter)
    if store_filter:
        query += ' AND e.current_store = ?'
        params.append(store_filter)
    # Optional incident time-frame (mirrors the uniforms/lay-bys exports, but on
    # the incident month rather than a plan start date). Fall back to the start
    # month when a row has no incident date, so a NULL-incident row is not
    # silently dropped the moment a range is applied.
    _uc_idx = ('(COALESCE(uc.incident_year, uc.start_year) * 12 '
               '+ COALESCE(uc.incident_month, uc.start_month))')
    if from_month and from_year:
        query += f' AND {_uc_idx} >= ?'
        params.append(from_year * 12 + from_month)
    if to_month and to_year:
        query += f' AND {_uc_idx} <= ?'
        params.append(to_year * 12 + to_month)
    query += ' ORDER BY e.current_store, uc.type DESC, e.full_name, uc.incident_year DESC, uc.incident_month DESC'
    rows = conn.execute(query, params).fetchall()
    # NOT closed here: the per-row loop below still prices each undercharge
    # through `conn` (get_undercharge_account / _uc_remaining). Closing at this
    # point raised "Cannot operate on a closed database" and 500'd the export —
    # it only stopped doing so once a request-shared connection made close() a
    # no-op, which is a coincidence, not a contract. Closed after the loop.

    # ── Styles (mirror the uniforms export) ──────────────────────────────────
    wb = openpyxl.Workbook()
    navy    = PatternFill('solid', start_color='1A2B4A')
    amber   = PatternFill('solid', start_color='B45309')
    alt_row = PatternFill('solid', start_color='F0F4FF')
    store_hdr_fill = PatternFill('solid', start_color='1E3A5F')
    white_bold = Font(bold=True, color='FFFFFF', size=11)
    store_font = Font(bold=True, color='BDD7FF', size=10, italic=True)
    thin = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )
    cur = 'R#,##0.00;-R#,##0.00;"-"'

    ws = wb.active
    ws.title = 'Undercharges'
    headers = [
        'Employee ID', 'Employee Name', 'Store', 'Job Title', 'Type',
        'Sale Number', 'Reason', 'Incident', 'Recovery', 'Progress',
        'Total (R)', 'Paid (R)', 'Remaining (R)', 'Status',
    ]
    widths = [12, 28, 20, 22, 12, 14, 34, 14, 10, 10, 14, 14, 16, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A3'

    # Title row
    ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
    title_cell = ws['A1']
    store_label  = f' · {store_filter}' if store_filter else ' · All Stores'
    status_label = {'outstanding': 'Outstanding', 'recovered': 'Recovered',
                    'written_off': 'Written Off'}.get(status_filter, 'All')
    type_label   = {'undercharge': 'Undercharges', 'overcharge': 'Overcharges'}.get(type_filter, 'Under/Overcharges')
    from_label = f'{db.MONTH_NAMES[from_month]} {from_year}' if from_month and from_year else ''
    to_label   = f'{db.MONTH_NAMES[to_month]} {to_year}' if to_month and to_year else ''
    date_range = f' · {from_label} – {to_label}' if (from_label or to_label) else ''
    title_cell.value = f'NORTHWIND {type_label} — {status_label}{store_label}{date_range}'
    title_cell.font = Font(bold=True, color='FFFFFF', size=13)
    title_cell.fill = navy
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24

    # Header row
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = white_bold
        c.fill = amber
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = thin
    ws.row_dimensions[2].height = 18

    data_row = 3
    last_store = None
    alt = False
    store_totals = {}  # store -> {total, paid, remaining, count}

    for r in rows:
        store   = r['current_store']
        uc_type = r['type'] or 'undercharge'
        total_amt = r['total_amount'] or 0
        account = None if uc_type == 'overcharge' else db.get_undercharge_account(r['id'], conn)
        rem_val = round(_uc_remaining(r, conn), 2)
        # Overcharges are on-record only; "paid" applies to recovered undercharges.
        paid_val = ('' if uc_type == 'overcharge'
                    else db.to_rands(account['net_employee_paid_cents']))
        recovery  = '—' if uc_type == 'overcharge' else (r['recovery_method'] or '').title()
        progress = (f"{account['payment_count']} paid"
                    if uc_type != 'overcharge' else '—')
        incident  = (f"{db.MONTH_NAMES[r['incident_month']]} {r['incident_year']}"
                     if r['incident_month'] and r['incident_year'] else '')

        # Store group header
        if store != last_store:
            last_store = store
            alt = False
            ws.merge_cells(f'A{data_row}:{get_column_letter(len(headers))}{data_row}')
            sc = ws.cell(row=data_row, column=1, value=f'  🏪  {store}')
            sc.font = store_font
            sc.fill = store_hdr_fill
            sc.alignment = Alignment(horizontal='left', vertical='center')
            ws.row_dimensions[data_row].height = 16
            data_row += 1
            store_totals.setdefault(store, {'total': 0, 'paid': 0, 'remaining': 0, 'count': 0})

        fill = alt_row if alt else None
        alt = not alt

        row_vals = [
            r['employee_id'], r['full_name'], store, r['job_title'] or '',
            'Overcharge' if uc_type == 'overcharge' else 'Undercharge',
            r['sale_number'] or '', r['reason'] or '', incident, recovery, progress,
            round(total_amt, 2), paid_val, rem_val,
            account['status'] if account else r['status'],
        ]
        for col, val in enumerate(map(db.xl_safe, row_vals), 1):
            c = ws.cell(row=data_row, column=col, value=val)
            c.border = thin
            c.alignment = Alignment(vertical='center')
            if fill:
                c.fill = fill
            if col in (11, 12):   # Total, Paid
                c.number_format = cur; c.alignment = Alignment(horizontal='right', vertical='center')
            elif col == 13:       # Remaining
                c.number_format = cur; c.alignment = Alignment(horizontal='right', vertical='center')
                c.font = Font(bold=True, color='C84B11' if rem_val > 0 else '059669')

        store_totals[store]['total']     += total_amt
        store_totals[store]['paid']      += (paid_val or 0)
        store_totals[store]['remaining'] += rem_val
        store_totals[store]['count']     += 1
        data_row += 1

    conn.close()   # last use of the connection — the rest is workbook assembly

    # Grand total row
    data_row += 1
    total_total     = sum(s['total']     for s in store_totals.values())
    total_paid      = sum(s['paid']      for s in store_totals.values())
    total_remaining = sum(s['remaining'] for s in store_totals.values())
    total_count     = sum(s['count']     for s in store_totals.values())
    gt_vals = ['', f'GRAND TOTAL  ({total_count} records)', '', '', '', '', '', '', '', '',
               round(total_total, 2), round(total_paid, 2), round(total_remaining, 2), '']
    for col, val in enumerate(map(db.xl_safe, gt_vals), 1):
        c = ws.cell(row=data_row, column=col, value=val)
        c.font = Font(bold=True, color='FFFFFF', size=11)
        c.fill = navy
        c.border = thin
        if col in (11, 12, 13):
            c.number_format = cur; c.alignment = Alignment(horizontal='right', vertical='center')

    out = io.BytesIO()
    wb.save(out); out.seek(0)
    now = datetime.now()
    store_slug = f'_{store_filter.replace(" ", "_")}' if store_filter else '_All_Stores'
    # Date-slug mirrors the uniforms/lay-bys exports (full range → From-To),
    # but stays robust if only one side of the incident range is supplied.
    if from_month and from_year and to_month and to_year:
        date_slug = f'_{from_year}{from_month:02d}-{to_year}{to_month:02d}'
    elif from_month and from_year:
        date_slug = f'_from{from_year}{from_month:02d}'
    elif to_month and to_year:
        date_slug = f'_to{to_year}{to_month:02d}'
    else:
        date_slug = ''
    filename = f'Undercharges{store_slug}{date_slug}_{status_filter}_{now.strftime("%Y%m%d")}.xlsx'
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)
