from flask import (render_template, request, redirect, url_for, flash,
                   jsonify, send_file, session)
from datetime import datetime
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from northwind.data import database as db
from northwind.core import app, employee_detail_url
from northwind.deductions import plans
from northwind.deductions.pagination import paginate


@app.route('/layby/add', methods=['POST'])
def add_layby():
    """Parse the basket form, then delegate to plans.create_layby_plan.

    The discount/monthly math, validation, lock check and inserts live in
    northwind/deductions/plans.py, shared with the MCP connector.
    """
    emp_id = request.form['employee_id']
    sector = db.get_employee_sector(emp_id)

    # Collect basket items off the indexed form fields.
    items = []
    i = 0
    while f'item_desc_{i}' in request.form:
        items.append({'description': request.form.get(f'item_desc_{i}', ''),
                      'unit_price': request.form.get(f'item_price_{i}', 0) or 0,
                      'quantity': request.form.get(f'item_qty_{i}', 1) or 1})
        i += 1

    conn = db.get_db()
    try:
        with conn:
            plans.create_layby_plan(
                conn, emp_id,
                items=items,
                term=request.form['term_months'],
                start_year=request.form['start_year'],
                start_month=request.form['start_month'],
                discount_pct=request.form.get('discount_pct', 40),
                sale_number=request.form.get('sale_number', ''),
                notes=request.form.get('notes', ''),
                actor=session.get('admin') or 'admin')
    except KeyError:
        flash('Please enter valid numbers for term, discount and start period.', 'danger')
        return redirect(employee_detail_url(emp_id, sector))
    except ValueError as exc:
        flash(str(exc), 'danger')
        return redirect(employee_detail_url(emp_id, sector))
    finally:
        conn.close()

    flash('Lay-by plan added.', 'success')
    return redirect(employee_detail_url(emp_id))


@app.route('/layby/<int:plan_id>/edit', methods=['POST'])
def edit_layby(plan_id):
    """Parse the edit form, then delegate to plans.edit_plan.

    The validation, status derivation, UPDATE and basket-row sync live in
    northwind/deductions/plans.py, shared with the MCP connector's edit_plan tool.
    """
    conn = db.get_db()
    try:
        with conn:
            result = plans.edit_plan(
                conn, 'layby', plan_id,
                actor=session.get('admin_username') or 'admin',
                description=request.form.get('description', ''),
                sale_number=request.form.get('sale_number', ''),
                basket_total=request.form.get('basket_total', 0) or 0,
                discount_pct=request.form.get('discount_pct', 40) or 40,
                total=request.form.get('total_amount'),
                monthly=request.form.get('monthly_amount'),
                balance_remaining=request.form.get('balance_remaining'),
                term_months=request.form.get('term_months'),
                payments_made=request.form.get('payments_made'),
                start_month=request.form.get('start_month'),
                start_year=request.form.get('start_year'),
                notes=request.form.get('notes', ''))
    except ValueError as exc:
        flash(str(exc), 'danger')
        return redirect(request.referrer or url_for('employees'))
    finally:
        conn.close()

    flash('Lay-by plan updated.', 'success')
    return redirect(employee_detail_url(result['employee_id']))


@app.route('/layby/<int:plan_id>/tick', methods=['POST'])
def tick_layby(plan_id):
    conn = db.get_db()
    plan = conn.execute("SELECT * FROM layby_deductions WHERE id=?", (plan_id,)).fetchone()
    if not plan:
        conn.close()
        if request.headers.get('Accept') == 'application/json':
            return jsonify({'success': False, 'message': 'Plan not found.'})
        flash('Plan not found.', 'danger')
        return redirect(request.referrer or url_for('employees'))

    if plan['payments_made'] >= plan['term_months'] or (plan['balance_remaining'] is not None and plan['balance_remaining'] <= 0.01):
        conn.close()
        if request.headers.get('Accept') == 'application/json':
            return jsonify({'success': False, 'message': 'This plan is already fully paid.'})
        flash('Plan is already fully paid.', 'warning')
        return redirect(request.referrer or employee_detail_url(plan['employee_id']))

    # Calculate target period
    start_m = plan['start_month']
    start_y = plan['start_year']
    payments = plan['payments_made']
    target_m = start_m + payments
    target_y = start_y + (target_m - 1) // 12
    target_m = ((target_m - 1) % 12) + 1

    if db.is_period_locked(target_y, target_m, db.get_employee_sector(plan['employee_id'], conn)):
        conn.close()
        if request.headers.get('Accept') == 'application/json':
            return jsonify({'success': False, 'message': 'The targeted payroll period is locked.'})
        flash('Cannot process payment. The targeted payroll period is locked.', 'danger')
        return redirect(request.referrer or employee_detail_url(plan['employee_id']))

    # Prevent duplicate payments for the target month
    exists = conn.execute('''
        SELECT 1 FROM deduction_transactions 
        WHERE plan_type = 'layby' AND plan_id = ? AND year = ? AND month = ? AND COALESCE(voided, 0) = 0
    ''', (plan_id, target_y, target_m)).fetchone()
    if exists:
        conn.close()
        if request.headers.get('Accept') == 'application/json':
            return jsonify({'success': False, 'message': f'A payment has already been allocated to {db.MONTH_FULL[target_m]} {target_y}.'})
        flash(f'A payment has already been allocated to {db.MONTH_FULL[target_m]} {target_y}.', 'warning')
        return redirect(request.referrer or employee_detail_url(plan['employee_id']))

    # Deduct the agreed regular monthly installment; the plan completes after
    # `term` payments. Matches the payroll export — we no longer shrink the
    # final month to a possibly-stale balance. A basket-only plan with no
    # per-month figure falls back to clearing the remaining balance.
    current_bal = plan['balance_remaining'] if plan['balance_remaining'] is not None else plan['total_amount']
    installment = round(plan['monthly_amount'], 2) if plan['monthly_amount'] else current_bal
    installment = round(max(0, installment), 2)

    new_balance  = round(max(0, current_bal - installment), 2)
    new_payments = payments + 1
    status = 'complete' if (new_payments >= plan['term_months'] or new_balance <= 0.01) else 'active'

    conn.execute(
        "UPDATE layby_deductions_cents SET payments_made=?, balance_remaining_cents=?, status=? WHERE id=?",
        (new_payments, db.to_cents(max(0, new_balance)), status, plan_id)
    )
    conn.execute('''
        INSERT INTO deduction_transactions_cents (plan_type, plan_id, employee_id, amount_cents, year, month)
        VALUES ('layby', ?, ?, ?, ?, ?)
    ''', (plan_id, plan['employee_id'], db.to_cents(installment), target_y, target_m))
    conn.commit()
    conn.close()

    if request.headers.get('Accept') == 'application/json':
        outstanding = db.get_outstanding_summary(plan['employee_id'])
        cat_totals = db.get_category_totals(plan['employee_id'])
        plan_paid = round((plan['total_amount'] or 0) - max(0, new_balance), 2)
        return jsonify({
            'success': True,
            'new_payments_made': new_payments,
            'status': status,
            'plan_remaining': max(0, new_balance),
            'plan_paid': plan_paid,
            'outstanding': outstanding,
            'cat_totals': cat_totals
        })
    return redirect(request.referrer or employee_detail_url(plan['employee_id']))


@app.route('/layby/<int:plan_id>/adjust', methods=['POST'])
def adjust_layby(plan_id):
    """Record an ad-hoc payment and recalculate remaining monthly amount."""
    conn = db.get_db()
    try:
        with conn:
            result = plans.adjust_plan(
                conn, 'layby', plan_id,
                amount=request.form['amount'],
                note=request.form.get('note', ''),
                new_term=request.form.get('new_term', '').strip() or None,
                actor=session.get('admin') or 'admin')
    except KeyError:
        flash('Please enter a valid payment amount.', 'danger')
        return redirect(request.referrer or url_for('employees'))
    except ValueError as exc:
        flash(str(exc), 'danger')
        return redirect(request.referrer or url_for('employees'))
    finally:
        conn.close()

    flash(f"Ad-hoc payment of R{result['amount']:.2f} recorded. "
          f"New monthly: R{result['new_monthly']:.2f}.", 'success')
    return redirect(employee_detail_url(result['employee_id']))


@app.route('/layby/<int:plan_id>/items')
def layby_items(plan_id):
    conn = db.get_db()
    items = conn.execute("SELECT * FROM layby_items WHERE layby_id=?", (plan_id,)).fetchall()
    adjustments = conn.execute(
        "SELECT * FROM plan_adjustments WHERE plan_type='layby' AND plan_id=? ORDER BY created_at",
        (plan_id,)
    ).fetchall()
    conn.close()
    return jsonify({
        'items': [dict(r) for r in items],
        'adjustments': [dict(r) for r in adjustments]
    })


@app.route('/layby/<int:plan_id>/write-off', methods=['POST'])
def write_off_layby(plan_id):
    conn = db.get_db()
    try:
        with conn:
            result = plans.write_off_plan(
                conn, 'layby', plan_id,
                reason=request.form.get('reason'),
                actor=session.get('admin') or 'admin')
    except ValueError as exc:
        flash(str(exc), 'danger')
        return redirect(request.referrer or url_for('employees'))
    finally:
        conn.close()

    flash('Lay-by plan written off.', 'warning')
    return redirect(employee_detail_url(result['employee_id']))


# ── Undercharges ────────────────────────────────────────────────────────────


@app.route('/laybys/export')
def export_laybys():
    conn = db.get_db()
    store_filter = request.args.get('store', '').strip()
    from_month   = request.args.get('from_month', type=int)
    from_year    = request.args.get('from_year',  type=int)
    to_month     = request.args.get('to_month',   type=int)
    to_year      = request.args.get('to_year',    type=int)

    query = '''
        SELECT l.*, e.full_name, e.current_store, e.job_title, e.id as emp_id
        FROM layby_deductions l
        JOIN employees e ON e.id = l.employee_id
        WHERE l.status = 'active' AND e.sector = 'retail'
          AND e.status = 'active'
          AND l.payments_made < l.term_months
          AND (l.balance_remaining IS NULL OR l.balance_remaining > 0.01)
    '''
    params = []
    if store_filter:
        query += ' AND e.current_store = ?'
        params.append(store_filter)
    if from_month and from_year:
        query += ' AND (l.start_year * 12 + l.start_month) >= ?'
        params.append(from_year * 12 + from_month)
    if to_month and to_year:
        query += ' AND (l.start_year * 12 + l.start_month) <= ?'
        params.append(to_year * 12 + to_month)
    query += ' ORDER BY e.current_store, e.full_name, l.id'
    rows = conn.execute(query, params).fetchall()
    conn.close()

    # ── Styles ──────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    navy    = PatternFill('solid', start_color='0F2A42')
    teal    = PatternFill('solid', start_color='0F5B52')
    alt_row = PatternFill('solid', start_color='F0FBF9')
    store_hdr_fill = PatternFill('solid', start_color='164E45')
    white_bold = Font(bold=True, color='FFFFFF', size=11)
    store_font = Font(bold=True, color='99F6E4', size=10, italic=True)
    thin = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )
    cur = 'R#,##0.00;-R#,##0.00;"-"'
    pct = '0.0"%"'

    # ── Sheet 1: Lay-by Plans ────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Lay-by Plans'
    headers = [
        'Employee ID', 'Employee Name', 'Store', 'Job Title',
        'Sale Number', 'Description',
        'Start', 'Term (mo)', 'Basket (R)', 'Disc %',
        'Total (R)', 'Regular Monthly (R)', 'Balance (R)',
        'Payments Made', 'Status',
    ]
    widths = [12, 28, 20, 22, 14, 36, 12, 10, 14, 8, 14, 18, 14, 14, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A3'

    # Title row
    ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
    title_cell = ws['A1']
    from_label = f'{db.MONTH_NAMES[from_month]} {from_year}' if from_month and from_year else ''
    to_label   = f'{db.MONTH_NAMES[to_month]} {to_year}' if to_month and to_year else ''
    date_range = f' · {from_label} – {to_label}' if from_label else ''
    store_label = f' · {store_filter}' if store_filter else ' · All Stores'
    title_cell.value = f'NORTHWIND Active Staff Lay-by Plans{store_label}{date_range}'
    title_cell.font = Font(bold=True, color='FFFFFF', size=13)
    title_cell.fill = navy
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24

    # Header row
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = white_bold; c.fill = teal
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = thin
    ws.row_dimensions[2].height = 18

    data_row = 3
    last_store = None
    alt = False
    store_totals = {}

    for r in rows:
        store = r['current_store']
        balance = r['balance_remaining'] if r['balance_remaining'] is not None else (r['term_months'] - r['payments_made']) * r['monthly_amount']
        balance = round(max(0, balance), 2)

        if store != last_store:
            last_store = store
            alt = False
            ws.merge_cells(f'A{data_row}:{get_column_letter(len(headers))}{data_row}')
            sc = ws.cell(row=data_row, column=1, value=f'  🏪  {store}')
            sc.font = store_font; sc.fill = store_hdr_fill
            sc.alignment = Alignment(horizontal='left', vertical='center')
            ws.row_dimensions[data_row].height = 16
            data_row += 1
            if store not in store_totals:
                store_totals[store] = {'monthly': 0, 'basket': 0, 'total': 0, 'balance': 0, 'count': 0}

        fill = alt_row if alt else None
        alt = not alt

        row_vals = [
            r['emp_id'], r['full_name'], store, r['job_title'] or '',
            r['sale_number'] or '', r['description'] or '',
            f"{db.MONTH_NAMES[r['start_month']]} {r['start_year']}",
            r['term_months'],
            r['basket_total'] or 0,
            r['discount_pct'] or 40,
            r['total_amount'] or 0,
            round(r['monthly_amount'], 2),
            balance,
            r['payments_made'], r['status'],
        ]
        for col, val in enumerate(map(db.xl_safe, row_vals), 1):
            c = ws.cell(row=data_row, column=col, value=val)
            c.border = thin; c.alignment = Alignment(vertical='center')
            if fill:
                c.fill = fill
            if col == 9:   # Basket
                c.number_format = cur; c.alignment = Alignment(horizontal='right', vertical='center')
            elif col == 10:  # Disc %
                c.number_format = pct; c.alignment = Alignment(horizontal='right', vertical='center')
            elif col == 11:  # Total
                c.number_format = cur; c.alignment = Alignment(horizontal='right', vertical='center')
            elif col == 12:  # Monthly
                c.number_format = cur; c.alignment = Alignment(horizontal='right', vertical='center')
            elif col == 13:  # Balance
                c.number_format = cur; c.alignment = Alignment(horizontal='right', vertical='center')
                c.font = Font(bold=True, color='C84B11' if balance > 0 else '059669')

        store_totals[store]['monthly'] += r['monthly_amount']
        store_totals[store]['basket']  += r['basket_total'] or 0
        store_totals[store]['total']   += r['total_amount'] or 0
        store_totals[store]['balance'] += balance
        store_totals[store]['count']   += 1
        data_row += 1

    # Grand total
    data_row += 1
    tm = sum(s['monthly'] for s in store_totals.values())
    tb = sum(s['basket']  for s in store_totals.values())
    tt = sum(s['total']   for s in store_totals.values())
    tbal = sum(s['balance'] for s in store_totals.values())
    tc = sum(s['count']   for s in store_totals.values())
    gt_vals = ['', f'GRAND TOTAL  ({tc} plans)', '', '', '', '', '', '', tb, '', tt, tm, tbal, '', '']
    for col, val in enumerate(map(db.xl_safe, gt_vals), 1):
        c = ws.cell(row=data_row, column=col, value=val)
        c.font = Font(bold=True, color='FFFFFF', size=11); c.fill = navy; c.border = thin
        if col in (9, 11, 12, 13):
            c.number_format = cur; c.alignment = Alignment(horizontal='right', vertical='center')

    # ── Sheet 2: Summary by Store ────────────────────────────────────────────
    ws2 = wb.create_sheet('Summary by Store')
    for i, w in enumerate([28, 10, 22, 16, 16, 16], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    ws2.merge_cells('A1:F1')
    t2 = ws2['A1']
    t2.value = f'Lay-by Plans — Summary by Store{store_label}{date_range}'
    t2.font = Font(bold=True, color='FFFFFF', size=12); t2.fill = navy
    t2.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 22

    s2_hdrs = ['Store', '# Plans', 'Regular Monthly (R)', 'Basket Total (R)', 'Net Total (R)', 'Balance (R)']
    for col, h in enumerate(s2_hdrs, 1):
        c = ws2.cell(row=2, column=col, value=h)
        c.font = white_bold; c.fill = teal
        c.alignment = Alignment(horizontal='center', vertical='center'); c.border = thin
    ws2.row_dimensions[2].height = 18

    for i, (store, s) in enumerate(sorted(store_totals.items()), 3):
        row_data = [store, s['count'], round(s['monthly'], 2), round(s['basket'], 2), round(s['total'], 2), round(s['balance'], 2)]
        fill = alt_row if i % 2 == 1 else None
        for col, val in enumerate(map(db.xl_safe, row_data), 1):
            c = ws2.cell(row=i, column=col, value=val)
            c.border = thin; c.alignment = Alignment(vertical='center')
            if fill: c.fill = fill
            if col in (3, 4, 5, 6):
                c.number_format = cur; c.alignment = Alignment(horizontal='right', vertical='center')

    last_r = 2 + len(store_totals) + 1
    gt_s = [f'TOTAL ({len(store_totals)} stores)', tc, round(tm, 2), round(tb, 2), round(tt, 2), round(tbal, 2)]
    for col, val in enumerate(map(db.xl_safe, gt_s), 1):
        c = ws2.cell(row=last_r, column=col, value=val)
        c.font = Font(bold=True, color='FFFFFF', size=11); c.fill = navy; c.border = thin
        if col in (3, 4, 5, 6):
            c.number_format = cur; c.alignment = Alignment(horizontal='right', vertical='center')

    ws2.freeze_panes = 'A3'

    out = io.BytesIO()
    wb.save(out); out.seek(0)
    now = datetime.now()
    store_slug = f'_{store_filter.replace(" ", "_")}' if store_filter else '_All_Stores'
    date_slug  = f'_{from_year}{from_month:02d}-{to_year}{to_month:02d}' if from_month and from_year else ''
    filename   = f'Layby_Plans{store_slug}{date_slug}_{now.strftime("%Y%m%d")}.xlsx'
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


@app.route('/laybys')
def laybys_list():
    conn = db.get_db()
    store_filter = request.args.get('store', '')
    month_filter = request.args.get('month', type=int)
    year_filter  = request.args.get('year',  type=int)
    status_filter = request.args.get('status', 'active')
    search = request.args.get('q', '').strip()
    if status_filter not in ('active', 'complete', 'written_off', 'all'):
        status_filter = 'active'

    query = '''
        SELECT l.*, e.full_name, e.current_store, e.job_title
        FROM layby_deductions l
        JOIN employees e ON e.id = l.employee_id
        WHERE e.sector = 'retail'
    '''
    params = []
    if status_filter != 'all':
        query += ' AND l.status = ?'
        params.append(status_filter)

    if month_filter and year_filter:
        query_idx = year_filter * 12 + month_filter
        # plan must have started by this month and not yet ended
        query += ' AND (l.start_year*12+l.start_month) <= ? AND (l.start_year*12+l.start_month+l.term_months) > ?'
        params.extend([query_idx, query_idx])

    if store_filter:
        query += ' AND e.current_store = ?'
        params.append(store_filter)
    if search:
        # The name/description box used to filter only the rows already in the
        # DOM. Now the list is windowed, a match on page 4 has to be found by
        # the server or the search would report "no results" for a live plan.
        query += ' AND (e.full_name LIKE ? OR IFNULL(l.description, "") LIKE ?)'
        params.extend([f'%{search}%', f'%{search}%'])
    query += ' ORDER BY e.current_store, e.full_name'
    rows = conn.execute(query, params).fetchall()
    stores = [r[0] for r in conn.execute(
        "SELECT DISTINCT e.current_store FROM layby_deductions l "
        "JOIN employees e ON e.id=l.employee_id WHERE l.status='active' AND e.sector='retail' ORDER BY e.current_store"
    ).fetchall()]
    plans = []
    for r in rows:
        balance = r['balance_remaining'] or ((r['term_months'] - r['payments_made']) * r['monthly_amount'])
        pmt_paid = None
        if month_filter and year_filter:
            pmt_index = (year_filter * 12 + month_filter) - (r['start_year'] * 12 + r['start_month'])
            pmt_paid  = pmt_index < r['payments_made']
        plans.append({
            'id': r['id'], 'employee_id': r['employee_id'],
            'full_name': r['full_name'], 'current_store': r['current_store'],
            'job_title': r['job_title'], 'description': r['description'] or '',
            'sale_number': r['sale_number'] or '',
            'basket_total': r['basket_total'] or 0, 'discount_pct': r['discount_pct'] or 40,
            'total_amount': r['total_amount'] or 0, 'monthly_amount': r['monthly_amount'],
            'balance': round(balance, 2), 'term_months': r['term_months'],
            'payments_made': r['payments_made'],
            'payments_left': r['term_months'] - r['payments_made'],
            'start_month': r['start_month'], 'start_year': r['start_year'],
            'pmt_paid': pmt_paid,
        })
    now = datetime.now()
    all_employees = [dict(r) for r in conn.execute(
        "SELECT id, full_name, current_store FROM employees WHERE status='active' AND sector='retail' ORDER BY current_store, full_name"
    ).fetchall()]
    schedule_issues = db.layby_schedule_issues(conn)
    conn.close()
    # Whole-filtered-set figures, taken before the display window: the stat
    # cards and the TOTAL row must not shrink to whatever page you are on.
    totals = {
        'plans': len(plans),
        'employees': len({p['employee_id'] for p in plans}),
        'monthly': sum(p['monthly_amount'] for p in plans),
        'basket': sum(p['basket_total'] for p in plans),
        'outstanding': sum(p['balance'] for p in plans),
    }
    plans, pager = paginate(plans, noun='plans')
    return render_template('laybys.html', plans=plans, stores=stores,
                           store_filter=store_filter, status_filter=status_filter,
                           month_filter=month_filter, year_filter=year_filter,
                           search=search,
                           total_outstanding=totals['outstanding'],
                           totals=totals, pager=pager,
                           all_employees=all_employees,
                           schedule_issues=schedule_issues,
                           now=now, MONTH_NAMES=db.MONTH_NAMES,
                           page_year=year_filter or now.year,
                           page_month=month_filter or now.month)
