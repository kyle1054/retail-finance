"""HQ Deductions — a second, separate workspace alongside Retail.

HQ mirrors the retail section with two fixed locations (HQ + DC) and carries
lay-bys plus staff allowances. Employees here have sector='hq' and never mix
with retail.
Listing / profile / monthly pages reuse the shared templates (parameterised by
`current_workspace`); the heavy lay-by mutation routes (tick/edit/adjust/
write-off/add) are shared with retail and already sector-aware via the
employee's sector. Only sector-scoped surfaces (lists, payroll, locks) live here.
"""
from flask import (render_template, request, redirect, url_for, flash,
                   send_file)
from datetime import datetime
import io
import sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from northwind.data import database as db
from northwind.core import app, HQ_STORES, HQ_JOB_TITLES
from northwind.deductions.routes_monthly import _tick_layby_plans

SECTOR = 'hq'


# ── Employees ────────────────────────────────────────────────────────────────

@app.route('/hq/employees')
def hq_employees():
    conn = db.get_db()
    store_filter = request.args.get('store', '')
    search = request.args.get('q', '').strip()
    status_filter = request.args.get('status', 'active')

    query = "SELECT * FROM employees WHERE sector='hq'"
    params = []
    if store_filter:
        query += " AND current_store = ?"
        params.append(store_filter)
    if search:
        query += " AND full_name LIKE ?"
        params.append(f'%{search}%')
    if status_filter and status_filter != 'all':
        query += " AND status = ?"
        params.append(status_filter)
    query += " ORDER BY current_store, full_name"

    emps = conn.execute(query, params).fetchall()
    conn.close()
    outstanding = db.get_all_outstanding_totals('hq')
    return render_template('employees.html', employees=emps, stores=HQ_STORES,
                           store_filter=store_filter, search=search,
                           status_filter=status_filter,
                           job_titles=HQ_JOB_TITLES, outstanding=outstanding)


@app.route('/hq/employees/add', methods=['POST'])
def hq_add_employee():
    conn = db.get_db()
    emp_id = db.next_employee_id(conn)
    full_name = request.form['full_name'].strip()
    store = request.form['store'].strip()
    job_title = request.form['job_title'].strip()
    if store not in HQ_STORES:
        store = HQ_STORES[0]
    conn.execute(
        "INSERT INTO employees (id, full_name, current_store, job_title, sector) VALUES (?, ?, ?, ?, 'hq')",
        (emp_id, full_name, store, job_title)
    )
    conn.execute(
        "INSERT INTO store_history (employee_id, store, from_date) VALUES (?, ?, ?)",
        (emp_id, store, datetime.now().isoformat())
    )
    conn.commit()
    conn.close()
    flash(f'Employee {full_name} added — ID: {emp_id}', 'success')
    return redirect(url_for('hq_employees'))


@app.route('/hq/employees/<emp_id>')
def hq_employee_detail(emp_id):
    conn = db.get_db()
    emp = conn.execute("SELECT * FROM employees WHERE id=?", (emp_id,)).fetchone()
    if not emp:
        conn.close()
        flash('Employee not found.', 'danger')
        return redirect(url_for('hq_employees'))
    # Keep sections separate: a retail employee opened on an HQ URL goes to Retail.
    if (emp['sector'] or 'retail') != 'hq':
        conn.close()
        return redirect(url_for('employee_detail', emp_id=emp_id))

    laybys = conn.execute(
        "SELECT * FROM layby_deductions WHERE employee_id=? ORDER BY start_year, start_month, id", (emp_id,)
    ).fetchall()
    history = conn.execute(
        "SELECT * FROM store_history WHERE employee_id=? ORDER BY from_date DESC", (emp_id,)
    ).fetchall()
    transactions = conn.execute(
        "SELECT * FROM deduction_transactions WHERE employee_id=? ORDER BY created_at DESC LIMIT 20", (emp_id,)
    ).fetchall()
    conn.close()

    outstanding = db.get_outstanding_summary(emp_id)
    cat_totals  = db.get_category_totals(emp_id)
    schedule    = db.get_employee_schedule(emp_id)

    allowance_year = request.args.get('allowance_year', type=int) or datetime.now().year
    allowance = db.get_allowance_summary(emp_id, allowance_year)
    allowance_purchases = db.get_allowance_purchases(emp_id, allowance_year)

    # uniforms/undercharges are empty for HQ; passed so the shared template's
    # (feature-flagged) references stay defined.
    return render_template('employee.html', emp=emp, uniforms=[], laybys=laybys,
                           undercharges=[], history=history, outstanding=outstanding,
                           cat_totals=cat_totals, schedule=schedule,
                           stores=HQ_STORES, job_titles=HQ_JOB_TITLES,
                           MONTH_NAMES=db.MONTH_NAMES, MONTH_FULL=db.MONTH_FULL,
                           transactions=transactions,
                           allowance=allowance, allowance_year=allowance_year,
                           allowance_purchases=allowance_purchases)


# ── Lay-bys ──────────────────────────────────────────────────────────────────

@app.route('/hq/laybys')
def hq_laybys_list():
    conn = db.get_db()
    store_filter = request.args.get('store', '')
    month_filter = request.args.get('month', type=int)
    year_filter  = request.args.get('year',  type=int)
    status_filter = request.args.get('status', 'active')
    if status_filter not in ('active', 'complete', 'written_off', 'all'):
        status_filter = 'active'

    query = '''
        SELECT l.*, e.full_name, e.current_store, e.job_title
        FROM layby_deductions l
        JOIN employees e ON e.id = l.employee_id
        WHERE e.sector = 'hq'
    '''
    params = []
    if status_filter != 'all':
        query += " AND l.status = ?"
        params.append(status_filter)
    if month_filter and year_filter:
        query_idx = year_filter * 12 + month_filter
        query += ' AND (l.start_year*12+l.start_month) <= ? AND (l.start_year*12+l.start_month+l.term_months) > ?'
        params.extend([query_idx, query_idx])
    if store_filter:
        query += ' AND e.current_store = ?'
        params.append(store_filter)
    query += ' ORDER BY e.current_store, e.full_name'
    rows = conn.execute(query, params).fetchall()

    plans = []
    for r in rows:
        balance = r['balance_remaining'] or ((r['term_months'] - r['payments_made']) * r['monthly_amount'])
        pmt_paid = None
        if month_filter and year_filter:
            pmt_index = (year_filter * 12 + month_filter) - (r['start_year'] * 12 + r['start_month'])
            pmt_paid  = pmt_index < r['payments_made']
        plans.append({
            'id': r['id'], 'employee_id': r['employee_id'],
            'full_name': r['full_name'], 'current_store': r['current_store'],
            'job_title': r['job_title'], 'description': r['description'] or '',
            'sale_number': r['sale_number'] or '',
            'basket_total': r['basket_total'] or 0, 'discount_pct': r['discount_pct'] or 40,
            'total_amount': r['total_amount'] or 0, 'monthly_amount': r['monthly_amount'],
            'balance': round(balance, 2), 'term_months': r['term_months'],
            'payments_made': r['payments_made'],
            'payments_left': r['term_months'] - r['payments_made'],
            'start_month': r['start_month'], 'start_year': r['start_year'],
            'pmt_paid': pmt_paid,
        })
    now = datetime.now()
    all_employees = [dict(r) for r in conn.execute(
        "SELECT id, full_name, current_store FROM employees WHERE status='active' AND sector='hq' ORDER BY current_store, full_name"
    ).fetchall()]
    conn.close()
    return render_template('laybys.html', plans=plans, stores=HQ_STORES,
                           store_filter=store_filter, status_filter=status_filter,
                           month_filter=month_filter, year_filter=year_filter,
                           total_outstanding=sum(p['balance'] for p in plans),
                           all_employees=all_employees,
                           now=now, MONTH_NAMES=db.MONTH_NAMES,
                           page_year=year_filter or now.year,
                           page_month=month_filter or now.month)


# ── Monthly payroll (lay-by only) ────────────────────────────────────────────

@app.route('/hq/monthly')
def hq_monthly_current():
    now = datetime.now()
    return redirect(url_for('hq_monthly_view', year=now.year, month=now.month))


@app.route('/hq/monthly/<int:year>/<int:month>')
def hq_monthly_view(year, month):
    if db.validate_month_year(year, month) is None:
        flash('Invalid period — month must be 1–12.', 'warning')
        return redirect(url_for('hq_monthly_current'))
    data = db.get_monthly_data(year, month, 'hq')
    store_filter = request.args.get('store', '')

    display_data = [d for d in data if d['layby_total'] > 0]
    if store_filter:
        display_data = [d for d in display_data if d['employee']['current_store'] == store_filter]

    prev_month, prev_year = (month - 1, year) if month > 1 else (12, year - 1)
    next_month, next_year = (month + 1, year) if month < 12 else (1, year + 1)
    all_stores = sorted(set(d['employee']['current_store'] for d in data if d['layby_total'] > 0))

    now = datetime.now()
    is_locked = db.is_period_locked(year, month, 'hq')
    return render_template('monthly.html', data=display_data, year=year, month=month,
                           month_name=db.MONTH_NAMES[month], month_full=db.MONTH_FULL[month],
                           cat='layby', now=now, page_year=year, page_month=month,
                           store_filter=store_filter, all_stores=all_stores,
                           prev_month=prev_month, prev_year=prev_year,
                           next_month=next_month, next_year=next_year,
                           total_uniform=0,
                           total_layby=sum(d['layby_total'] for d in display_data),
                           total_undercharges=0,
                           grand_total=sum(d['layby_total'] for d in display_data),
                           invoice_data=[], invoice_total=0,
                           is_locked=is_locked)


@app.route('/hq/monthly/<int:year>/<int:month>/pay-layby/<emp_id>', methods=['POST'])
def hq_pay_layby_monthly(year, month, emp_id):
    if db.validate_month_year(year, month) is None:
        flash('Invalid period.', 'warning')
        return redirect(url_for('hq_monthly_current'))
    if db.is_period_locked(year, month, 'hq'):
        flash('Cannot process payment. This payroll period is locked.', 'danger')
        return redirect(url_for('hq_monthly_view', year=year, month=month))
    conn = db.get_db()
    try:
        with conn:
            _tick_layby_plans(conn, emp_id, year, month)
        flash('Payment processed successfully.', 'success')
    except sqlite3.IntegrityError:
        flash('A payment has already been allocated for this period.', 'warning')
    except Exception as e:
        flash(f'Error processing payment: {e}', 'danger')
    finally:
        conn.close()
    return redirect(url_for('hq_monthly_view', year=year, month=month, store=request.form.get('store', '')))


@app.route('/hq/monthly/<int:year>/<int:month>/pay-all', methods=['POST'])
def hq_pay_all_monthly(year, month):
    store_filter = request.form.get('store', '')
    if db.validate_month_year(year, month) is None:
        flash('Invalid period.', 'warning')
        return redirect(url_for('hq_monthly_current'))
    if db.is_period_locked(year, month, 'hq'):
        flash('Cannot process batch payments. This payroll period is locked.', 'danger')
        return redirect(url_for('hq_monthly_view', year=year, month=month, store=store_filter))
    conn = db.get_db()
    try:
        with conn:
            emp_query = "SELECT id FROM employees WHERE status='active' AND sector='hq'"
            emp_params = []
            if store_filter:
                emp_query += " AND current_store=?"
                emp_params.append(store_filter)
            emp_ids = [r['id'] for r in conn.execute(emp_query, emp_params).fetchall()]
            skipped = 0
            for emp_id in emp_ids:
                # Per-employee savepoint: a duplicate-payment IntegrityError rolls
                # back only this employee instead of aborting the whole batch.
                conn.execute("SAVEPOINT pay_emp")
                try:
                    _tick_layby_plans(conn, emp_id, year, month)
                except sqlite3.IntegrityError:
                    conn.execute("ROLLBACK TO SAVEPOINT pay_emp")
                    skipped += 1
                finally:
                    conn.execute("RELEASE SAVEPOINT pay_emp")
        label = f' for {store_filter}' if store_filter else ''
        if skipped:
            flash(f'Lay-by deductions{label} marked as paid for {db.MONTH_FULL[month]} {year}; '
                  f'{skipped} employee(s) skipped (already allocated this period).', 'warning')
        else:
            flash(f'All Lay-by deductions{label} marked as paid for {db.MONTH_FULL[month]} {year}.', 'success')
    except Exception as e:
        flash(f'Error processing batch payments: {e}', 'danger')
    finally:
        conn.close()
    return redirect(url_for('hq_monthly_view', year=year, month=month, store=store_filter))


@app.route('/hq/monthly/<int:year>/<int:month>/pay-selected', methods=['POST'])
def hq_pay_selected_monthly(year, month):
    """HQ: mark lay-by deductions paid for a chosen subset of employees (per-row
    checkboxes). Mirrors hq_pay_all_monthly, scoped to the posted emp_ids."""
    store_filter = request.form.get('store', '')
    emp_ids = [e for e in request.form.getlist('emp_ids') if e]
    if db.validate_month_year(year, month) is None:
        flash('Invalid period.', 'warning')
        return redirect(url_for('hq_monthly_current'))
    if db.is_period_locked(year, month, 'hq'):
        flash('Cannot process payments. This payroll period is locked.', 'danger')
        return redirect(url_for('hq_monthly_view', year=year, month=month, store=store_filter))
    if not emp_ids:
        flash('No employees were selected.', 'warning')
        return redirect(url_for('hq_monthly_view', year=year, month=month, store=store_filter))
    conn = db.get_db()
    try:
        with conn:
            valid = {r['id'] for r in conn.execute(
                "SELECT id FROM employees WHERE status='active' AND sector='hq'").fetchall()}
            targets = [e for e in emp_ids if e in valid]
            done = 0
            skipped = 0
            for emp_id in targets:
                conn.execute("SAVEPOINT pay_emp")
                try:
                    _tick_layby_plans(conn, emp_id, year, month)
                    done += 1
                except sqlite3.IntegrityError:
                    conn.execute("ROLLBACK TO SAVEPOINT pay_emp")
                    skipped += 1
                finally:
                    conn.execute("RELEASE SAVEPOINT pay_emp")
        msg = f'Lay-by deductions marked as paid for {done} selected employee(s).'
        if skipped:
            msg += f' {skipped} skipped (already allocated this period).'
        flash(msg, 'warning' if skipped else 'success')
    except Exception as e:
        flash(f'Error processing payments: {e}', 'danger')
    finally:
        conn.close()
    return redirect(url_for('hq_monthly_view', year=year, month=month, store=store_filter))


@app.route('/hq/monthly/<int:year>/<int:month>/lock', methods=['POST'])
def hq_lock_period(year, month):
    if db.validate_month_year(year, month) is None:
        flash('Invalid period.', 'warning')
        return redirect(url_for('hq_monthly_current'))
    conn = db.get_db()
    try:
        conn.execute("INSERT OR IGNORE INTO locked_periods (sector, year, month) VALUES ('hq', ?, ?)", (year, month))
        conn.commit()
        flash(f'HQ payroll period {db.MONTH_FULL[month]} {year} has been locked.', 'success')
    except Exception as e:
        flash(f'Error locking period: {e}', 'danger')
    finally:
        conn.close()
    return redirect(url_for('hq_monthly_view', year=year, month=month, store=request.form.get('store', '')))


@app.route('/hq/monthly/<int:year>/<int:month>/unlock', methods=['POST'])
def hq_unlock_period(year, month):
    if db.validate_month_year(year, month) is None:
        flash('Invalid period.', 'warning')
        return redirect(url_for('hq_monthly_current'))
    conn = db.get_db()
    try:
        conn.execute("DELETE FROM locked_periods WHERE sector='hq' AND year = ? AND month = ?", (year, month))
        conn.commit()
        flash(f'HQ payroll period {db.MONTH_FULL[month]} {year} has been unlocked.', 'warning')
    except Exception as e:
        flash(f'Error unlocking period: {e}', 'danger')
    finally:
        conn.close()
    return redirect(url_for('hq_monthly_view', year=year, month=month, store=request.form.get('store', '')))


# ── Exports ──────────────────────────────────────────────────────────────────

def _layby_export(rows, title):
    """Shared Excel builder for HQ lay-by exports (list of view-model dicts)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'HQ Lay-bys'
    navy = PatternFill('solid', start_color='0F2A42')
    teal = PatternFill('solid', start_color='0F5B52')
    white_bold = Font(bold=True, color='FFFFFF', size=11)
    thin = Border(*[Side(style='thin', color='D1D5DB')] * 4)
    cur = 'R#,##0.00;-R#,##0.00;"-"'

    headers = ['Employee ID', 'Employee Name', 'Location', 'Job Title', 'Sale Number',
               'Description', 'Start', 'Term (mo)', 'Regular Monthly (R)', 'Balance (R)',
               'Payments Made', 'Status']
    widths = [12, 26, 14, 22, 14, 34, 12, 10, 18, 14, 14, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
    tc = ws['A1']; tc.value = title; tc.fill = navy
    tc.font = Font(bold=True, color='FFFFFF', size=13)
    tc.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = white_bold; c.fill = teal; c.border = thin
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.freeze_panes = 'A3'

    r = 3
    tot_monthly = tot_balance = 0.0
    for row in rows:
        vals = [row['emp_id'], row['full_name'], row['current_store'], row['job_title'] or '',
                row['sale_number'] or '', row['description'] or '',
                f"{db.MONTH_NAMES[row['start_month']]} {row['start_year']}",
                row['term_months'], round(row['monthly_amount'], 2), round(row['balance'], 2),
                row['payments_made'], row['status']]
        for col, v in enumerate(map(db.xl_safe, vals), 1):
            c = ws.cell(row=r, column=col, value=v); c.border = thin
            if col in (9, 10):
                c.number_format = cur; c.alignment = Alignment(horizontal='right')
        tot_monthly += row['monthly_amount']; tot_balance += row['balance']
        r += 1
    r += 1
    gt = ['', f'TOTAL ({len(rows)} plans)', '', '', '', '', '', '', round(tot_monthly, 2), round(tot_balance, 2), '', '']
    for col, v in enumerate(gt, 1):
        c = ws.cell(row=r, column=col, value=v); c.fill = navy; c.border = thin
        c.font = Font(bold=True, color='FFFFFF', size=11)
        if col in (9, 10):
            c.number_format = cur; c.alignment = Alignment(horizontal='right')

    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out


@app.route('/hq/laybys/export')
def hq_export_laybys():
    conn = db.get_db()
    store_filter = request.args.get('store', '').strip()
    query = '''
        SELECT l.*, e.full_name, e.current_store, e.job_title, e.id as emp_id
        FROM layby_deductions l
        JOIN employees e ON e.id = l.employee_id
        WHERE l.status = 'active' AND e.sector = 'hq'
          AND e.status = 'active'
          AND l.payments_made < l.term_months
          AND (l.balance_remaining IS NULL OR l.balance_remaining > 0.01)
    '''
    params = []
    if store_filter:
        query += ' AND e.current_store = ?'
        params.append(store_filter)
    query += ' ORDER BY e.current_store, e.full_name, l.id'
    raw = conn.execute(query, params).fetchall()
    conn.close()
    rows = []
    for r in raw:
        balance = r['balance_remaining'] if r['balance_remaining'] is not None else (r['term_months'] - r['payments_made']) * r['monthly_amount']
        rows.append({'emp_id': r['emp_id'], 'full_name': r['full_name'],
                     'current_store': r['current_store'], 'job_title': r['job_title'],
                     'sale_number': r['sale_number'], 'description': r['description'],
                     'start_month': r['start_month'], 'start_year': r['start_year'],
                     'term_months': r['term_months'], 'monthly_amount': r['monthly_amount'],
                     'balance': round(max(0, balance), 2), 'payments_made': r['payments_made'],
                     'status': r['status']})
    label = f' · {store_filter}' if store_filter else ''
    out = _layby_export(rows, f'NORTHWIND HQ Active Lay-by Plans{label}')
    fname = f'HQ_Layby_Plans_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)


@app.route('/hq/monthly/<int:year>/<int:month>/export')
def hq_export_monthly(year, month):
    if db.validate_month_year(year, month) is None:
        flash('Invalid period.', 'warning')
        return redirect(url_for('hq_monthly_current'))
    store_filter = request.args.get('store', '').strip()
    data = db.get_monthly_data(year, month, 'hq')
    rows = []
    for d in data:
        if d['layby_total'] <= 0:
            continue
        if store_filter and d['employee']['current_store'] != store_filter:
            continue
        for p in d['layby_plans']:
            rows.append({'emp_id': d['employee']['id'], 'full_name': d['employee']['full_name'],
                         'current_store': d['employee']['current_store'], 'job_title': d['employee']['job_title'],
                         'sale_number': p.get('sale_number'), 'description': p.get('description'),
                         'start_month': p['start_month'], 'start_year': p['start_year'],
                         'term_months': p['term_months'], 'monthly_amount': p['monthly_amount'],
                         'balance': 0.0, 'payments_made': p['payments_made'], 'status': 'active'})
    out = _layby_export(rows, f'NORTHWIND HQ Lay-by Deductions — {db.MONTH_FULL[month]} {year}')
    label = f'_{store_filter.replace(" ", "_")}' if store_filter else ''
    fname = f'HQ_Deductions_{db.MONTH_NAMES[month]}_{year}{label}.xlsx'
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)
