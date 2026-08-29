from flask import (render_template, request, redirect, url_for, flash,
                   jsonify, send_file)
from datetime import datetime
import io
import sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from northwind.data import database as db
from northwind.core import app


@app.route('/monthly')
def monthly_current():
    now = datetime.now()
    return redirect(url_for('monthly_view', year=now.year, month=now.month))


@app.route('/monthly/<int:year>/<int:month>')
def monthly_view(year, month):
    if db.validate_month_year(year, month) is None:
        flash('Invalid period — month must be 1–12.', 'warning')
        return redirect(url_for('monthly_current'))
    data = db.get_monthly_data(year, month)
    cat = request.args.get('cat', 'all')
    store_filter = request.args.get('store', '')

    cat_key = {'uniform': 'uniform_total', 'layby': 'layby_total', 'undercharges': 'undercharge_total'}
    if cat in cat_key:
        display_data = [d for d in data if d[cat_key[cat]] > 0]
    elif cat == 'invoices':
        display_data = [d for d in data if d['uniform_total'] > 0]
    else:
        display_data = [d for d in data if d['total'] > 0]

    if store_filter:
        display_data = [d for d in display_data if d['employee']['current_store'] == store_filter]

    # The template groups by store with per-store subtotals using row adjacency,
    # so rows must be contiguous per store. get_monthly_data returns employee-id
    # order, so sort by store (then name) — same ordering as the Excel export.
    display_data.sort(key=lambda d: ((d['employee']['current_store'] or '').lower(),
                                     (d['employee']['full_name'] or '').lower()))

    prev_month, prev_year = (month - 1, year) if month > 1 else (12, year - 1)
    next_month, next_year = (month + 1, year) if month < 12 else (1, year + 1)

    all_stores = sorted(set(d['employee']['current_store'] for d in data if d['total'] > 0))

    # Invoice data — always computed for the tab badge, filtered by store if set
    invoice_data = db.get_monthly_invoice_data(year, month)
    if store_filter:
        invoice_data = [i for i in invoice_data if store_filter in i['store']]

    now = datetime.now()
    is_locked = db.is_period_locked(year, month)
    return render_template('monthly.html', data=display_data, year=year, month=month,
                           month_name=db.MONTH_NAMES[month], month_full=db.MONTH_FULL[month],
                           cat=cat, now=now, page_year=year, page_month=month,
                           store_filter=store_filter, all_stores=all_stores,
                           prev_month=prev_month, prev_year=prev_year,
                           next_month=next_month, next_year=next_year,
                           total_uniform=sum(d['uniform_total'] for d in display_data),
                           total_layby=sum(d['layby_total'] for d in display_data),
                           total_undercharges=sum(d['undercharge_total'] for d in display_data),
                           grand_total=sum(d['total'] for d in display_data),
                           invoice_data=invoice_data,
                           invoice_total=sum(i['monthly_total'] for i in invoice_data),
                           is_locked=is_locked)


def _tick_uniform_plans(conn, emp_id, year, month):
    # Single source of truth in database.py so this path and the reconcile
    # 'tick all' can never diverge.
    return db.tick_uniform_due(conn, emp_id, year, month)


def _tick_layby_plans(conn, emp_id, year, month):
    # Single source of truth in database.py so this path and the reconcile
    # 'tick all' can never diverge.
    return db.tick_layby_due(conn, emp_id, year, month)


def _tick_undercharge_rows(conn, emp_id, year, month):
    # Single source of truth for ticking undercharges lives in database.py so
    # this path and the reconcile 'tick all' can never diverge.
    db.tick_undercharges_due(conn, emp_id, year, month)


@app.route('/monthly/<int:year>/<int:month>/pay-uniform/<emp_id>', methods=['POST'])
def pay_uniform_monthly(year, month, emp_id):
    if db.validate_month_year(year, month) is None:
        flash('Invalid period.', 'warning')
        return redirect(url_for('monthly_current'))
    if db.is_period_locked(year, month):
        flash('Cannot process payment. This payroll period is locked.', 'danger')
        return redirect(url_for('monthly_view', year=year, month=month, cat=request.form.get('cat', 'all')))

    conn = db.get_db()
    try:
        with conn:
            _tick_uniform_plans(conn, emp_id, year, month)
        flash('Payment processed successfully.', 'success')
    except sqlite3.IntegrityError:
        flash('A payment has already been allocated for this period.', 'warning')
    except Exception as e:
        flash(f'Error processing payment: {e}', 'danger')
    finally:
        conn.close()
    return redirect(url_for('monthly_view', year=year, month=month, cat=request.form.get('cat', 'all')))


@app.route('/monthly/<int:year>/<int:month>/pay-layby/<emp_id>', methods=['POST'])
def pay_layby_monthly(year, month, emp_id):
    if db.validate_month_year(year, month) is None:
        flash('Invalid period.', 'warning')
        return redirect(url_for('monthly_current'))
    if db.is_period_locked(year, month):
        flash('Cannot process payment. This payroll period is locked.', 'danger')
        return redirect(url_for('monthly_view', year=year, month=month, cat=request.form.get('cat', 'all')))

    conn = db.get_db()
    try:
        with conn:
            _tick_layby_plans(conn, emp_id, year, month)
        flash('Payment processed successfully.', 'success')
    except sqlite3.IntegrityError:
        flash('A payment has already been allocated for this period.', 'warning')
    except Exception as e:
        flash(f'Error processing payment: {e}', 'danger')
    finally:
        conn.close()
    return redirect(url_for('monthly_view', year=year, month=month, cat=request.form.get('cat', 'all')))


@app.route('/monthly/<int:year>/<int:month>/pay-undercharge/<emp_id>', methods=['POST'])
def pay_undercharge_monthly(year, month, emp_id):
    if db.validate_month_year(year, month) is None:
        flash('Invalid period.', 'warning')
        return redirect(url_for('monthly_current'))
    if db.is_period_locked(year, month):
        flash('Cannot process payment. This payroll period is locked.', 'danger')
        return redirect(url_for('monthly_view', year=year, month=month, cat=request.form.get('cat', 'all')))

    conn = db.get_db()
    try:
        with conn:
            _tick_undercharge_rows(conn, emp_id, year, month)
        flash('Payment processed successfully.', 'success')
    except sqlite3.IntegrityError:
        flash('A payment has already been allocated for this period.', 'warning')
    except Exception as e:
        flash(f'Error processing payment: {e}', 'danger')
    finally:
        conn.close()
    return redirect(url_for('monthly_view', year=year, month=month, cat=request.form.get('cat', 'all')))


@app.route('/monthly/<int:year>/<int:month>/pay-all', methods=['POST'])
def pay_all_monthly(year, month):
    """Mark every active deduction for this month as paid (optionally filtered by store)."""
    store_filter = request.form.get('store', '')
    cat = request.form.get('cat', 'all')
    if db.validate_month_year(year, month) is None:
        flash('Invalid period.', 'warning')
        return redirect(url_for('monthly_current'))
    if db.is_period_locked(year, month):
        flash('Cannot process batch payments. This payroll period is locked.', 'danger')
        return redirect(url_for('monthly_view', year=year, month=month, cat=cat, store=store_filter))

    conn = db.get_db()
    try:
        with conn:
            emp_query = "SELECT id FROM employees WHERE status='active' AND sector='retail'"
            emp_params = []
            if store_filter:
                emp_query += " AND current_store=?"
                emp_params.append(store_filter)
            emp_ids = [r['id'] for r in conn.execute(emp_query, emp_params).fetchall()]

            skipped = 0
            for emp_id in emp_ids:
                # Per-employee savepoint: a duplicate-payment IntegrityError for
                # one employee (concurrent double-submit) rolls back only that
                # employee instead of aborting the entire batch.
                conn.execute("SAVEPOINT pay_emp")
                try:
                    if cat in ('all', 'uniform'):
                        _tick_uniform_plans(conn, emp_id, year, month)
                    if cat in ('all', 'layby'):
                        _tick_layby_plans(conn, emp_id, year, month)
                    if cat in ('all', 'undercharges'):
                        _tick_undercharge_rows(conn, emp_id, year, month)
                except sqlite3.IntegrityError:
                    conn.execute("ROLLBACK TO SAVEPOINT pay_emp")
                    skipped += 1
                finally:
                    conn.execute("RELEASE SAVEPOINT pay_emp")
        cat_label = 'Lay-by ' if cat == 'layby' else 'Uniform ' if cat == 'uniform' else 'Undercharge ' if cat == 'undercharges' else ''
        label = f' for {store_filter}' if store_filter else ''
        if skipped:
            flash(f'{cat_label or ""}Deductions{label} marked as paid for {db.MONTH_FULL[month]} {year}; '
                  f'{skipped} employee(s) skipped (already allocated this period).', 'warning')
        else:
            flash(f'All {cat_label}deductions{label} marked as paid for {db.MONTH_FULL[month]} {year}.', 'success')
    except Exception as e:
        flash(f'Error processing batch payments: {e}', 'danger')
    finally:
        conn.close()
    return redirect(url_for('monthly_view', year=year, month=month, cat=cat, store=store_filter))


@app.route('/monthly/<int:year>/<int:month>/pay-selected', methods=['POST'])
def pay_selected_monthly(year, month):
    """Mark deductions paid for a CHOSEN subset of employees (the per-row
    checkboxes + 'Tick selected' button). Same per-employee savepoint loop as
    pay-all, just scoped to the posted emp_ids instead of the whole store/sector."""
    store_filter = request.form.get('store', '')
    cat = request.form.get('cat', 'all')
    emp_ids = [e for e in request.form.getlist('emp_ids') if e]
    if db.validate_month_year(year, month) is None:
        flash('Invalid period.', 'warning')
        return redirect(url_for('monthly_current'))
    if db.is_period_locked(year, month):
        flash('Cannot process payments. This payroll period is locked.', 'danger')
        return redirect(url_for('monthly_view', year=year, month=month, cat=cat, store=store_filter))
    if not emp_ids:
        flash('No employees were selected.', 'warning')
        return redirect(url_for('monthly_view', year=year, month=month, cat=cat, store=store_filter))

    conn = db.get_db()
    try:
        with conn:
            # Scope to real, active retail employees — never trust the posted ids.
            valid = {r['id'] for r in conn.execute(
                "SELECT id FROM employees WHERE status='active' AND sector='retail'").fetchall()}
            targets = [e for e in emp_ids if e in valid]
            done = 0
            skipped = 0
            for emp_id in targets:
                conn.execute("SAVEPOINT pay_emp")
                try:
                    if cat in ('all', 'uniform'):
                        _tick_uniform_plans(conn, emp_id, year, month)
                    if cat in ('all', 'layby'):
                        _tick_layby_plans(conn, emp_id, year, month)
                    if cat in ('all', 'undercharges'):
                        _tick_undercharge_rows(conn, emp_id, year, month)
                    done += 1
                except sqlite3.IntegrityError:
                    conn.execute("ROLLBACK TO SAVEPOINT pay_emp")
                    skipped += 1
                finally:
                    conn.execute("RELEASE SAVEPOINT pay_emp")
        cat_label = ('Lay-by ' if cat == 'layby' else 'Uniform ' if cat == 'uniform'
                     else 'Undercharge ' if cat == 'undercharges' else '')
        msg = f'{cat_label}deductions marked as paid for {done} selected employee(s).'
        if skipped:
            msg += f' {skipped} skipped (already allocated this period).'
        flash(msg, 'warning' if skipped else 'success')
    except Exception as e:
        flash(f'Error processing payments: {e}', 'danger')
    finally:
        conn.close()
    return redirect(url_for('monthly_view', year=year, month=month, cat=cat, store=store_filter))


def _store_display(store):
    """Label for a store-group header / subtotal row, tolerating a NULL or blank
    `current_store` (renders '(No store)' instead of 'None — subtotal')."""
    s = str(store).strip() if store is not None else ''
    return s if s else '(No store)'


def _send_workbook(wb, prefix, month_name, year, store_filter):
    """Serialise `wb` and return it as an .xlsx download, preserving the
    per-category filename convention: `{prefix}_{month}_{year}{store}.xlsx`
    where the store tail is `_<store>` (spaces→_) or `_All_Stores`."""
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    store_label = f"_{store_filter.replace(' ', '_')}" if store_filter else "_All_Stores"
    return send_file(out,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=f'{prefix}_{month_name}_{year}{store_label}.xlsx')


def _category_export_styles(hdr_color, alt_color, total_color):
    """openpyxl styles for a per-category monthly export sheet. The store-group
    and subtotal styling matches the approved 'export all' sheet; the header /
    zebra / total colours keep each category's own identity (navy uniforms,
    teal lay-bys, red undercharges)."""
    return {
        'hdr_fill': PatternFill('solid', start_color=hdr_color),
        'hdr_font': Font(bold=True, color='FFFFFF', size=11),
        'alt_fill': PatternFill('solid', start_color=alt_color),
        'total_fill': PatternFill('solid', start_color=total_color),
        'sub_fill': PatternFill('solid', start_color='F0FDF4'),
        'sub_font': Font(bold=True, size=10),
        'store_fill': PatternFill('solid', start_color='F8FAFC'),
        'store_font': Font(bold=True, size=10, color='475569'),
        'thin': Border(left=Side(style='thin', color='CCCCCC'),
                       right=Side(style='thin', color='CCCCCC'),
                       top=Side(style='thin', color='CCCCCC'),
                       bottom=Side(style='thin', color='CCCCCC')),
        'currency_fmt': 'R#,##0.00;-R#,##0.00;"-"',
    }


def _write_category_sheet(ws, data, headers, col_widths, money_cols, center_cols,
                          row_extractor, styles):
    """Store-grouped layout shared by the three per-category monthly exports.

    `data` is already filtered to the category and sorted by (store, name).
    `row_extractor(d)` yields one value-list per output row for employee `d`
    (an employee can produce several rows — one per plan / undercharge).

    Money columns (1-indexed) are currency-formatted + right-aligned and feed a
    per-store subtotal row and ONE literal grand TOTAL. Every subtotal and the
    grand total is a LITERAL Python sum of the emitted 2dp row values — never
    `=SUM()` over the column, because the per-store subtotal rows live in those
    same columns and a column SUM would double-count them (same reason the
    'export all' sheet accumulates literals). Totals are the sum of the DISPLAYED
    2-decimal row amounts (round-each-row-then-sum), so the visible rows always
    add up to their subtotal and the subtotals to the grand total. For a
    fractional split undercharge (e.g. R10/3) the per-row 2dp value is not exact,
    so this grand total can differ from get_monthly_data's per-employee-rounded
    tab total by up to ~1c per such row — that is intentional and the more
    faithful projection; do NOT switch to sum-then-round."""
    hdr_fill, hdr_font = styles['hdr_fill'], styles['hdr_font']
    alt_fill, total_fill = styles['alt_fill'], styles['total_fill']
    sub_fill, sub_font = styles['sub_fill'], styles['sub_font']
    store_fill, store_font = styles['store_fill'], styles['store_font']
    thin, currency_fmt = styles['thin'], styles['currency_fmt']
    ncols = len(headers)

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.border = thin
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22

    current_store = None
    store_tot = {c: 0.0 for c in money_cols}
    grand_tot = {c: 0.0 for c in money_cols}
    data_row = 2

    def _emit_subtotal(store):
        nonlocal data_row
        for col in range(1, ncols + 1):
            c = ws.cell(row=data_row, column=col)
            c.fill = sub_fill
            c.border = thin
            c.font = sub_font
        ws.cell(row=data_row, column=1, value=f'{_store_display(store)} — subtotal')
        for col in money_cols:
            c = ws.cell(row=data_row, column=col, value=round(store_tot[col], 2))
            c.number_format = currency_fmt
            c.font = sub_font
            c.alignment = Alignment(horizontal='right', vertical='center')
        data_row += 1

    for d in data:
        store = d['employee']['current_store']
        if store != current_store:
            if current_store is not None:
                _emit_subtotal(current_store)
                for col in money_cols:
                    store_tot[col] = 0.0
            for col in range(1, ncols + 1):
                c = ws.cell(row=data_row, column=col)
                c.fill = store_fill
                c.border = thin
            ws.cell(row=data_row, column=1, value=f'  {_store_display(store)}').font = store_font
            data_row += 1
            current_store = store

        for vals in row_extractor(d):
            row_fill = alt_fill if data_row % 2 == 0 else None
            for col, val in enumerate(map(db.xl_safe, vals), 1):
                c = ws.cell(row=data_row, column=col, value=val)
                c.border = thin
                c.alignment = Alignment(vertical='center')
                if row_fill:
                    c.fill = row_fill
                if col in money_cols:
                    c.number_format = currency_fmt
                    c.alignment = Alignment(horizontal='right', vertical='center')
                elif col in center_cols:
                    c.alignment = Alignment(horizontal='center', vertical='center')
            for col in money_cols:
                v = vals[col - 1]
                if isinstance(v, (int, float)):
                    store_tot[col] += v
                    grand_tot[col] += v
            data_row += 1

    if current_store is not None:
        _emit_subtotal(current_store)

    for col in range(1, ncols + 1):
        c = ws.cell(row=data_row, column=col)
        c.fill = total_fill
        c.border = thin
        c.font = Font(bold=True)
    ws.cell(row=data_row, column=1, value='TOTALS')
    for col in money_cols:
        c = ws.cell(row=data_row, column=col, value=round(grand_tot[col], 2))
        c.number_format = currency_fmt
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='right', vertical='center')
    ws.row_dimensions[data_row].height = 22
    ws.freeze_panes = 'A2'


@app.route('/monthly/<int:year>/<int:month>/export')
def export_monthly(year, month):
    if db.validate_month_year(year, month) is None:
        flash('Invalid period.', 'warning')
        return redirect(url_for('monthly_current'))
    cat = request.args.get('cat', 'all').strip()
    store_filter = request.args.get('store', '').strip()
    month_name = db.MONTH_NAMES[month]

    if cat in ('uniform', 'invoices'):
        # Derive straight from get_monthly_data so the sheet's figures are
        # identical to the on-screen Uniform tab and the "export all" sheet
        # (active + retail scoping, completed-plan handling and per-month
        # installments all come for free). One row per active uniform plan.
        data = db.get_monthly_data(year, month, 'retail')
        rows = [d for d in data if d['uniform_total'] != 0]
        if store_filter:
            rows = [d for d in rows if d['employee']['current_store'] == store_filter]
        rows.sort(key=lambda d: ((d['employee']['current_store'] or '').lower(),
                                 (d['employee']['full_name'] or '').lower()))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Staff Uniforms'
        styles = _category_export_styles('1F3864', 'EBF3FB', 'D6E4F0')  # navy
        headers = ['Employee ID', 'Employee Name', 'Store', 'Job Title',
                   'Sale Number', 'SKU', 'Item Description',
                   'Deducted This Month (R)', 'Total Plan Amount (R)', 'Remaining Balance (R)']
        col_widths = [15, 30, 25, 22, 15, 22, 38, 22, 22, 22]

        def uniform_rows(d):
            emp = d['employee']
            for p in d['uniform_plans']:
                if p['monthly_amount'] == 0:
                    continue  # completed / not-scheduled this month
                # Remaining balance: total − payments_made × the agreed regular
                # monthly (preserved by get_monthly_data before monthly_amount
                # was overwritten with this month's installment). Same rule as
                # the previous sheet.
                regular = p.get('regular_monthly') or 0
                total_amt = (p['total_amount'] if p['total_amount'] is not None
                             else round((p['term_months'] or 0) * regular, 2))
                pmt = p['payments_made'] or 0
                if pmt >= (p['term_months'] or 0):
                    rem_val = 0.0
                else:
                    rem_val = total_amt - pmt * round(regular, 2)
                rem_val = round(max(0, rem_val), 2)
                yield [emp['id'], emp['full_name'], emp['current_store'], emp['job_title'] or '',
                       p.get('sale_number') or '', p.get('sku') or '', p.get('description') or '',
                       p['monthly_amount'], total_amt, rem_val]

        _write_category_sheet(ws, rows, headers, col_widths,
                              money_cols=[8, 9, 10], center_cols=[1, 5, 6],
                              row_extractor=uniform_rows, styles=styles)

        # Second worksheet: the per-sale-number summary (moved OFF the main
        # table — the old side-by-side layout was the confusing bit removed).
        ws2 = wb.create_sheet('By Sale Number')
        so_groups = {}
        for d in rows:
            for p in d['uniform_plans']:
                if p['monthly_amount'] == 0:
                    continue
                so = p.get('sale_number') or '(no sale number)'
                g = so_groups.setdefault(so, {'total': 0.0, 'employees': []})
                g['total'] += p['monthly_amount']
                if d['employee']['full_name'] not in g['employees']:
                    g['employees'].append(d['employee']['full_name'])
        s_headers = ['Sale Number', 'Total Installment (R)', 'Employees']
        s_widths = [18, 24, 60]
        for col, (h, w) in enumerate(zip(s_headers, s_widths), 1):
            c = ws2.cell(row=1, column=col, value=h)
            c.font = styles['hdr_font']
            c.fill = styles['hdr_fill']
            c.border = styles['thin']
            c.alignment = Alignment(horizontal='center', vertical='center')
            ws2.column_dimensions[get_column_letter(col)].width = w
        ws2.row_dimensions[1].height = 22
        r2 = 2
        summary_total = 0.0
        for so in sorted(so_groups):
            g = so_groups[so]
            fill = styles['alt_fill'] if r2 % 2 == 0 else None
            cells = [ws2.cell(row=r2, column=1, value=db.xl_safe(so)),
                     ws2.cell(row=r2, column=2, value=round(g['total'], 2)),
                     ws2.cell(row=r2, column=3, value=db.xl_safe(' / '.join(g['employees'])))]
            for c in cells:
                c.border = styles['thin']
                c.alignment = Alignment(vertical='center')
                if fill:
                    c.fill = fill
            cells[0].alignment = Alignment(horizontal='center', vertical='center')
            cells[1].number_format = styles['currency_fmt']
            cells[1].alignment = Alignment(horizontal='right', vertical='center')
            summary_total += g['total']
            r2 += 1
        for col in range(1, 4):
            c = ws2.cell(row=r2, column=col)
            c.fill = styles['total_fill']
            c.border = styles['thin']
            c.font = Font(bold=True)
        ws2.cell(row=r2, column=1, value='TOTALS')
        ct = ws2.cell(row=r2, column=2, value=round(summary_total, 2))
        ct.number_format = styles['currency_fmt']
        ct.font = Font(bold=True)
        ct.alignment = Alignment(horizontal='right', vertical='center')
        ws2.freeze_panes = 'A2'

        return _send_workbook(wb, 'StaffUniforms', month_name, year, store_filter)

    elif cat == 'layby':
        # Derive from get_monthly_data — same figures as the on-screen Lay-by tab
        # and the "export all" sheet (HQ lay-bys can no longer leak in, and the
        # per-month installment is capped at months_left). One row per lay-by plan.
        data = db.get_monthly_data(year, month, 'retail')
        rows = [d for d in data if d['layby_total'] != 0]
        if store_filter:
            rows = [d for d in rows if d['employee']['current_store'] == store_filter]
        rows.sort(key=lambda d: ((d['employee']['current_store'] or '').lower(),
                                 (d['employee']['full_name'] or '').lower()))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Staff Lay-bys'
        styles = _category_export_styles('0F5B52', 'F0FBF9', 'D1FAE5')  # teal
        headers = ['Employee ID', 'Employee Name', 'Store', 'Job Title',
                   'Sale Number', 'Description', 'Deducted This Month (R)']
        col_widths = [12, 28, 20, 22, 14, 38, 20]

        def layby_rows(d):
            emp = d['employee']
            for p in d['layby_plans']:
                if p['monthly_amount'] == 0:
                    continue
                yield [emp['id'], emp['full_name'], emp['current_store'], emp['job_title'] or '',
                       p.get('sale_number') or '', p.get('description') or '', p['monthly_amount']]

        _write_category_sheet(ws, rows, headers, col_widths,
                              money_cols=[7], center_cols=[1, 5],
                              row_extractor=layby_rows, styles=styles)

        return _send_workbook(wb, 'StaffLaybys', month_name, year, store_filter)

    elif cat == 'undercharges':
        # Derive from get_monthly_data — same rows and signed amounts as the
        # on-screen Undercharges tab and the "export all" sheet. Each row carries
        # a pre-computed month_amount (reimbursements negative, recovered/
        # accounted-for 0, a row ticked this month uses its real transaction
        # amount) so the sheet sums exactly to the tab.
        data = db.get_monthly_data(year, month, 'retail')
        rows = [d for d in data if d['undercharge_total'] != 0]
        if store_filter:
            rows = [d for d in rows if d['employee']['current_store'] == store_filter]
        rows.sort(key=lambda d: ((d['employee']['current_store'] or '').lower(),
                                 (d['employee']['full_name'] or '').lower()))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Staff Undercharges'
        styles = _category_export_styles('DC2626', 'FEF2F2', 'FEE2E2')  # red
        headers = ['Employee ID', 'Employee Name', 'Store', 'Job Title',
                   'Sale Number', 'Reason / Type', 'Deducted This Month (R)', 'Status']
        col_widths = [12, 28, 20, 22, 14, 38, 20, 16]

        # Human-readable status labels (unchanged from the previous sheet).
        status_labels = {
            'pending': 'To deduct', 'partial': 'Part-paid',
            'recovered': 'Recovered', 'accounted_for': 'Accounted for',
            'paid_by_customer': 'Reimbursement', 'reimbursed': 'Reimbursed',
        }

        def undercharge_rows(d):
            emp = d['employee']
            for r in d['undercharge_rows']:
                reason_str = r['reason'] or 'Undercharge'
                if r['status'] in ('paid_by_customer', 'reimbursed'):
                    # Preserve the "(Reimbursement)" suffix + negative amount.
                    reason_str = f"{reason_str} (Reimbursement)"
                yield [emp['id'], emp['full_name'], emp['current_store'], emp['job_title'] or '',
                       r.get('sale_number') or '', reason_str, round(r['month_amount'], 2),
                       status_labels.get(r['status'], r['status'])]

        _write_category_sheet(ws, rows, headers, col_widths,
                              money_cols=[7], center_cols=[1, 5, 8],
                              row_extractor=undercharge_rows, styles=styles)

        return _send_workbook(wb, 'StaffUndercharges', month_name, year, store_filter)

    else:
        # Full summary deductions sheet (cat == 'all')
        data = db.get_monthly_data(year, month)
        rows = [d for d in data if d['total'] != 0]
        # The store subtotal logic below assumes rows are contiguous per store,
        # but get_monthly_data returns them in employee-id order. Sort by store
        # (then name) so each store forms a single block — otherwise a store that
        # reappears later in id order spawns duplicate headers/subtotals and its
        # staff get split across several partial groups.
        rows.sort(key=lambda d: ((d['employee']['current_store'] or '').lower(),
                                 (d['employee']['full_name'] or '').lower()))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f'{month_name} {year}'

        hdr_fill = PatternFill('solid', start_color='1F3864')
        hdr_font = Font(bold=True, color='FFFFFF', size=11)
        alt_fill = PatternFill('solid', start_color='EBF3FB')
        total_fill = PatternFill('solid', start_color='D6E4F0')
        currency_fmt = 'R#,##0.00;-R#,##0.00;"-"'
        thin = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )

        headers = ['Employee ID', 'Employee', 'Store', 'Job Title',
                   'Staff Uniform (R)', 'Lay-by (R)', 'Undercharges (R)',
                   'Ref #', 'Reason', 'Total Deductions (R)']
        col_widths = [12, 35, 25, 22, 18, 14, 18, 18, 38, 20]

        for col, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[1].height = 22

        current_store = None
        store_uniform = store_layby = store_uc = 0
        # Grand totals accumulated from EMPLOYEE rows only. The TOTALS row must
        # not =SUM() the column, because the per-store subtotal rows also live in
        # those columns and would be counted a second time (doubling the total).
        grand_uniform = grand_layby = grand_uc = 0.0
        data_row = 2

        sub_fill  = PatternFill('solid', start_color='F0FDF4')
        sub_font  = Font(bold=True, size=10)
        store_fill = PatternFill('solid', start_color='F8FAFC')
        store_font = Font(bold=True, size=10, color='475569')

        for d in rows:
            emp = d['employee']
            store = emp['current_store']

            # Filter by store if filter is active
            if store_filter and store != store_filter:
                continue

            if store != current_store:
                if current_store is not None:
                    # subtotal row for previous store
                    ws.cell(row=data_row, column=1, value=f'{current_store} — subtotal').font = sub_font
                    for col in range(1, 11):
                        ws.cell(row=data_row, column=col).fill = sub_fill
                        ws.cell(row=data_row, column=col).border = thin
                        ws.cell(row=data_row, column=col).font = sub_font
                    ws.cell(row=data_row, column=5, value=store_uniform).number_format = currency_fmt
                    ws.cell(row=data_row, column=5).alignment = Alignment(horizontal='right')
                    ws.cell(row=data_row, column=6, value=store_layby).number_format = currency_fmt
                    ws.cell(row=data_row, column=6).alignment = Alignment(horizontal='right')
                    ws.cell(row=data_row, column=7, value=store_uc).number_format = currency_fmt
                    ws.cell(row=data_row, column=7).alignment = Alignment(horizontal='right')
                    ws.cell(row=data_row, column=10, value=store_uniform + store_layby + store_uc).number_format = currency_fmt
                    ws.cell(row=data_row, column=10).alignment = Alignment(horizontal='right')
                    data_row += 1
                    store_uniform = store_layby = store_uc = 0

                # store header row
                ws.cell(row=data_row, column=1, value=f'  {store}').font = store_font
                for col in range(1, 11):
                    ws.cell(row=data_row, column=col).fill = store_fill
                    ws.cell(row=data_row, column=col).border = thin
                data_row += 1
                current_store = store

            row_fill = alt_fill if data_row % 2 == 0 else None

            reimbursement_amt = 0.0
            for r in d['undercharge_rows']:
                if r.get('month_amount', 0) < 0:
                    reimbursement_amt += -r['month_amount']

            emp_name = emp['full_name']
            if reimbursement_amt > 0:
                if reimbursement_amt == int(reimbursement_amt):
                    emp_name = f"{emp_name} (Pay {int(reimbursement_amt)})"
                else:
                    emp_name = f"{emp_name} (Pay {reimbursement_amt:.2f})"

            # Collect references and reasons from all active plans for this employee for this month
            u_refs = [p['sale_number'] for p in d['uniform_plans'] if p.get('sale_number')]
            u_descs = [p['description'] for p in d['uniform_plans'] if p.get('description')]

            l_refs = [p['sale_number'] for p in d['layby_plans'] if p.get('sale_number')]
            l_descs = [p['description'] for p in d['layby_plans'] if p.get('description')]

            uc_refs = [p['sale_number'] for p in d['undercharge_rows'] if p.get('sale_number')]
            uc_descs = [
                f"{p['reason']} (Reimbursement)" if p['status'] == 'paid_by_customer' else p['reason']
                for p in d['undercharge_rows'] if p.get('reason')
            ]

            all_refs = u_refs + l_refs + uc_refs
            combined_ref = ' / '.join(filter(None, dict.fromkeys(all_refs)))

            all_reasons = u_descs + l_descs + uc_descs
            combined_reason = ' / '.join(filter(None, dict.fromkeys(all_reasons)))

            vals = [emp['id'], emp_name, store, emp['job_title'],
                    d['uniform_total'] or None, d['layby_total'] or None,
                    d['undercharge_total'] or None,
                    combined_ref or None, combined_reason or None,
                    d['total'] or None]
            for col, val in enumerate(map(db.xl_safe, vals), 1):
                cell = ws.cell(row=data_row, column=col, value=val)
                cell.border = thin
                if row_fill:
                    cell.fill = row_fill
                if col in (5, 6, 7, 10):
                    cell.number_format = currency_fmt
                    cell.alignment = Alignment(horizontal='right')

            store_uniform += d['uniform_total'] or 0
            store_layby   += d['layby_total'] or 0
            store_uc      += d['undercharge_total'] or 0
            grand_uniform += d['uniform_total'] or 0
            grand_layby   += d['layby_total'] or 0
            grand_uc      += d['undercharge_total'] or 0
            data_row += 1

        # final store subtotal
        if current_store and (not store_filter or current_store == store_filter):
            ws.cell(row=data_row, column=1, value=f'{current_store} — subtotal').font = sub_font
            for col in range(1, 11):
                ws.cell(row=data_row, column=col).fill = sub_fill
                ws.cell(row=data_row, column=col).border = thin
                ws.cell(row=data_row, column=col).font = sub_font
            ws.cell(row=data_row, column=5, value=store_uniform).number_format = currency_fmt
            ws.cell(row=data_row, column=5).alignment = Alignment(horizontal='right')
            ws.cell(row=data_row, column=6, value=store_layby).number_format = currency_fmt
            ws.cell(row=data_row, column=6).alignment = Alignment(horizontal='right')
            ws.cell(row=data_row, column=7, value=store_uc).number_format = currency_fmt
            ws.cell(row=data_row, column=7).alignment = Alignment(horizontal='right')
            ws.cell(row=data_row, column=10, value=store_uniform + store_layby + store_uc).number_format = currency_fmt
            ws.cell(row=data_row, column=10).alignment = Alignment(horizontal='right')
            data_row += 1

        tr = data_row
        ws.cell(row=tr, column=1, value='TOTALS').font = Font(bold=True)
        for col in range(1, 11):
            cell = ws.cell(row=tr, column=col)
            cell.fill = total_fill
            cell.border = thin
            cell.font = Font(bold=True)
        for col, val in [(5, grand_uniform), (6, grand_layby), (7, grand_uc),
                         (10, grand_uniform + grand_layby + grand_uc)]:
            cell = ws.cell(row=tr, column=col, value=round(val, 2))
            cell.number_format = currency_fmt
            cell.alignment = Alignment(horizontal='right')
            cell.font = Font(bold=True)
            cell.fill = total_fill

        ws.freeze_panes = 'A2'

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        store_label = f"_{store_filter.replace(' ', '_')}" if store_filter else ""
        return send_file(out,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True,
                         download_name=f'Deductions_{month_name}_{year}{store_label}.xlsx')


@app.route('/monthly/<int:year>/<int:month>/lock', methods=['POST'])
def lock_period(year, month):
    if db.validate_month_year(year, month) is None:
        flash('Invalid period.', 'warning')
        return redirect(url_for('monthly_current'))
    conn = db.get_db()
    try:
        conn.execute("INSERT OR IGNORE INTO locked_periods (sector, year, month) VALUES ('retail', ?, ?)", (year, month))
        conn.commit()
        flash(f'Payroll period {db.MONTH_FULL[month]} {year} has been locked successfully.', 'success')
    except Exception as e:
        flash(f'Error locking period: {e}', 'danger')
    finally:
        conn.close()
    return redirect(url_for('monthly_view', year=year, month=month, cat=request.form.get('cat', 'all'), store=request.form.get('store', '')))


@app.route('/monthly/<int:year>/<int:month>/unlock', methods=['POST'])
def unlock_period(year, month):
    if db.validate_month_year(year, month) is None:
        flash('Invalid period.', 'warning')
        return redirect(url_for('monthly_current'))
    conn = db.get_db()
    try:
        conn.execute("DELETE FROM locked_periods WHERE sector='retail' AND year = ? AND month = ?", (year, month))
        conn.commit()
        flash(f'Payroll period {db.MONTH_FULL[month]} {year} has been unlocked.', 'warning')
    except Exception as e:
        flash(f'Error unlocking period: {e}', 'danger')
    finally:
        conn.close()
    return redirect(url_for('monthly_view', year=year, month=month, cat=request.form.get('cat', 'all'), store=request.form.get('store', '')))


@app.route('/transactions/<int:tx_id>/void', methods=['POST'])
def void_transaction(tx_id):
    conn = db.get_db()
    tx = conn.execute("SELECT * FROM deduction_transactions WHERE id=?", (tx_id,)).fetchone()
    if not tx:
        conn.close()
        if request.headers.get('Accept') == 'application/json':
            return jsonify({'success': False, 'message': 'Transaction not found.'})
        flash('Transaction not found.', 'danger')
        return redirect(request.referrer or url_for('employees'))

    if tx['voided'] == 1:
        conn.close()
        if request.headers.get('Accept') == 'application/json':
            return jsonify({'success': False, 'message': 'This transaction is already voided.'})
        flash('Transaction is already voided.', 'warning')
        return redirect(request.referrer or url_for('employee_detail', emp_id=tx['employee_id']))

    if db.is_period_locked(tx['year'], tx['month']):
        conn.close()
        if request.headers.get('Accept') == 'application/json':
            return jsonify({'success': False, 'message': 'The payroll period for this transaction is locked.'})
        flash('Cannot void transaction. The associated payroll period is locked.', 'danger')
        return redirect(request.referrer or url_for('employee_detail', emp_id=tx['employee_id']))

    try:
        plan_type = tx['plan_type']
        plan_id = tx['plan_id']
        amount = tx['amount']

        # Wrap all writes in a transaction so the plan update and the
        # transaction void are atomic — no half-applied state on crash.
        with conn:
            if plan_type == 'uniform':
                plan = conn.execute("SELECT * FROM uniform_deductions WHERE id=?", (plan_id,)).fetchone()
                if plan:
                    new_count = max(0, plan['payments_made'] - 1)
                    # Restore the balance too (mirrors the layby branch) — the
                    # uniform tick writes balance_remaining_cents, so voiding must
                    # add the amount back or the outstanding stays understated.
                    # Coalesce a NULL balance to 0 (legacy plans may carry one)
                    # so the void can never fail with a TypeError.
                    new_bal = (plan['balance_remaining'] or 0) + amount
                    conn.execute(
                        "UPDATE uniform_deductions_cents SET payments_made=?, balance_remaining_cents=?, status='active' WHERE id=?",
                        (new_count, db.to_cents(new_bal), plan_id)
                    )
            elif plan_type == 'layby':
                plan = conn.execute("SELECT * FROM layby_deductions WHERE id=?", (plan_id,)).fetchone()
                if plan:
                    new_count = max(0, plan['payments_made'] - 1)
                    new_bal = (plan['balance_remaining'] or 0) + amount
                    conn.execute(
                        "UPDATE layby_deductions_cents SET payments_made=?, balance_remaining_cents=?, status='active' WHERE id=?",
                        (new_count, db.to_cents(new_bal), plan_id)
                    )
            elif plan_type == 'undercharge':
                plan = conn.execute("SELECT * FROM undercharges WHERE id=?", (plan_id,)).fetchone()
                if plan:
                    item = conn.execute(
                        "SELECT id FROM undercharge_schedule_items WHERE transaction_id=?",
                        (tx_id,)).fetchone()
                    if item:
                        conn.execute(
                            "UPDATE undercharge_schedule_items SET transaction_id=NULL "
                            "WHERE id=?", (item['id'],))

            conn.execute("UPDATE deduction_transactions_cents SET voided=1 WHERE id=?", (tx_id,))
            if plan_type == 'undercharge':
                db.sync_undercharge_state(conn, plan_id)
        
        flash('Transaction voided successfully and balance restored.', 'success')
        success = True
    except Exception as e:
        conn.rollback()
        flash(f'Error voiding transaction: {e}', 'danger')
        success = False
    finally:
        conn.close()

    if request.headers.get('Accept') == 'application/json':
        outstanding = db.get_outstanding_summary(tx['employee_id'])
        cat_totals = db.get_category_totals(tx['employee_id'])
        return jsonify({
            'success': success,
            'outstanding': outstanding,
            'cat_totals': cat_totals
        })
    return redirect(request.referrer or url_for('employee_detail', emp_id=tx['employee_id']))
